from __future__ import annotations
# pyright: ignore[reportMissingImports]
from info_produto_crud import InfoProdutoCRUDMixin
from arranjo_crud import ArranjoCRUDMixin
from deposito_crud import DepositoCRUDMixin
from categoria_crud import CategoriaCRUDMixin
from estoque_crud import EstoqueCRUDMixin
from estrutura_crud import EstruturaCRUDMixin
from fornecedor_crud import FornecedorCRUDMixin
from produtos_crud import ProdutosCRUDMixin
from situacao_crud import SituacaoCRUDMixin
from reportlab.lib.pagesizes import mm
from reportlab.pdfgen import canvas

import os
import sys
import json
import threading
import warnings
import traceback
import subprocess
from dataclasses import dataclass, asdict
from datetime import datetime, date
from collections import defaultdict
from math import ceil
from typing import Optional, Dict, Any, List, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

import requests
import psycopg2
from psycopg2 import errors
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ============================================================
# PATHS
# ============================================================


def get_app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = get_app_dir()

# Ajuste conforme sua realidade
BASE_DIR = r"Z:\Planilhas_OP"

if not os.path.exists(BASE_DIR):
    try:
        os.makedirs(BASE_DIR)
    except Exception:
        pass

CAMINHO_MODELO = os.path.join(BASE_DIR, "pedido-de-compra v2.xlsx")
CAMINHO_SAIDA = os.path.join(BASE_DIR, "saida_pedido-de-compra v2.xlsx")

# ou "" para desabilitar
N8N_WEBHOOK_URL = "http://localhost:56789/webhook/ordem-producao"

warnings.filterwarnings(
    "ignore", message="Cannot parse header or footer so it will be ignored")
warnings.filterwarnings(
    "ignore", message="Data Validation extension is not supported and will be removed")


# ============================================================
# LOG
# ============================================================

def log_exception(err: Exception, context: str = "") -> str:
    try:
        texto = "".join(traceback.format_exception(
            type(err), err, err.__traceback__))
        log_path = os.path.join(BASE_DIR, "erro_app.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n" + ("=" * 80) + "\n")
            if context:
                f.write(f"{context}\n")
            f.write(texto)
        return log_path
    except Exception:
        return os.path.join(BASE_DIR, "erro_app.log")


# ============================================================
# ÍCONE
# ============================================================

def find_icon_path() -> Optional[str]:
    candidates = [
        os.path.join(BASE_DIR, "imagens", "favicon.ico"),
        os.path.join(BASE_DIR, "favicon.ico"),
        os.path.join(APP_DIR, "imagens", "favicon.ico"),
        os.path.join(APP_DIR, "favicon.ico"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return None


def apply_window_icon(win) -> None:
    try:
        ico = find_icon_path()  # seu .ico
        if ico:
            try:
                win.iconbitmap(default=ico)
            except Exception:
                pass

        # Para a taskbar, o que ajuda MESMO é iconphoto com PNG
        png_candidates = [
            os.path.join(BASE_DIR, "imagens", "favicon.png"),
            os.path.join(APP_DIR, "imagens", "favicon.png"),
        ]
        png = next((p for p in png_candidates if os.path.isfile(p)), None)
        if png:
            img = tk.PhotoImage(file=png)
            win.iconphoto(True, img)
            win._icon_img = img  # mantém referência (senão some)
    except Exception:
        pass

# ============================================================
# CONFIG (JSON)
# ============================================================


@dataclass
class AppConfig:
    # banco
    db_host: str = "10.0.0.154"
    db_database: str = "postgresekenox"
    db_user: str = "postgresekenox"
    db_password: str = "Ekenox5426"
    db_port: int = 55432

    # F7
    f7_geometry: str = "1100x560"
    # (no F7: saldo sempre do banco; fica aqui por compatibilidade)
    usar_saldo_do_banco: bool = True

    # pedido compra
    caminho_modelo: str = CAMINHO_MODELO
    caminho_saida: str = CAMINHO_SAIDA

    # bling
    bling_base_url: str = "https://api.bling.com.br/Api/v3"
    bling_token: str = ""  # cole aqui o access_token (Bearer)
    bling_timeout: int = 20


def config_path() -> str:
    return os.path.join(BASE_DIR, "config_op.json")


def load_config() -> AppConfig:
    p = config_path()
    if not os.path.exists(p):
        return AppConfig()
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        return AppConfig(**data)
    except Exception:
        return AppConfig()


def save_config(cfg: AppConfig) -> None:
    p = config_path()
    with open(p, "w", encoding="utf-8") as f:
        json.dump(asdict(cfg), f, ensure_ascii=False, indent=2)


# ============================================================
# SPLASH
# ============================================================

class Splash(tk.Toplevel):
    def __init__(self, master: tk.Tk, titulo="Inicializando..."):
        super().__init__(master)
        self.title(titulo)
        self.geometry("430x170")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        apply_window_icon(self)

        self.lbl = ttk.Label(self, text="Iniciando...",
                             font=("Segoe UI", 11, "bold"))
        self.lbl.pack(pady=(18, 8))

        self.pb = ttk.Progressbar(self, mode="indeterminate")
        self.pb.pack(fill="x", padx=18, pady=8)

        self.info = ttk.Label(self, text="", foreground="gray")
        self.info.pack(pady=(8, 0))

        self.pb.start(10)
        self._center()

    def set_text(self, t: str):
        self.lbl.config(text=t)
        self.update_idletasks()

    def set_info(self, t: str):
        self.info.config(text=t)
        self.update_idletasks()

    def _center(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (self.winfo_width() // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")


# ============================================================
# EXCEL (Pedido de compra)
# ============================================================

def _nome_aba_excel_valido(nome: str) -> str:
    invalidos = ['\\', '/', '?', '*', '[', ']']
    for ch in invalidos:
        nome = nome.replace(ch, " ")
    return nome[:31]


def _escrever_celula_segura(ws, coord: str, valor):
    try:
        cell = ws[coord]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if coord in merged_range:
                    top_left_coord = merged_range.coord.split(":")[0]
                    ws[top_left_coord] = valor
                    return
        else:
            cell.value = valor
    except Exception:
        pass


def _set_cell_segura_rc(ws, row: int, col: int, valor):
    cell = ws.cell(row=row, column=col)
    coord = cell.coordinate
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                ws.cell(row=merged_range.min_row,
                        column=merged_range.min_col).value = valor
                return
    else:
        cell.value = valor


def gerar_abas_fornecedor_pedido(dados: List[Dict[str, Any]], nome_aba_modelo: str = "Pedido de Compra",
                                 caminho_modelo: str = CAMINHO_MODELO, caminho_saida: str = CAMINHO_SAIDA):
    if os.path.exists(caminho_saida):
        wb = load_workbook(caminho_saida)
    else:
        wb = load_workbook(caminho_modelo)

    if nome_aba_modelo not in wb.sheetnames:
        raise ValueError(
            f"Aba de modelo '{nome_aba_modelo}' não encontrada. Abas: {wb.sheetnames}")

    aba_modelo = wb[nome_aba_modelo]

    tmp = defaultdict(list)
    for item in dados:
        fornecedor = item["fornecedor"]
        numero_pedido = item["numero_pedido"]
        data_pedido = item.get("data_pedido")
        tmp[(fornecedor, numero_pedido, data_pedido)].append(item)

    for (fornecedor, numero_pedido, data_pedido), linhas in tmp.items():
        ws = wb.copy_worksheet(aba_modelo)
        titulo_aba = f"{numero_pedido} - {str(fornecedor)[:15]}"
        ws.title = _nome_aba_excel_valido(titulo_aba)

        _escrever_celula_segura(ws, "D6", numero_pedido)
        _escrever_celula_segura(ws, "D8", data_pedido or date.today())
        _escrever_celula_segura(ws, "D10", str(fornecedor))

        # limpa área da tabela (ajuste se seu modelo tiver outra área)
        for r in range(16, 43):
            for c in range(2, 10):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        linha = 16
        numero_item = 1

        for item in linhas:
            descricao = item["produto"]
            quantidade = float(item["quantidade"])

            estoque_atual = float(item.get("estoque_atual", 0.0) or 0.0)
            estoque_minimo = float(item.get("estoque_minimo", 0.0) or 0.0)
            estoque_maximo = float(item.get("estoque_maximo", 0.0) or 0.0)
            valor_unitario = item.get("valor_unitario")

            _set_cell_segura_rc(ws, linha, 2, numero_item)        # B
            _set_cell_segura_rc(ws, linha, 3, descricao)          # C
            _set_cell_segura_rc(ws, linha, 4, estoque_atual)      # D
            _set_cell_segura_rc(ws, linha, 5, estoque_minimo)     # E
            _set_cell_segura_rc(ws, linha, 6, estoque_maximo)     # F

            if valor_unitario is not None:
                vu = float(valor_unitario)
                _set_cell_segura_rc(ws, linha, 7, vu)             # G
                _set_cell_segura_rc(ws, linha, 9, quantidade * vu)  # I

            _set_cell_segura_rc(ws, linha, 8, quantidade)         # H

            numero_item += 1
            linha += 1

    wb.save(caminho_saida)


# ============================================================
# BANCO / REGRAS (inclui F7 lendo Ekenox.estrutura/infoProduto/fornecedor)
# ============================================================

class SistemaOrdemProducao(InfoProdutoCRUDMixin, ArranjoCRUDMixin, CategoriaCRUDMixin, DepositoCRUDMixin,
                           EstoqueCRUDMixin, EstruturaCRUDMixin, FornecedorCRUDMixin, ProdutosCRUDMixin,
                           SituacaoCRUDMixin):

    def __init__(self, cfg: AppConfig):
        self.cfg = cfg
        self.conn = None
        self.cursor = None
        self.ultimo_erro: Optional[str] = None
        self._ultimo_erro_bd: Optional[str] = None

    def conectar(self) -> bool:
        self.ultimo_erro = None
        try:
            self.conn = psycopg2.connect(
                host=self.cfg.db_host,
                database=self.cfg.db_database,
                user=self.cfg.db_user,
                password=self.cfg.db_password,
                port=int(self.cfg.db_port),
                connect_timeout=5,
            )
            self.cursor = self.conn.cursor()
            return True
        except Exception as e:
            self.ultimo_erro = f"{type(e).__name__}: {e}"
            return False

    def desconectar(self):
        try:
            if self.cursor:
                self.cursor.close()
            if self.conn:
                self.conn.close()
        except Exception:
            pass

    def _q(self, sql: str, params: Tuple = ()):
        if not self.cursor:
            raise RuntimeError("Sem cursor (não conectado).")
        self.cursor.execute(sql, params)

    from decimal import Decimal

    def validar_estoque_insumos_para_producao(
        self,
        fkproduto: int,
        qtd_produzir: float,
        bloquear_se_saldo_negativo: bool = True,
        bloquear_se_insuficiente: bool = True,
    ) -> Dict[str, Any]:
        """
        Valida estoque dos insumos (estrutura) para produzir 'qtd_produzir' do produto final.
        Regras:
          - se saldo < 0: bloqueia (se bloquear_se_saldo_negativo=True)
          - se saldo < necessario: bloqueia (se bloquear_se_insuficiente=True)

        Retorno:
          {
            "ok": bool,
            "problemas": [
                {
                  "componente": int,
                  "nome": str,
                  "qtd_base": float,
                  "necessario": float,
                  "saldo": float,
                  "falta": float
                }, ...
            ]
          }
        """
        problemas: List[Dict[str, Any]] = []

        try:
            # estrutura do produto final
            # [(componente, qtd_base), ...]
            itens = self.f7_buscar_estrutura(int(fkproduto))

            # Se não tiver estrutura, aqui você decide:
            # - pode permitir salvar (ok=True) ou bloquear. Vou deixar permitir.
            if not itens:
                return {"ok": True, "problemas": []}

            for (componente, qtd_base) in itens:
                comp_id = int(componente)
                qtd_base_f = float(qtd_base or 0.0)
                necessario = qtd_base_f * float(qtd_produzir)

                saldo = float(self.saldo_fisico(comp_id) or 0.0)

                # nome do componente (pra mostrar no erro)
                prod_comp = self.validar_produto(comp_id) or {}
                nome = (prod_comp.get("nomeproduto") or "").strip()

                falta = max(0.0, necessario - saldo)

                # regra 1: saldo negativo
                if bloquear_se_saldo_negativo and saldo < 0:
                    problemas.append({
                        "componente": comp_id,
                        "nome": nome,
                        "qtd_base": qtd_base_f,
                        "necessario": necessario,
                        "saldo": saldo,
                        "falta": falta,
                        "motivo": "Saldo negativo",
                    })
                    continue

                # regra 2: insuficiente
                if bloquear_se_insuficiente and saldo < necessario:
                    problemas.append({
                        "componente": comp_id,
                        "nome": nome,
                        "qtd_base": qtd_base_f,
                        "necessario": necessario,
                        "saldo": saldo,
                        "falta": falta,
                        "motivo": "Saldo insuficiente",
                    })

            ok = (len(problemas) == 0)
            return {"ok": ok, "problemas": problemas}

        except Exception:
            if self.conn:
                self.conn.rollback()
            # em caso de erro técnico, é mais seguro bloquear
            return {"ok": False, "problemas": [{"motivo": "Erro ao validar estrutura/estoque"}]}

    def relatorio_bling_insumos_produto(self, produto_id: int, qtd_produzir: float):
        """
        Retorna 6 colunas:
        descricao, codigo(produtoId), valor_unit, unidade, qtde_un, qtde_total
        """
        sql = """
            SELECT
                COALESCE(pcomp."nomeProduto", '')                         AS descricao,
                pcomp."produtoId"::bigint                                 AS codigo,

                -- ✅ valor unitário do insumo (prioriza precoCompra se existir)
                CASE
                WHEN ipcomp."precoCompra" IS NOT NULL AND ipcomp."precoCompra" > 0
                    THEN ipcomp."precoCompra"::numeric
                WHEN pcomp."custo" IS NOT NULL AND pcomp."custo" > 0
                    THEN pcomp."custo"::numeric
                ELSE 0::numeric
                END                                                      AS valor_unit,

                COALESCE(ipcomp."unidadeMedida", ipcomp."unidade", '')    AS unidade,
                COALESCE(e."quantidade", 0)::numeric                      AS qtde_un,
                (COALESCE(e."quantidade", 0) * %s)::numeric               AS qtde_total
            FROM "Ekenox"."estrutura" e
            LEFT JOIN "Ekenox"."produtos" pcomp
                ON pcomp."produtoId"::bigint = e."componente"::bigint
            LEFT JOIN "Ekenox"."infoProduto" ipcomp
                ON ipcomp."fkProduto"::bigint = e."componente"::bigint
            WHERE e."fkproduto"::bigint = %s
            ORDER BY descricao;
        """
        self._q(sql, (float(qtd_produzir), int(produto_id)))
        return self.cursor.fetchall() or []

    def buscar_ordem_producao_por_numero(self, numero: str | int):
        try:
            num_int = int(str(numero).strip())

            sql = """
                SELECT
                    o."id", o."numero", o."deposito_destino", o."deposito_origem",
                    o."situacao_id", o."responsavel", o."fkprodutoid",
                    o."data_previsao_inicio", o."data_previsao_final",
                    o."data_inicio", o."data_fim", o."valor", o."observacao", o."quantidade"
                FROM "Ekenox"."ordem_producao" o
                WHERE o."numero" = %s
                LIMIT 1;
            """
            self._q(sql, (num_int,))
            r = self.cursor.fetchone()
            if not r:
                return None

            return {
                "id": r[0],
                "numero": r[1],
                "deposito_destino": r[2],
                "deposito_origem": r[3],
                "situacao_id": r[4],
                "responsavel": r[5],
                "fkprodutoid": r[6],
                "data_previsao_inicio": r[7],
                "data_previsao_final": r[8],
                "data_inicio": r[9],
                "data_fim": r[10],
                "valor": r[11],
                "observacao": r[12],
                "quantidade": r[13],
            }

        except Exception as e:
            if self.conn:
                self.conn.rollback()
            return None

    def buscar_qtd_produzir_por_sku(self, sku: str) -> float:
        """
        Soma a quantidade no arranjo para um SKU (tenta com N e sem N).
        Retorna 0.0 se não encontrar.
        """
        try:
            sku_raw = (sku or "")
            sku_norm = sku_raw.strip().upper()

            if not sku_norm:
                return 0.0

            # Gera candidatos: SKU como está, +N, e sem N
            candidatos = {sku_norm}
            if sku_norm.endswith("N"):
                candidatos.add(sku_norm[:-1])
            else:
                candidatos.add(sku_norm + "N")

            candidatos = sorted(candidatos)  # só pra debug ficar estável

            sql = """
                SELECT COALESCE(SUM(a."quantidade"), 0)
                FROM "Ekenox"."arranjo" a
                WHERE UPPER(TRIM(a."sku")) = ANY(%s);
            """
            self._q(sql, (candidatos,))
            row = self.cursor.fetchone()

            return float(row[0]) if row and row[0] is not None else 0.0

        except Exception as e:
            print("Erro buscar_qtd_produzir_por_sku:", e)
            if self.conn:
                self.conn.rollback()
            return 0.0

    def validar_produto(self, produto_id: int) -> Optional[Dict[str, Any]]:
        try:
            sql = """
                SELECT p."produtoId", p."nomeProduto", p."sku", p."preco", p."tipo"
                  FROM "Ekenox"."produtos" p
                 WHERE p."produtoId" = %s;
            """
            self._q(sql, (str(produto_id),))
            r = self.cursor.fetchone()
            if not r:
                return None
            return {"produtoid": r[0], "nomeproduto": r[1], "sku": r[2], "preco": r[3], "tipo": r[4]}
        except Exception:
            return None

    def media_vendas_mensal(self, fk_produto: int, ano: int | None = None, mes: int | None = None) -> float:
        try:
            if ano is None or mes is None:
                hoje = date.today()
                ano = hoje.year
                mes = hoje.month

            mes_inicio = date(int(ano), int(mes), 1)

            sql = """
                SELECT v.media_vendas
                FROM vw_media_vendas_mensal v
                WHERE v.fkProduto = %s
                  AND date_trunc('month', v.dataVenda)::date = %s
                LIMIT 1
            """
            self._q(sql, (int(fk_produto), mes_inicio))
            row = self.cursor.fetchone()
            return float(row[0]) if row and row[0] is not None else 0.0

        except Exception:
            if self.conn:
                self.conn.rollback()
            return 0.0

    def validar_situacao(self, situacao_id: int) -> Optional[Dict[str, Any]]:
        try:
            sql = """SELECT s."id", s."nome" FROM "Ekenox"."situacao" s WHERE s."id" = %s;"""
            self._q(sql, (situacao_id,))
            r = self.cursor.fetchone()
            if not r:
                return None
            return {"id": r[0], "nome": r[1]}
        except Exception:
            return None

    def gerar_id_ordem(self) -> int:
        try:
            sql = """SELECT COALESCE(MAX(id),0) + 1 FROM "Ekenox"."ordem_producao";"""
            self._q(sql)
            r = self.cursor.fetchone()
            return int(r[0]) if r and r[0] is not None else 1
        except Exception:
            return 1

    def gerar_numero_ordem(self) -> int:
        try:
            sql = """SELECT COALESCE(MAX(numero),0) + 1 FROM "Ekenox"."ordem_producao";"""
            self._q(sql)
            r = self.cursor.fetchone()
            return int(r[0]) if r and r[0] is not None else 1
        except Exception:
            return 1

    # ✅ ESTE MÉTODO PRECISA EXISTIR NA CLASSE
    def listar_depositos_disponiveis(self, limite: Optional[int] = None):
        try:
            if limite is None or limite <= 0:
                self._q("""
                    SELECT d."id", d."descricao", d."situacao", d."padrao", d."desconsiderarsaldo"
                      FROM "Ekenox"."deposito" d
                     ORDER BY d."descricao";
                """)
            else:
                self._q("""
                    SELECT d."id", d."descricao", d."situacao", d."padrao", d."desconsiderarsaldo"
                      FROM "Ekenox"."deposito" d
                     ORDER BY d."descricao"
                     LIMIT %s;
                """, (limite,))
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def listar_situacoes_disponiveis(self, limite: Optional[int] = None):
        try:
            if limite is None or limite <= 0:
                self._q(
                    """SELECT s."id", s."nome" FROM "Ekenox"."situacao" s ORDER BY s."nome";""")
            else:
                self._q(
                    """SELECT s."id", s."nome" FROM "Ekenox"."situacao" s ORDER BY s."nome" LIMIT %s;""", (limite,))
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def listar_produtos_disponiveis(self, limite: Optional[int] = None):
        try:
            base_sql = """
                SELECT p."produtoId", p."nomeProduto", p."sku", p."preco", p."tipo"
                  FROM "Ekenox"."produtos" AS p
                  JOIN "Ekenox"."infoProduto" AS ip
                    ON ip."fkProduto" = p."produtoId"
                 WHERE ip."fkCategoria" NOT IN (
                    3844533,3983855,7879429,3869959,4241123,
                    3870601,3844542,7651801,3983399,959867,
                    897565,3984869,3862825,7879102,7911660,
                    4828356,6568231
                 )
                 ORDER BY p."nomeProduto"
            """
            if limite is None or limite <= 0:
                self._q(base_sql)
            else:
                self._q(base_sql + " LIMIT %s", (limite,))
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    # ✅ inserção retornando (bool, msg)
    def inserir_ordem_producao(self, dados: Dict[str, Any]) -> tuple[bool, str]:
        self._ultimo_erro_bd = None

        try:
            produto = self.validar_produto(dados["fkprodutoid"])
            if not produto:
                raise ValueError(
                    f"Produto ID {dados['fkprodutoid']} não encontrado.")

            situacao = self.validar_situacao(dados["situacao_id"])
            if not situacao:
                raise ValueError(
                    f"Situação ID {dados['situacao_id']} não encontrada.")

            if not dados.get("id"):
                dados["id"] = self.gerar_id_ordem()

            query = """
                INSERT INTO "Ekenox"."ordem_producao" (
                    id, numero, deposito_destino, deposito_origem, situacao_id,
                    responsavel, fkprodutoid, data_previsao_inicio, data_previsao_final,
                    data_inicio, data_fim, valor, observacao, quantidade
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """

            valores = (
                dados["id"],
                dados["numero"],
                dados["deposito_id_destino"],
                dados["deposito_id_origem"],
                dados["situacao_id"],
                dados.get("responsavel"),
                dados["fkprodutoid"],
                dados.get("data_previsao_inicio"),
                dados.get("data_previsao_final"),
                dados.get("data_inicio"),
                dados.get("data_fim"),
                dados.get("valor"),
                dados.get("observacao"),
                dados["quantidade"],
            )

            self.cursor.execute(query, valores)
            self.conn.commit()
            return True, ""

        except errors.UniqueViolation as e:
            self.conn.rollback()
            msg = (
                "NÚMERO DE ORDEM JÁ EXISTENTE.\n\n"
                f"Número: {dados.get('numero')}\n"
                f"Constraint: {getattr(e.diag, 'constraint_name', '')}"
            )
            self._ultimo_erro_bd = msg
            return False, msg

        except errors.ForeignKeyViolation as e:
            self.conn.rollback()
            tabela = getattr(e.diag, "table_name", "desconhecida")
            constraint = getattr(e.diag, "constraint_name", "desconhecida")
            msg = (
                "VIOLAÇÃO DE CHAVE ESTRANGEIRA.\n\n"
                f"Tabela alvo: {tabela}\n"
                f"Constraint: {constraint}\n\n"
                "Provavelmente algum ID não existe (depósitos, situação ou produto)."
            )
            self._ultimo_erro_bd = msg
            return False, msg

        except Exception as e:
            if self.conn:
                self.conn.rollback()
            msg = f"Erro ao inserir OP: {e}"
            self._ultimo_erro_bd = msg
            return False, msg
    # ---------- Listagens básicas ----------

    def listar_produtos_disponiveis(self, limite: Optional[int] = None):
        try:
            query = """
                    SELECT p."produtoId", p."nomeProduto", p."sku", p."preco", p."tipo"
                    FROM "Ekenox"."produtos" AS p
                    JOIN "Ekenox"."infoProduto" AS ip
                        ON ip."fkProduto" = p."produtoId"
                    WHERE ip."fkCategoria" NOT IN (
                        3844533,3983855,7879429,3869959,4241123,
                        3870601,3844542,7651801,3983399,959867,
                        897565,3984869,3862825,7879102,7911660,
                        4828356,6568231
                    )
                ORDER BY p."nomeProduto"
                """
            if limite is not None and limite > 0:
                query += " LIMIT %s"
                self._q(query, (limite,))
            else:
                self._q(query)
            return self.cursor.fetchall()
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def listar_situacoes_disponiveis(self, limite: Optional[int] = None):
        try:
            sql = """SELECT s."id", s."nome" FROM "Ekenox"."situacao" s ORDER BY s."nome" """
            if limite is not None and limite > 0:
                sql += " LIMIT %s"
                self._q(sql, (limite,))
            else:
                self._q(sql)
            return self.cursor.fetchall()
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def listar_depositos_disponiveis(self, limite: Optional[int] = None):
        try:
            sql = """
                SELECT d."id", d."descricao", d."situacao", d."padrao", d."desconsiderarsaldo"
                FROM "Ekenox"."deposito" d
            ORDER BY d."descricao"
            """
            if limite is not None and limite > 0:
                sql += " LIMIT %s"
                self._q(sql, (limite,))
            else:
                self._q(sql)
            return self.cursor.fetchall()
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    # ---------- Estoque (saldo) ----------
    def saldo_fisico(self, produto_id: int) -> float:
        """Saldo físico do produto (precisa existir Ekenox.estoque com fkProduto e saldoFisico)."""
        try:
            sql = """
                SELECT COALESCE(SUM(e."saldoFisico"), 0)
                  FROM "Ekenox"."estoque" e
                 WHERE e."fkProduto"::bigint = %s;
            """
            self._q(sql, (str(produto_id),))
            r = self.cursor.fetchone()
            return float(r[0]) if r and r[0] is not None else 0.0
        except Exception:
            return 0.0

    # ============================================================
    # F7 - Leitura do BANCO
    # ============================================================

    def f7_buscar_estrutura(self, fkproduto: int):
        try:
            sql = """
                SELECT e."componente", e."quantidade"
                FROM "Ekenox"."estrutura" e
                WHERE e."fkproduto"::bigint = %s
            ORDER BY e."componente";
            """
            self._q(sql, (str(fkproduto),))  # str() pra evitar text=bigint
            return self.cursor.fetchall() or []
        except Exception as e:
            # rollback extra de segurança (caso alguém tenha usado execute direto)
            if self.conn:
                self.conn.rollback()
            raise

    def f6_buscar_estrutura(self, fkproduto: int):
        try:
            sql = """
                select ip."estoqueMaximo"
                from "Ekenox"."infoProduto" as ip
                WHERE ip."fkProduto"::bigint = %s
                limit 1;
            """
            self._q(sql, (str(fkproduto),))  # str() pra evitar text=bigint
            return self.cursor.fetchall() or []
        except Exception as e:
            # rollback extra de segurança (caso alguém tenha usado execute direto)
            if self.conn:
                self.conn.rollback()
            raise

    def f7_buscar_info_produto(self, fkproduto: int) -> Dict[str, Any]:
        """
        Lê Ekenox.infoProduto (colunas relevantes mostradas por você):
          estoqueMinimo, estoqueMaximo, precoCompra, fkFornecedor, fkProduto
        """
        sql = """
            SELECT
                i."estoqueMinimo",
                i."estoqueMaximo",
                i."precoCompra",
                i."fkFornecedor",
                i."fkProduto"
              FROM "Ekenox"."infoProduto" i
             WHERE i."fkProduto"::bigint = %s;
        """
        self._q(sql, (int(fkproduto),))
        r = self.cursor.fetchone()
        if not r:
            return {}
        return {
            "estoqueMinimo": float(r[0] or 0),
            "estoqueMaximo": float(r[1] or 0),
            "precoCompra": float(r[2] or 0),
            "fkFornecedor": int(r[3] or 0) if r[3] is not None else 0,
            "fkProduto": int(r[4] or fkproduto),
        }

    def f7_buscar_fornecedor(self, id_fornecedor: int) -> Dict[str, Any]:
        """
        Lê Ekenox.fornecedor:
          nome, codigo, telefone, celular, idFornecedor...
        """
        if not id_fornecedor:
            return {}
        sql = """
            SELECT f."idFornecedor", f."nome", f."codigo", f."telefone", f."celular"
              FROM "Ekenox"."fornecedor" f
             WHERE f."idFornecedor" = %s;
        """
        self._q(sql, (int(id_fornecedor),))
        r = self.cursor.fetchone()
        if not r:
            return {}
        return {"idFornecedor": r[0], "nome": r[1], "codigo": r[2], "telefone": r[3], "celular": r[4]}

    # ============================
    # F10 - Listar ordens
    # ============================
    def listar_ordens_producao(self):
        try:
            sql = """
                SELECT
                    o."id",
                    o."numero",
                    o."fkprodutoid",
                    p."nomeProduto" AS produto_nome,
                    o."situacao_id",
                    s."nome" AS situacao_nome,
                    o."quantidade",
                    o."data_inicio",
                    o."data_fim"
                FROM "Ekenox"."ordem_producao" o
                LEFT JOIN "Ekenox"."produtos" p
                       ON p."produtoId" = o."fkprodutoid"
                LEFT JOIN "Ekenox"."situacao" s
                       ON s."id" = o."situacao_id"
                ORDER BY o."id" DESC;
            """
            self._q(sql)
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def excluir_ordem_producao(self, ordem_id: int) -> bool:
        try:
            sql = 'DELETE FROM "Ekenox"."ordem_producao" WHERE "id" = %s;'
            self._q(sql, (int(ordem_id),))
            ok = (self.cursor.rowcount or 0) > 0
            self.conn.commit()
            return ok
        except Exception:
            if self.conn:
                self.conn.rollback()
            return False

    # ============================
    # F11 - Pendentes / finalizar
    # ============================
    def listar_ordens_sem_data_fim(self):
        try:
            sql = """
                SELECT
                    o."id",
                    o."numero",
                    o."fkprodutoid",
                    p."nomeProduto" AS produto_nome,
                    o."situacao_id",
                    s."nome" AS situacao_nome,
                    o."quantidade",
                    o."data_inicio"
                FROM "Ekenox"."ordem_producao" o
                LEFT JOIN "Ekenox"."produtos" p
                       ON p."produtoId" = o."fkprodutoid"
                LEFT JOIN "Ekenox"."situacao" s
                       ON s."id" = o."situacao_id"
                WHERE o."data_fim" IS NULL OR o."data_fim" = '1970-01-01'
                ORDER BY o."id" DESC;
            """
            self._q(sql)
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def finalizar_ordem_individual(self, ordem_id: int) -> bool:
        """
        Marca data_fim = hoje e situacao_id = 18162 (ajuste se precisar)
        """
        try:
            hoje = date.today()
            sql = """
                UPDATE "Ekenox"."ordem_producao"
                   SET "data_fim" = %s,
                       situacao_id = 18162
                 WHERE "id" = %s
                   AND ("data_fim" IS NULL OR "data_fim" = '1970-01-01');
            """
            self._q(sql, (hoje, int(ordem_id)))
            if (self.cursor.rowcount or 0) == 0:
                self.conn.rollback()
                return False
            self.conn.commit()
            return True
        except Exception:
            if self.conn:
                self.conn.rollback()
            return False

# ============================================================
# MÓDULO ETIQUETAS (abre com F12 no app principal)
# - NÃO cria novo Tk()
# - Usa Toplevel (janela filha)
# ============================================================


class EtiquetasModule:
    def __init__(self, master: tk.Tk, cfg: AppConfig):
        self.master = master
        self.cfg = cfg
        self.win: Optional[tk.Toplevel] = None

    def open(self, event=None):
        # Se já existe, só traz pra frente
        if self.win is not None and self.win.winfo_exists():
            try:
                self.win.deiconify()
                self.win.lift()
                self.win.focus_force()
            except Exception:
                pass
            return

        self.win = tk.Toplevel(self.master)
        apply_window_icon(self.win)
        self.win.title("Gerador de Etiquetas EKENOX")
        self.win.geometry("680x720")
        self.win.transient(self.master)
        self.win.grab_set()  # modal (se não quiser modal, remova)

        self.win.bind("<Escape>", lambda e: self.close())

        self._montar_interface(self.win)

        # F12 dentro da janela de etiquetas = abrir seleção de produtos
        self.win.bind("<F12>", lambda e: self.listar_produtos())

    def close(self):
        try:
            if self.win and self.win.winfo_exists():
                self.win.destroy()
        except Exception:
            pass
        self.win = None

    # ---------- Banco (usa cfg do app principal) ----------
    def _conn(self):
        return psycopg2.connect(
            host=self.cfg.db_host,
            database=self.cfg.db_database,
            user=self.cfg.db_user,
            password=self.cfg.db_password,
            port=int(self.cfg.db_port),
            connect_timeout=5,
        )

    @staticmethod
    def _ultimo_caractere(texto: str):
        return texto[-1] if texto else None

    # ---------- Seleção de produtos ----------
    def listar_produtos(self):
        if self.win is None or not self.win.winfo_exists():
            return

        try:
            conn = self._conn()
            cur = conn.cursor()

            sql = '''
                SELECT 
                    p."produtoId"                        AS Produto,
                    p."nomeProduto"                      AS Nome,
                    p."sku"                              AS SKU,
                    COALESCE(NULLIF(TRIM(p."descImetro"), ''), p."nomeProduto") AS Imetro,
                    (
                        SELECT RIGHT(ped."numero"::text, 4) AS Numero
                        FROM "Ekenox"."itens"   AS i
                        JOIN "Ekenox"."pedidos" AS ped
                          ON ped."idPedido" = i."fkPedido"
                        WHERE i."fkProduto" = p."produtoId"
                        ORDER BY ped."data" DESC
                        LIMIT 1
                    ) AS numero_pedido
                FROM "Ekenox"."produtos" AS p
                LEFT JOIN "Ekenox"."infoProduto" AS ip
                  ON p."produtoId" = ip."fkProduto"
                WHERE (p."descImetro"  IS NOT NULL AND TRIM(p."descImetro")  <> '')
                ORDER BY p."nomeProduto" ASC;
            '''
            cur.execute(sql)
            produtos = cur.fetchall()
            cur.close()
            conn.close()

            if not produtos:
                messagebox.showinfo(
                    "Produtos",
                    "Nenhum produto encontrado com os filtros configurados.",
                    parent=self.win,
                )
                return

        except Exception as e:
            messagebox.showerror(
                "Erro ao buscar produtos",
                f"Ocorreu um erro ao consultar o banco de dados:\n{e}",
                parent=self.win,
            )
            return

        # janela seleção
        janela = tk.Toplevel(self.win)
        apply_window_icon(janela)
        janela.title("Selecionar Produto")
        janela.geometry("900x400")
        janela.transient(self.win)
        janela.grab_set()

        frame = tk.Frame(janela, padx=10, pady=10)
        frame.pack(fill="both", expand=True)

        scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        cols = ("ID", "Nome", "SKU", "DescInmetro", "Pedido")
        tree = ttk.Treeview(
            frame, columns=cols, show="headings", yscrollcommand=scrollbar.set
        )

        tree.heading("ID", text="ID")
        tree.heading("Nome", text="Nome do Produto")
        tree.heading("SKU", text="SKU")
        tree.heading("DescInmetro", text="Desc. Inmetro")
        tree.heading("Pedido", text="Nº Pedido (último)")

        tree.column("ID", width=60, anchor="center")
        tree.column("Nome", width=300, anchor="w")
        tree.column("SKU", width=120, anchor="w")
        tree.column("DescInmetro", width=260, anchor="w")
        tree.column("Pedido", width=110, anchor="center")

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=tree.yview)

        for prod_id, nome, sku, desc_inm, num_ped in produtos:
            tree.insert("", tk.END, values=(prod_id, nome or "",
                        sku or "", desc_inm or "", num_ped or ""))

        def selecionar_produto(event=None):
            sel = tree.selection()
            if not sel:
                return
            valores = tree.item(sel[0])["values"]

            sku_val = (valores[2] or "").strip()
            desc_inmetro = (valores[3] or "").strip()
            numero_pedido = valores[4]

            # remove N/n final do SKU
            if sku_val and self._ultimo_caractere(sku_val).upper() == "N":
                sku_val = sku_val[:-1]

            self.entry_produto.delete(0, tk.END)
            self.entry_produto.insert(0, desc_inmetro)

            self.entry_modelo.delete(0, tk.END)
            self.entry_modelo.insert(0, sku_val)

            if numero_pedido not in (None, ""):
                self.entry_serie.delete(0, tk.END)
                self.entry_serie.insert(0, str(numero_pedido).strip())

            janela.destroy()

        tree.bind("<Double-Button-1>", selecionar_produto)
        tree.bind("<Return>", selecionar_produto)

        janela.bind("<Escape>", lambda e: janela.destroy())
        janela.update_idletasks()
        x = (janela.winfo_screenwidth() // 2) - (janela.winfo_width() // 2)
        y = (janela.winfo_screenheight() // 2) - (janela.winfo_height() // 2)
        janela.geometry(f"+{x}+{y}")

    # ---------- Gerar PDF ----------
    def gerar_etiquetas(self):
        try:
            empresa = {
                "company_name": self.entry_empresa.get().strip(),
                "company_address": self.entry_endereco.get().strip(),
                "company_district": self.entry_bairro.get().strip(),
                "company_city": self.entry_cidade.get().strip(),
                "company_state": self.entry_estado.get().strip(),
                "company_cep": self.entry_cep.get().strip(),
                "company_phone": self.entry_telefone.get().strip(),
                "company_email": self.entry_email.get().strip(),
            }

            produto = {
                "product_title": self.entry_produto.get().strip(),
                "product_model": self.entry_modelo.get().strip(),
                "product_classe": self.entry_classe.get().strip(),
                "voltage": self.combo_tensao.get().strip(),
                "power": self.combo_potencia.get().strip(),
                "temperature": self.entry_temperatura.get().strip(),
                "frequency": self.entry_frequencia.get().strip(),
            }

            if not produto["product_title"]:
                messagebox.showerror(
                    "Erro", "O campo 'Produto' deve ser preenchido (selecione na lista).", parent=self.win)
                return

            quantidade_str = self.entry_quantidade.get().strip()
            try:
                quantidade = int(quantidade_str)
                if quantidade <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror(
                    "Erro", "A quantidade deve ser um inteiro > 0.", parent=self.win)
                return

            serie_base = self.entry_serie.get().strip()
            if not serie_base:
                messagebox.showerror(
                    "Erro", "Preencha o 'Número de Série (prefixo/base)'.", parent=self.win)
                return

            # salva em BASE_DIR (evita cair em pasta desconhecida)
            pdf_path = os.path.join(BASE_DIR, "etiquetas.pdf")

            largura, altura = 100 * mm, 75 * mm
            c = canvas.Canvas(pdf_path, pagesize=(largura, altura))

            x_titulo = 10
            x_valor = 70
            espaco = 10

            for i in range(quantidade):
                serial = f"{serie_base}-{i+1:03d}"

                c.setLineWidth(1)
                c.rect(5, 5, largura - 10, altura - 10)

                c.setFont("Helvetica-Bold", 12)
                c.drawCentredString(largura / 2, altura -
                                    15, empresa["company_name"])

                y = altura - 30

                campos_empresa = [
                    ("Endereço:", empresa["company_address"]),
                    ("Bairro:", empresa["company_district"]),
                    ("Cidade:",
                     f"{empresa['company_city']} - {empresa['company_state']}"),
                    ("CEP:", empresa["company_cep"]),
                    ("Telefone:", empresa["company_phone"]),
                    ("Email SAC:", empresa["company_email"]),
                ]

                for titulo, valor in campos_empresa:
                    c.setFont("Helvetica-Bold", 9)
                    c.drawString(x_titulo, y, titulo)
                    c.setFont("Helvetica", 9)
                    c.drawString(x_valor, y, valor)
                    y -= espaco

                c.line(x_titulo, y, largura - 10, y)
                y -= espaco

                produto_campos = [
                    ("Produto:", produto["product_title"]),
                    ("Modelo:", produto["product_model"]),
                    ("Classe:", produto["product_classe"]),
                    ("Tensão:", produto["voltage"]),
                    ("Potência:", produto["power"]),
                    ("Temp:", produto["temperature"]),
                    ("Freq:", produto["frequency"]),
                ]

                for titulo, valor in produto_campos:
                    c.setFont("Helvetica-Bold", 9)
                    c.drawString(x_titulo, y, titulo)
                    c.setFont("Helvetica", 9)
                    c.drawString(x_valor, y, valor)
                    y -= espaco

                c.line(x_titulo, y, largura - 10, y)
                y -= espaco * 2

                c.setFont("Helvetica-Bold", 12)
                c.drawCentredString(largura / 2, y, f"Nº Série: {serial}")

                c.showPage()

            c.save()
            messagebox.showinfo(
                "Sucesso", f"PDF gerado:\n{pdf_path}", parent=self.win)

            # opcional: abrir automaticamente no Windows
            try:
                if os.name == "nt":
                    os.startfile(pdf_path)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Erro", str(e), parent=self.win)

    # ---------- UI ----------
    def _montar_interface(self, root):
        # Empresa
        frame_empresa = tk.LabelFrame(
            root, text="Informações da Empresa", padx=10, pady=10)
        frame_empresa.pack(fill="both", padx=10, pady=5)

        def add_row(r, label, default=""):
            tk.Label(frame_empresa, text=label).grid(
                row=r, column=0, sticky="e")
            e = tk.Entry(frame_empresa, width=50)
            if default:
                e.insert(0, default)
            e.grid(row=r, column=1, pady=2)
            return e

        self.entry_empresa = add_row(
            0, "Nome da Empresa:", "EKENOX DISTRIBUIDORA DE COZ. IND. LTDA")
        self.entry_endereco = add_row(
            1, "Endereço:", "Rua: José de Ribamar Souza, 499")
        self.entry_bairro = add_row(2, "Bairro:", "Pq. Industrial")
        self.entry_cidade = add_row(3, "Cidade:", "Catanduva")
        self.entry_estado = add_row(4, "Estado:", "SP")
        self.entry_cep = add_row(5, "CEP:", "15803-290")
        self.entry_telefone = add_row(6, "Telefone:", "(11)98740-3669")
        self.entry_email = add_row(7, "Email SAC:", "sac@ekenox.com.br")

        # Produto
        frame_prod = tk.LabelFrame(
            root, text="Informações do Produto", padx=10, pady=10)
        frame_prod.pack(fill="both", padx=10, pady=5)

        tk.Label(frame_prod, text="Produto:").grid(row=0, column=0, sticky="e")
        self.entry_produto = tk.Entry(frame_prod, width=45)
        self.entry_produto.insert(0, "BUFFET TÉRMICO")
        self.entry_produto.grid(row=0, column=1, pady=2, sticky="w")

        tk.Button(frame_prod, text="Selecionar...", command=self.listar_produtos).grid(
            row=0, column=2, padx=5, pady=2)

        tk.Label(frame_prod, text="Classe:").grid(row=1, column=0, sticky="e")
        self.entry_classe = tk.Entry(frame_prod, width=50)
        self.entry_classe.insert(0, "IPX4")
        self.entry_classe.grid(
            row=1, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Modelo (SKU):").grid(
            row=2, column=0, sticky="e")
        self.entry_modelo = tk.Entry(frame_prod, width=50)
        self.entry_modelo.insert(0, "VIX8368")
        self.entry_modelo.grid(
            row=2, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Tensão:").grid(row=3, column=0, sticky="e")
        self.combo_tensao = ttk.Combobox(
            frame_prod, values=["127V", "220V"], state="readonly", width=47)
        self.combo_tensao.grid(
            row=3, column=1, columnspan=2, pady=2, sticky="w")
        self.combo_tensao.set("127V")

        tk.Label(frame_prod, text="Potência:").grid(
            row=4, column=0, sticky="e")
        self.combo_potencia = ttk.Combobox(
            frame_prod, values=["1000W", "2000W"], state="readonly", width=47)
        self.combo_potencia.grid(
            row=4, column=1, columnspan=2, pady=2, sticky="w")
        self.combo_potencia.set("2000W")

        tk.Label(frame_prod, text="Temperatura:").grid(
            row=5, column=0, sticky="e")
        self.entry_temperatura = tk.Entry(frame_prod, width=50)
        self.entry_temperatura.insert(0, "30°C a 120°C")
        self.entry_temperatura.grid(
            row=5, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Frequência:").grid(
            row=6, column=0, sticky="e")
        self.entry_frequencia = tk.Entry(frame_prod, width=50)
        self.entry_frequencia.insert(0, "60Hz")
        self.entry_frequencia.grid(
            row=6, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Número de Série (prefixo/base):").grid(row=7,
                                                                          column=0, sticky="e")
        self.entry_serie = tk.Entry(frame_prod, width=50)
        self.entry_serie.insert(0, "EKX2024")
        self.entry_serie.grid(
            row=7, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Quantidade de etiquetas:").grid(
            row=8, column=0, sticky="e")
        self.entry_quantidade = tk.Entry(frame_prod, width=50)
        self.entry_quantidade.insert(0, "5")
        self.entry_quantidade.grid(
            row=8, column=1, columnspan=2, pady=2, sticky="w")

        frame_btn = tk.Frame(root, pady=10)
        frame_btn.pack(fill="x")

        tk.Button(frame_btn, text="Gerar PDF", command=self.gerar_etiquetas,
                  bg="#2563eb", fg="white", font=("Arial", 12, "bold"), width=15).pack(side="left", padx=(40, 10))

        tk.Button(frame_btn, text="Fechar", command=self.close,
                  bg="#ef4444", fg="white", font=("Arial", 12, "bold"), width=15).pack(side="left")


# ============================================================
# UI
# ============================================================

class OrdemProducaoApp(tk.Tk):

    def buscar_ordem_por_numero(self, event=None):
        if not self.connected:
            messagebox.showerror("Buscar OP", "Sem conexão com o banco.")
            return

        num_txt = (self.numero_var.get() or "").strip()
        if not num_txt:
            messagebox.showwarning("Buscar OP", "Informe o Número da Ordem.")
            return

        op = self.sistema.buscar_ordem_producao_por_numero(num_txt)
        if not op:
            messagebox.showinfo(
                "Buscar OP", f"OP nº {num_txt} não encontrada.")
            return

        # Preenche campos
        self.numero_var.set(str(op.get("numero") or ""))

        self.deposito_destino_var.set(str(op.get("deposito_destino") or ""))
        self.deposito_origem_var.set(str(op.get("deposito_origem") or ""))
        self.situacao_id_var.set(str(op.get("situacao_id") or ""))
        self.produto_id_var.set(str(op.get("fkprodutoid") or ""))

        self.responsavel_var.set(str(op.get("responsavel") or ""))

        qtd = op.get("quantidade")
        if qtd is None:
            self.quantidade_var.set("")
        else:
            try:
                self.quantidade_var.set(f"{float(qtd):.2f}")
            except Exception:
                self.quantidade_var.set(str(qtd))

        val = op.get("valor")
        if val is None:
            self.valor_var.set("")
        else:
            try:
                self.valor_var.set(f"{float(val):.2f}")
            except Exception:
                self.valor_var.set(str(val))

        obs = (op.get("observacao") or "").strip()
        self.observacao_text.delete("1.0", tk.END)
        if obs:
            self.observacao_text.insert("1.0", obs)

        # Importante:
        # NÃO chamar atualizar_quantidade_producao() aqui,
        # porque ela recalcula e sobrescreve a quantidade da OP buscada.

        messagebox.showinfo(
            "Buscar OP", f"OP nº {op.get('numero')} carregada com sucesso.")

    def arredondar_para_multiplo_arranjo(self, qtd: float, arranjo: float) -> float:
        """
       Arredonda 'qtd' para CIMA no próximo múltiplo de 'arranjo'.
        Ex:
        qtd=6, arranjo=5 -> 10
        qtd=10, arranjo=5 -> 10
        qtd=0, arranjo=5 -> 0
        """
        try:
            qtd = float(qtd or 0.0)
            arranjo = float(arranjo or 0.0)
            if qtd <= 0:
                return 0.0
            if arranjo <= 0:
                return qtd
            return ceil(qtd / arranjo) * arranjo
        except Exception:
            return float(qtd or 0.0)

    def exportar_relatorio_bling_excel_f9(self, event=None):
        if not self.connected:
            messagebox.showerror("F9 - Relatório", "Sem conexão com o banco.")
            return

        prod_txt = (self.produto_id_var.get() or "").strip()
        qtd_txt = (self.quantidade_var.get() or "").strip()

        if not prod_txt:
            messagebox.showerror("F9 - Relatório", "Informe o Produto (ID).")
            return

        try:
            produto_id = int(prod_txt)
        except ValueError:
            messagebox.showerror("F9 - Relatório", "Produto (ID) inválido.")
            return

        try:
            qtd_produzir = float(qtd_txt.replace(",", ".")) if qtd_txt else 0.0
        except Exception:
            qtd_produzir = 0.0

        if qtd_produzir <= 0:
            messagebox.showerror(
                "F9 - Relatório", "Quantidade para produzir deve ser > 0.")
            return

        produto = self.sistema.validar_produto(produto_id) or {}
        if not produto:
            messagebox.showerror("F9 - Relatório", "Produto não encontrado.")
            return

        prod_nome = (produto.get("nomeproduto") or "").strip()
        prod_codigo = int(produto_id)

        try:
            insumos = self.sistema.relatorio_bling_insumos_produto(
                produto_id, qtd_produzir)
        except Exception as e:
            messagebox.showerror(
                "F9 - Relatório", f"Erro ao consultar estrutura:\n{e}")
            return

        if not insumos:
            messagebox.showinfo(
                "F9 - Relatório", "Sem estrutura cadastrada para este produto.")
            return

        # ---------- helpers ----------
        def D(x) -> Decimal:
            if x is None:
                return Decimal("0")
            if isinstance(x, Decimal):
                return x
            try:
                return Decimal(str(x).replace(",", "."))
            except Exception:
                return Decimal("0")

        def money2(x: Decimal) -> Decimal:
            return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        qtd_prod_D = D(qtd_produzir)

        # ---------- cálculo conferido ----------
        total_geral = Decimal("0")
        for (desc, cod, valor_unit, un, qt_un, qt_total) in insumos:
            v = D(valor_unit)
            qt_tot = D(qt_total)
            total_item = money2(v * qt_tot)
            total_geral += total_item

        unitario = money2(
            total_geral / qtd_prod_D) if qtd_prod_D > 0 else Decimal("0")

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        def auto_col_width(ws, min_w=10, max_w=60):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    v = cell.value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[col_letter].width = max(
                    min_w, min(max_w, max_len + 2))

        # Estilos
        blue = PatternFill("solid", fgColor="1F4E79")
        gray = PatternFill("solid", fgColor="F2F2F2")
        header_font = Font(bold=True, color="FFFFFF")
        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        qty_fmt = '#,##0.00'
        money_fmt = '"R$" #,##0.00'

        # ---------- título ----------
        ws.merge_cells("A1:G1")
        ws["A1"] = "Relatório de Produção"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.append([""] * 7)

        # ---------- Item para produção ----------
        ws.append(["Item para produção"] + [""] * 6)
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=7)
        ws.cell(row=ws.max_row, column=1).font = bold

        ws.append(["Descrição", "Código",
                  "Quantidade a produzir", "", "", "", ""])
        r = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

        ws.append([prod_nome, prod_codigo, float(
            qtd_produzir), "", "", "", ""])
        r = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c == 3:
                cell.number_format = qty_fmt
                cell.alignment = Alignment(
                    horizontal="right", vertical="center")

        ws.append([""] * 7)

        # ---------- Insumos ----------
        ws.append(["Insumos (matéria prima)"] + [""] * 6)
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=7)
        ws.cell(row=ws.max_row, column=1).font = bold

        headers = [
            "Descrição", "Código", "Valor do insumo", "Qtde un.",
            "Valor (Qtde un * compra)", "Qtde total", "Total insumo"
        ]
        ws.append(headers)

        header_row = ws.max_row
        for c in range(1, 8):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = blue
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

        start_data = ws.max_row + 1

        for (desc, cod, valor_unit, un, qt_un, qt_total) in insumos:
            v = D(valor_unit)
            q_un = D(qt_un)
            q_tot = D(qt_total)

            valor_qt_un = money2(v * q_un)   # ✅ por 1 unidade do produto final
            total_item = money2(v * q_tot)  # ✅ total do lote

            ws.append([
                desc or "",
                int(cod or 0),
                float(v),
                float(q_un),
                float(valor_qt_un),
                float(q_tot),
                float(total_item),
            ])

            r = ws.max_row
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.border = border

                # C, E, G = dinheiro | D, F = quantidade
                if c in (3, 5, 7):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center")
                elif c in (4, 6):
                    cell.number_format = qty_fmt
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True)

        end_data = ws.max_row
        ws.auto_filter.ref = f"A{header_row}:G{end_data}"
        ws.freeze_panes = f"A{start_data}"

        # Linha TOTAL (geral)
        ws.append(["", "", "", "", "", "TOTAL:", float(total_geral)])
        r = ws.max_row
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=6).font = bold
        ws.cell(row=r, column=7).font = bold
        ws.cell(row=r, column=7).number_format = money_fmt
        ws.cell(row=r, column=7).alignment = Alignment(
            horizontal="right", vertical="center")

        ws.append([""] * 7)

        # ---------- Observação ----------
        ws.append(["Observação"] + [""] * 6)
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=7)
        ws.cell(row=ws.max_row, column=1).font = bold

        # Unitário em E (coluna 5) e Total Geral em G (coluna 7)
        ws.append(["", "", "Total Unitário:", "", float(
            unitario), "Total Geral:", float(total_geral)])
        obs_row = ws.max_row

        # ✅ E (unitário) em moeda
        ws.cell(row=obs_row, column=5).number_format = money_fmt
        # ✅ G (total) em moeda
        ws.cell(row=obs_row, column=7).number_format = money_fmt
        ws.cell(row=obs_row, column=5).alignment = Alignment(
            horizontal="right", vertical="center")
        ws.cell(row=obs_row, column=7).alignment = Alignment(
            horizontal="right", vertical="center")

        for c in range(1, 8):
            cell = ws.cell(row=obs_row, column=c)
            cell.border = border
            cell.fill = gray
            cell.font = bold
            if c in (5, 7):
                cell.alignment = Alignment(
                    horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center")

        # Larguras
        auto_col_width(ws)
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 14

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_bling_{produto_id}_{ts}.xlsx"
        caminho = os.path.join(BASE_DIR, nome_arquivo)

        wb.save(caminho)
        messagebox.showinfo(
            "F9 - Relatório", f"Relatório gerado em:\n{caminho}")

        try:
            if os.name == "nt":
                os.startfile(caminho)
        except Exception:
            pass

    def _on_f12(self, event=None):
        # evita abrir enquanto está fechando o app
        if getattr(self, "_closing", False):
            return

        # abre a janela de etiquetas
        try:
            self.mod_etiquetas.open()
        except Exception as e:
            messagebox.showerror(
                "Etiquetas (F12)", f"Erro ao abrir Etiquetas:\n{e}")

    def __init__(self):

        from sistema_loader import SistemaOrdemProducao

        self.sistema = SistemaOrdemProducao(self.cfg)
        ok, err = self.sistema.conectar()
        self.connected = ok

        if not ok:
            messagebox.showerror("Banco", f"Falha ao conectar:\n{err}")
            return

        super().__init__()

        self._closing = False
        self.cfg = load_config()
        self.mod_etiquetas = EtiquetasModule(self, self.cfg)

        self.title("Sistema de Ordem de Produção - Ekenox")
        self.geometry("1150x650")
        self.minsize(1150, 650)
        apply_window_icon(self)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        # esconde enquanto inicializa
        self.withdraw()

        self.sistema = SistemaOrdemProducao(self.cfg)
        self.connected = False

        # totais
        self.total_produtos = 0
        self.total_situacoes = 0
        self.total_depositos = 0
        self.total_ordens = 0

        self.variaveis_quantidade: Optional[Dict[str, Any]] = None

        # splash init sem travar UI
        self.after(50, self._startup_with_splash)

    # ---------------- startup ----------------

    def _startup_with_splash(self):
        splash = Splash(self, "Inicializando")

        def worker():
            splash.set_text("Conectando ao banco...")
            ok = self.sistema.conectar()
            self.connected = ok
            if ok:
                splash.set_info("Conectado")
            else:
                splash.set_info(self.sistema.ultimo_erro or "Falha na conexão")

            splash.set_text("Carregando interface...")
            self.after(0, self._build_ui)

            splash.set_text("Pronto")
            self.after(350, splash.destroy)
#            self.after(400, self.mostrar_tela_entrada)

        threading.Thread(target=worker, daemon=True).start()

    # ---------------- tela entrada ----------------

    def _carregar_avatar_tk(self, caminho: str, max_lado: int = 280) -> tk.PhotoImage | None:
        try:
            img = tk.PhotoImage(file=caminho)
            w, h = img.width(), img.height()
            maior = max(w, h)
            if maior > max_lado:
                fator = max(1, int(maior / max_lado))
                img = img.subsample(fator, fator)
            return img
        except Exception:
            return None

    def mostrar_tela_entrada(self):
        if self._closing:
            return

        tela = tk.Toplevel(self)
        apply_window_icon(tela)
        tela.title("Ekenox - Entrada")
        tela.resizable(False, False)
        tela.configure(bg="#121212")
        tela.protocol("WM_DELETE_WINDOW", self.on_closing)

        frame = tk.Frame(tela, bg="#121212", padx=30, pady=25)
        frame.pack(fill="both", expand=True)

        candidatos = [
            os.path.join(BASE_DIR, "imagens", "avatar_ekenox.png"),
            os.path.join(BASE_DIR, "avatar_ekenox.png"),
            os.path.join(BASE_DIR, "imagens", "Ekenox.png"),
            os.path.join(BASE_DIR, "Ekenox.png"),
        ]
        caminho_avatar = next(
            (p for p in candidatos if os.path.isfile(p)), None)

        avatar_img = self._carregar_avatar_tk(
            caminho_avatar, max_lado=260) if caminho_avatar else None
        if avatar_img:
            lbl_img = tk.Label(frame, image=avatar_img, bg="#121212")
            lbl_img.image = avatar_img
            lbl_img.pack(pady=(0, 15))
        else:
            tk.Label(frame, text="(Avatar não encontrado)", bg="#121212",
                     fg="#aaaaaa", font=("Segoe UI", 10)).pack(pady=(0, 15))

        tk.Label(frame, text="Sistema de Ordem de Produção", bg="#121212",
                 fg="#ffffff", font=("Segoe UI", 14, "bold")).pack()
        tk.Label(frame, text="Ekenox", bg="#121212", fg="#ff9f1a",
                 font=("Segoe UI", 18, "bold")).pack(pady=(2, 18))

        status = "Conectado ao banco" if self.connected else f"ERRO BD: {self.sistema.ultimo_erro or ''}"
        tk.Label(frame, text=status, bg="#121212", fg=(
            "#34d399" if self.connected else "#f87171"), font=("Segoe UI", 10, "bold")).pack(pady=(0, 14))

        botoes = tk.Frame(frame, bg="#121212")
        botoes.pack(fill="x")

        def entrar(event=None):
            try:
                tela.destroy()
            except Exception:
                pass
            self.deiconify()
            self.lift()
            try:
                self.deposito_origem_entry.focus_set()
            except Exception:
                pass

        ttk.Button(botoes, text="Entrar", command=entrar).pack(
            side="left", expand=True, fill="x", padx=(0, 8))
        ttk.Button(botoes, text="Sair", command=self.on_closing).pack(
            side="left", expand=True, fill="x")

        tela.bind("<Return>", entrar)
        tela.bind("<Escape>", lambda e: self.on_closing())

        tela.update_idletasks()
        w, h = tela.winfo_width(), tela.winfo_height()
        x = (tela.winfo_screenwidth() // 2) - (w // 2)
        y = (tela.winfo_screenheight() // 2) - (h // 2)
        tela.geometry(f"+{x}+{y}")

    # ---------------- UI ----------------

    def _build_ui(self):
        try:
            main_frame = ttk.Frame(self, padding=10)
            main_frame.pack(fill=tk.BOTH, expand=True)

            status_frame = ttk.Frame(main_frame)
            status_frame.pack(fill=tk.X)

            self.status_label = ttk.Label(
                status_frame,
                text=(
                    "Conectado ao banco de dados" if self.connected else "Erro ao conectar ao banco"),
                foreground=("green" if self.connected else "red"),
                font=("Segoe UI", 10, "bold"),
            )
            self.status_label.pack(side=tk.LEFT)

            ttk.Separator(main_frame, orient=tk.HORIZONTAL).pack(
                fill=tk.X, pady=10)

            form_frame = ttk.LabelFrame(
                main_frame, text="Nova Ordem de Produção")
            form_frame.pack(fill=tk.BOTH, expand=True)

            for col in range(6):
                form_frame.columnconfigure(
                    col, weight=1 if col in (1, 4) else 0)
            form_frame.rowconfigure(6, weight=1)

            # vars
            self.numero_var = tk.StringVar()
            self.deposito_origem_var = tk.StringVar()
            self.deposito_destino_var = tk.StringVar()
            self.situacao_id_var = tk.StringVar()
            self.produto_id_var = tk.StringVar()
            self.responsavel_var = tk.StringVar()
            self.quantidade_var = tk.StringVar()
            self.valor_var = tk.StringVar()
            self.data_previsao_inicio_var = tk.StringVar()
            self.data_previsao_final_var = tk.StringVar()
            self.data_inicio_var = tk.StringVar()
            self.data_fim_var = tk.StringVar()
            self.saldo_var = tk.StringVar()
            self.media_vendas_var = tk.StringVar()
            self.dia = tk.StringVar()
            self.media_mes = tk.StringVar()
            self.multiplicador = tk.StringVar()
            self.producao_mes = tk.StringVar()
            self.qtd_prod = tk.StringVar()
            self.sugestao_qtd = tk.StringVar()
            self.qtde_sku_var = tk.StringVar()

            # linha 0
            ttk.Label(form_frame, text="Número da Ordem:*").grid(
                row=0, column=0, sticky="e", padx=5, pady=5
            )

            num_frame = ttk.Frame(form_frame)
            num_frame.grid(row=0, column=1, sticky="w", padx=(0, 10), pady=5)

            self.numero_entry = ttk.Entry(
                num_frame, textvariable=self.numero_var, width=15)
            self.numero_entry.grid(row=0, column=0, sticky="w")

            ttk.Button(
                num_frame,
                text="Buscar (F5)",
                width=12,
                command=self.buscar_ordem_por_numero
            ).grid(row=0, column=1, padx=(6, 0), sticky="w")

            ttk.Label(form_frame, text="Responsável:").grid(
                row=0, column=3, sticky="e", padx=5, pady=5
            )
            ttk.Entry(form_frame, textvariable=self.responsavel_var, width=25).grid(
                row=0, column=4, sticky="ew", padx=(0, 10), pady=5
            )

            # (opcional) Enter no campo também busca
            self.numero_entry.bind("<Return>", self.buscar_ordem_por_numero)

            ttk.Label(form_frame, text="Responsável:").grid(
                row=0, column=3, sticky="e", padx=5, pady=5)
            ttk.Entry(form_frame, textvariable=self.responsavel_var, width=25).grid(
                row=0, column=4, sticky="ew", padx=(0, 10), pady=5)

            # linha 1 depósitos
            ttk.Label(form_frame, text="Depósito Origem (ID):*").grid(row=1,
                                                                      column=0, sticky="e", padx=5, pady=5)
            origem_frame = ttk.Frame(form_frame)
            origem_frame.grid(row=1, column=1, columnspan=2,
                              sticky="w", padx=(0, 10), pady=5)

            self.deposito_origem_entry = ttk.Entry(
                origem_frame, textvariable=self.deposito_origem_var, width=18)
            self.deposito_origem_entry.grid(row=0, column=0, sticky="w")

            ttk.Button(origem_frame, text="Listar Depósitos (F4)",
                       command=self.mostrar_depositos_origem).grid(row=0, column=1, padx=(5, 0))

            ttk.Label(form_frame, text="Depósito Destino (ID):*").grid(row=1,
                                                                       column=3, sticky="e", padx=5, pady=5)
            destino_frame = ttk.Frame(form_frame)
            destino_frame.grid(row=1, column=4, columnspan=2,
                               sticky="w", padx=(0, 10), pady=5)

            self.deposito_destino_entry = ttk.Entry(
                destino_frame, textvariable=self.deposito_destino_var, width=18)
            self.deposito_destino_entry.grid(row=0, column=0, sticky="w")

            ttk.Button(destino_frame, text="Listar Depósitos (F8)",
                       command=self.mostrar_depositos_destino).grid(row=0, column=1, padx=(5, 0))

            # linha 2 situação / produto
            ttk.Label(form_frame, text="Situação (ID):*").grid(row=2,
                                                               column=0, sticky="e", padx=5, pady=5)
            situacao_frame = ttk.Frame(form_frame)
            situacao_frame.grid(row=2, column=1, columnspan=2,
                                sticky="w", padx=(0, 10), pady=5)

            self.situacao_entry = ttk.Entry(
                situacao_frame, textvariable=self.situacao_id_var, width=18)
            self.situacao_entry.grid(row=0, column=0, sticky="w")

            ttk.Button(situacao_frame, text="Listar Situações (F3)",
                       command=self.mostrar_situacoes).grid(row=0, column=1, padx=(5, 0))

            ttk.Label(form_frame, text="Produto (ID):*").grid(row=2,
                                                              column=3, sticky="e", padx=5, pady=5)
            prod_frame = ttk.Frame(form_frame)
            prod_frame.grid(row=2, column=4, columnspan=2,
                            sticky="w", padx=(0, 10), pady=5)

            self.produto_entry = ttk.Entry(
                prod_frame, textvariable=self.produto_id_var, width=18)
            self.produto_entry.grid(row=0, column=0, sticky="w")

            ttk.Button(prod_frame, text="Listar Produtos (F2)",
                       command=self.mostrar_produtos).grid(row=0, column=1, padx=(5, 0))
            self.produto_entry.bind(
                "<Return>", self.atualizar_quantidade_producao)
            self.produto_entry.bind(
                "<FocusOut>", self.atualizar_quantidade_producao)

            # linha 3 quantidade
            ttk.Label(form_frame, text="Quantidade:*").grid(row=3,
                                                            column=0, sticky="e", padx=5, pady=5)
            qtd_frame = ttk.Frame(form_frame)
            qtd_frame.grid(row=3, column=1, columnspan=5,
                           sticky="w", padx=(0, 10), pady=5)

            self.quantidade_entry = ttk.Entry(
                qtd_frame, textvariable=self.quantidade_var, width=18)
            self.quantidade_entry.grid(
                row=0, column=0, sticky="w", padx=(0, 6))

            ttk.Button(qtd_frame, text="Detalhes (F6)", command=self.mostrar_detalhes_quantidade).grid(
                row=0, column=1, padx=(0, 6))
            ttk.Button(qtd_frame, text="Analisar Estrutura (F7)",
                       command=self.analisar_estrutura_f7).grid(row=0, column=2)

            # linha 6 observação
            ttk.Label(form_frame, text="Observação:").grid(
                row=6, column=0, sticky="ne", padx=5, pady=5)
            self.observacao_text = tk.Text(form_frame, height=6)
            self.observacao_text.grid(
                row=6, column=1, columnspan=5, sticky="nsew", padx=(0, 10), pady=5)

            # botões
            botoes = ttk.Frame(form_frame)
            botoes.grid(row=7, column=0, columnspan=6, pady=15)

            ttk.Button(botoes, text="Salvar Ordem",
                       command=self.salvar_ordem).pack(side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Limpar", command=self.limpar_formulario).pack(
                side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Relatório (F9)",
                       command=self.exportar_relatorio_bling_excel_f9).pack(side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Ordens Existentes (F10)",
                       command=self.mostrar_ordens_producao).pack(side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Finalizar Pendentes (F11)",
                       command=self.finalizar_producoes_pendentes).pack(side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Etiquetas (F12)",
                       command=self.mod_etiquetas.open).pack(side=tk.LEFT, padx=5)

            ttk.Button(botoes, text="Fechar", command=self.on_closing).pack(
                side=tk.LEFT, padx=5)

            ttk.Label(main_frame, text="Campos marcados com * são obrigatórios.",
                      foreground="gray").pack(side=tk.BOTTOM, anchor="w", pady=(5, 0))

            # binds globais (funcionam mesmo com Toplevel em foco)
            for seq, func in [
                ("<F2>",  self.mostrar_produtos),
                ("<F3>",  self.mostrar_situacoes),
                ("<F4>",  self.mostrar_depositos_origem),
                ("<F5>",  self.buscar_ordem_por_numero),   # <-- aqui
                ("<F6>",  self.mostrar_detalhes_quantidade),
                ("<F7>",  self.analisar_estrutura_f7),
                ("<F8>",  self.mostrar_depositos_destino),
                ("<F9>",  self.exportar_relatorio_bling_excel_f9),
                ("<F10>", self.mostrar_ordens_producao),
                ("<F11>", self.finalizar_producoes_pendentes),
                ("<F12>", self._on_f12),
            ]:
                self.bind_all(seq, func)

            # número
            if self.connected:
                self.numero_var.set(str(self.sistema.gerar_numero_ordem()))

        except Exception as e:
            log_path = log_exception(e, "Falha ao montar UI")
            messagebox.showerror(
                "Erro", f"Falha ao montar interface.\nLog: {log_path}")

    # ---------------- helpers ----------------

    def parse_int(self, valor_str: str, campo: str) -> int:
        if not valor_str.strip():
            raise ValueError(f"O campo '{campo}' é obrigatório.")
        return int(valor_str.strip())

    def parse_float(self, valor_str: str, campo: str, obrigatorio: bool = False) -> Optional[float]:
        if not valor_str.strip():
            if obrigatorio:
                raise ValueError(f"O campo '{campo}' é obrigatório.")
            return None
        return float(valor_str.replace(",", "."))

    # ---------------- atualizar quantidade (simples) ----------------

    def atualizar_quantidade_producao(self, event=None):
        try:
            pid_str = self.produto_id_var.get().strip()
            if not pid_str:
                self.quantidade_var.set("0")
                self.variaveis_quantidade = None
                return

            if not self.connected:
                self.quantidade_var.set("1.00")
                self.variaveis_quantidade = {
                    "obs": "Sem conexão: quantidade sugerida = 1"}
                return

            pid = int(pid_str)
            produto = self.sistema.validar_produto(pid)
            if not produto:
                self.quantidade_var.set("0")
                self.variaveis_quantidade = {"erro": "Produto não encontrado"}
                return

            preco = float(produto.get("preco") or 0)
            self.valor_var.set(f"{preco:.2f}")

            # saldo / media
            saldo = float(self.sistema.saldo_fisico(pid) or 0.0)
            self.saldo_var.set(f"{saldo:.2f}")

            media_vendas = float(self.sistema.media_vendas_mensal(pid) or 0.0)
            self.media_vendas_var.set(f"{media_vendas:.2f}")

            dia = int(date.today().day or 1)
            if dia <= 0:
                dia = 1
            self.dia.set(dia)

            # qtd_prod = estoqueMaximo (F6)
            qtd_prod_rows = self.sistema.f6_buscar_estrutura(pid)
            qtd_prod = float(
                qtd_prod_rows[0][0]) if qtd_prod_rows and qtd_prod_rows[0][0] is not None else 0.0
            self.qtd_prod.set(f"{qtd_prod:.2f}")

            # média diária
            if media_vendas <= 0:
                media_dia = 1.0
            else:
                media_dia = media_vendas / dia

            multiplicador = 7  # 7 dias de produção
            producao_media = media_dia * multiplicador

            # sugestão "crua" (como era)
            sugestao_qtd_calc = max(0.0, qtd_prod - producao_media)

            # --------------------------
            # ARRANJO: buscar tamanho do lote e arredondar para múltiplo
            # --------------------------
            sku = (produto.get("sku") or "").strip()

            # regra do N:
            # - no salvar_ordem você usa sku com N (quando existir)
            # - aqui vamos usar a MESMA lógica para consultar o arranjo
            if sku and sku[-1].upper() == "N":
                sku_arranjo = sku
            else:
                sku_arranjo = sku

            qtd_arranjo = float(
                self.sistema.buscar_qtd_produzir_por_sku(sku_arranjo) or 0.0)
            self.qtde_sku_var.set(f"{qtd_arranjo:.2f}")

            # arredondar para múltiplo do arranjo (sempre pra cima)
            if sugestao_qtd_calc <= 0:
                sugestao_qtd_final = 0.0
            elif qtd_arranjo > 0:
                sugestao_qtd_final = ceil(
                    sugestao_qtd_calc / qtd_arranjo) * qtd_arranjo
            else:
                # se não houver arranjo cadastrado, mantém o valor calculado
                sugestao_qtd_final = sugestao_qtd_calc

            self.sugestao_qtd.set(f"{sugestao_qtd_final:.2f}")

            # o campo Quantidade vai receber SEMPRE o múltiplo do arranjo
            self.quantidade_var.set(f"{sugestao_qtd_final:.2f}")

            # detalhes (F6)
            self.variaveis_quantidade = {
                "Produto id                   ": pid,
                "Produto nome                 ": produto.get("nomeproduto"),
                "sku                          ": produto.get("sku"),
                "Preco                        ": preco,
                "Saldo                        ": saldo,
                "Media vendas mês             ": media_vendas,
                "Dia atual                    ": dia,
                "Producao media (dia)         ": media_dia,
                "Previsão dias                ": multiplicador,
                "Producao média (7 dias)      ": producao_media,
                "Quantidade produção máxima   ": qtd_prod,

                # antes/depois arranjo
                "Sugestão calculada           ": sugestao_qtd_calc,
                "Arranjo (lote)               ": qtd_arranjo,
                "Sugestão (múltiplo arranjo)  ": sugestao_qtd_final,

                "obs": (
                    "Sugestão = (Quantidade produção máxima - (média diária * 7)). "
                    "Depois, arredonda para CIMA no próximo múltiplo do ARRANJO (se existir)."
                ),
            }

        except Exception as e:
            self.quantidade_var.set("0")
            self.variaveis_quantidade = {"erro": str(e)}

    def mostrar_detalhes_quantidade(self, event=None):
        if not self.variaveis_quantidade:
            messagebox.showinfo("Detalhes da Quantidade",
                                "Nenhum cálculo realizado ainda.")
            return

        janela = tk.Toplevel(self)
        janela.title("Detalhes do cálculo da quantidade")
        janela.geometry("600x500")
        janela.transient(self)
        janela.grab_set()

        frame = ttk.Frame(janela, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        txt = tk.Text(frame, wrap="word")
        txt.pack(fill=tk.BOTH, expand=True)

        txt.insert(tk.END, "Variáveis:\n\n")
        for k, v in self.variaveis_quantidade.items():
            txt.insert(tk.END, f"{k}: {v}\n")
        txt.config(state="disabled")

        ttk.Button(frame, text="Fechar", command=janela.destroy).pack(pady=5)
        janela.bind("<Escape>", lambda e: janela.destroy())

    # ============================================================
    # F7 - ANALISAR ESTRUTURA (lendo do BANCO)
    # ============================================================

    def analisar_estrutura_f7(self, event=None):
        if not self.connected:
            messagebox.showerror("F7 - Estrutura", "Sem conexão com o banco.")
            return

        prod = (self.produto_id_var.get() or "").strip()
        qtd_txt = (self.quantidade_var.get() or "").strip()

        if not prod:
            messagebox.showerror("F7 - Estrutura", "Informe o Produto (ID).")
            return

        try:
            produto_id = int(prod)
        except ValueError:
            messagebox.showerror("F7 - Estrutura", "Produto (ID) inválido.")
            return

        try:
            qtd_produzir = float(qtd_txt.replace(",", ".")) if qtd_txt else 0.0
        except ValueError:
            qtd_produzir = 0.0

        if qtd_produzir <= 0:
            messagebox.showerror(
                "F7 - Estrutura", "Quantidade para produzir deve ser maior que zero.")
            return

        try:
            itens = self.sistema.f7_buscar_estrutura(produto_id)
        except Exception as e:
            messagebox.showerror(
                "F7 - Estrutura", f"Erro ao ler Ekenox.estrutura:\n{e}")
            return

        if not itens:
            messagebox.showinfo(
                "F7 - Estrutura", "Sem estrutura cadastrada para este produto.")
            return

        # monta linhas com: componente, nome, qtd_base, necessária, saldo, falta, min, max, fornecedor, precoCompra
        linhas = []
        faltantes = 0
        itens_faltantes_para_pedido: List[Dict[str, Any]] = []

        for (componente, qtd_base) in itens:
            qtd_necessaria = float(qtd_base) * float(qtd_produzir)
            saldo = self.sistema.saldo_fisico(int(componente))
            falta = max(0.0, qtd_necessaria - saldo)

            info = self.sistema.f7_buscar_info_produto(int(componente))
            est_min = float(info.get("estoqueMinimo", 0.0) or 0.0)
            est_max = float(info.get("estoqueMaximo", 0.0) or 0.0)
            preco_compra = float(info.get("precoCompra", 0.0) or 0.0)

            fornecedor_nome = ""
            fk_fornecedor = int(info.get("fkFornecedor", 0) or 0)
            if fk_fornecedor:
                forn = self.sistema.f7_buscar_fornecedor(fk_fornecedor)
                fornecedor_nome = (forn.get("nome") or "").strip()

            prod_comp = self.sistema.validar_produto(int(componente))
            comp_nome = (prod_comp.get("nomeproduto")
                         if prod_comp else "") or ""

            if falta > 0:
                faltantes += 1

                # sugestão de compra: arredonda pra cima
                qtd_comprar = ceil(falta)

                itens_faltantes_para_pedido.append({
                    "fornecedor": fornecedor_nome or "SEM FORNECEDOR",
                    "descricao": f"{componente} - {comp_nome}".strip(" -"),
                    "qtd_comprar": float(qtd_comprar),
                    "estoque_atual": float(saldo),
                    "estoque_minimo": float(est_min),
                    "estoque_maximo": float(est_max),
                    "valor_unitario": float(preco_compra) if preco_compra > 0 else None,
                })

            linhas.append((
                int(componente),
                comp_nome,
                float(qtd_base),
                float(qtd_necessaria),
                float(saldo),
                float(falta),
                float(est_min),
                float(est_max),
                fornecedor_nome,
                float(preco_compra),
            ))

        win = tk.Toplevel(self)
        win.title(f"F7 - Estrutura | Produto {produto_id}")
        win.geometry(self.cfg.f7_geometry)
        win.minsize(1050, 520)
        win.transient(self)
        win.grab_set()
        apply_window_icon(win)

        top = ttk.Frame(win, padding=10)
        top.pack(fill=tk.X)

        ttk.Label(
            top,
            text=f"Produto: {produto_id} | Produzir: {qtd_produzir:.2f} | Itens: {len(linhas)} | Faltando: {faltantes}",
            font=("Segoe UI", 10, "bold"),
        ).pack(side=tk.LEFT)

        def gerar_pedido():
            if not itens_faltantes_para_pedido:
                messagebox.showinfo(
                    "Pedido de Compra", "Não há faltantes para gerar pedido.", parent=win)
                return
            if not os.path.exists(self.cfg.caminho_modelo):
                messagebox.showerror(
                    "Pedido de Compra", f"Modelo não encontrado:\n{self.cfg.caminho_modelo}", parent=win)
                return

            numero_inicial = simpledialog.askinteger(
                "Pedido de Compra",
                "Informe o número inicial do pedido:",
                parent=win,
                minvalue=1
            )
            if not numero_inicial:
                return

            dados_excel = []
            hoje = date.today()
            numero_atual = int(numero_inicial)

            grupos = defaultdict(list)
            for it in itens_faltantes_para_pedido:
                grupos[it["fornecedor"]].append(it)

            for fornecedor, itens_f in grupos.items():
                for it in itens_f:
                    dados_excel.append({
                        "fornecedor": fornecedor,
                        "numero_pedido": numero_atual,
                        "data_pedido": hoje,
                        "produto": it["descricao"],
                        "quantidade": float(it["qtd_comprar"]),
                        "estoque_atual": float(it.get("estoque_atual", 0.0) or 0.0),
                        "estoque_minimo": float(it.get("estoque_minimo", 0.0) or 0.0),
                        "estoque_maximo": float(it.get("estoque_maximo", 0.0) or 0.0),
                        "valor_unitario": it.get("valor_unitario"),
                    })
                numero_atual += 1

            try:
                gerar_abas_fornecedor_pedido(
                    dados=dados_excel,
                    nome_aba_modelo="Pedido de Compra",
                    caminho_modelo=self.cfg.caminho_modelo,
                    caminho_saida=self.cfg.caminho_saida,
                )
                messagebox.showinfo(
                    "Pedido de Compra", f"Gerado em:\n{self.cfg.caminho_saida}", parent=win)
            except Exception as e:
                messagebox.showerror("Pedido de Compra",
                                     f"Falha ao gerar Excel:\n{e}", parent=win)

        ttk.Button(top, text="Gerar Pedido (faltantes)",
                   command=gerar_pedido).pack(side=tk.RIGHT)

        # tabela
        frame = ttk.Frame(win, padding=(10, 0, 10, 10))
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("comp", "nome", "base", "necessaria", "saldo",
                "falta", "min", "max", "forn", "preco")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        headings = {
            "comp": "Componente (ID)",
            "nome": "Nome",
            "base": "Qtd Estrutura",
            "necessaria": "Qtd Necessária",
            "saldo": "Saldo",
            "falta": "Falta",
            "min": "Est. Mín",
            "max": "Est. Máx",
            "forn": "Fornecedor",
            "preco": "Preço Compra",
        }
        for c in cols:
            tree.heading(c, text=headings[c])

        tree.column("comp", width=130, anchor="w")
        tree.column("nome", width=260, anchor="w")
        tree.column("base", width=120, anchor="e")
        tree.column("necessaria", width=130, anchor="e")
        tree.column("saldo", width=110, anchor="e")
        tree.column("falta", width=110, anchor="e")
        tree.column("min", width=90, anchor="e")
        tree.column("max", width=90, anchor="e")
        tree.column("forn", width=210, anchor="w")
        tree.column("preco", width=110, anchor="e")

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree.tag_configure("faltando", foreground="red")

        for (comp, nome, base, nec, saldo, falta, est_min, est_max, forn, preco) in linhas:
            vals = (
                str(comp),
                nome or "",
                f"{base:.4f}",
                f"{nec:.4f}",
                f"{saldo:.4f}",
                f"{falta:.4f}",
                f"{est_min:.2f}",
                f"{est_max:.2f}",
                forn or "",
                f"{preco:.2f}" if preco else "",
            )
            if falta > 0:
                tree.insert("", tk.END, values=vals, tags=("faltando",))
            else:
                tree.insert("", tk.END, values=vals)

        def close():
            try:
                self.cfg.f7_geometry = win.geometry()
                save_config(self.cfg)
            except Exception:
                pass
            win.destroy()

        win.bind("<Escape>", lambda e: close())
        win.protocol("WM_DELETE_WINDOW", close)

    # ============================================================
    # F12 - Etiquetas (abre etiqueta.exe)
    # ============================================================
    def bring_window_to_front_by_title(self, title_contains: str) -> bool:
        """Windows: tenta achar janela pelo título e trazer para frente."""
        try:
            import ctypes
            from ctypes import wintypes

            user32 = ctypes.windll.user32
            EnumWindows = user32.EnumWindows
            EnumWindowsProc = ctypes.WINFUNCTYPE(
                ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
            GetWindowTextW = user32.GetWindowTextW
            GetWindowTextLengthW = user32.GetWindowTextLengthW
            IsWindowVisible = user32.IsWindowVisible
            ShowWindow = user32.ShowWindow
            SetForegroundWindow = user32.SetForegroundWindow

            found = {"hwnd": None}

            def foreach(hwnd, lParam):
                if not IsWindowVisible(hwnd):
                    return True
                length = GetWindowTextLengthW(hwnd)
                if length <= 0:
                    return True
                buf = ctypes.create_unicode_buffer(length + 1)
                GetWindowTextW(hwnd, buf, length + 1)
                if title_contains.lower() in buf.value.lower():
                    found["hwnd"] = hwnd
                    return False
                return True

            EnumWindows(EnumWindowsProc(foreach), 0)

            if found["hwnd"]:
                SW_RESTORE = 9
                ShowWindow(found["hwnd"], SW_RESTORE)
                SetForegroundWindow(found["hwnd"])
                return True
        except Exception:
            pass
        return False

    def abrir_programa_etiqueta(self, event=None):
        exe_path = os.path.join(BASE_DIR, "etiqueta.exe")
        if not os.path.isfile(exe_path):
            messagebox.showerror(
                "Etiquetas (F12)", f"Não encontrei:\n{exe_path}")
            return
        try:
            if os.name == "nt":
                os.startfile(exe_path)
            else:
                subprocess.Popen([exe_path], cwd=BASE_DIR)

            def _retry(attempt=0):
                ok = self.bring_window_to_front_by_title("etiqueta")
                if not ok and attempt < 12:
                    self.after(500, lambda: _retry(attempt + 1))

            self.after(500, lambda: _retry(0))
        except Exception as e:
            messagebox.showerror(
                "Etiquetas (F12)", f"Falha ao abrir etiqueta.exe:\n{e}")

    # ============================================================
    # F10 - Ordens existentes
    # ============================================================

    def mostrar_ordens_producao(self, event=None):
        if not self.connected:
            messagebox.showerror("F10 - Ordens", "Não há conexão com o banco.")
            return

        ordens = self.sistema.listar_ordens_producao()
        if not ordens:
            messagebox.showinfo("F10 - Ordens", "Nenhuma ordem encontrada.")
            return

        win = tk.Toplevel(self)
        win.title("F10 - Ordens Existentes")
        win.geometry("1050x560")
        win.transient(self)
        win.grab_set()
        apply_window_icon(win)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("id", "numero", "produto_id", "produto_nome",
                "situacao", "quantidade", "data_inicio", "data_fim")
        tree = ttk.Treeview(frame, columns=cols,
                            show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        for c in cols:
            tree.heading(c, text=c)

        tree.column("id", width=70, anchor="center")
        tree.column("numero", width=90, anchor="center")
        tree.column("produto_id", width=110, anchor="center")
        tree.column("produto_nome", width=280, anchor="w")
        tree.column("situacao", width=180, anchor="w")
        tree.column("quantidade", width=120, anchor="e")
        tree.column("data_inicio", width=120, anchor="center")
        tree.column("data_fim", width=120, anchor="center")

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        def fmt(dt):
            if not dt:
                return ""
            if isinstance(dt, (datetime, date)):
                return dt.strftime("%d/%m/%Y")
            return str(dt)

        for (oid, numero, produto_id, produto_nome, _, situacao_nome, quantidade, data_inicio, data_fim) in ordens:
            tree.insert(
                "",
                tk.END,
                values=(
                    oid,
                    numero,
                    produto_id,
                    produto_nome or "",
                    situacao_nome or "",
                    f"{float(quantidade):.2f}" if quantidade is not None else "",
                    fmt(data_inicio),
                    fmt(data_fim),
                ),
            )

        btns = ttk.Frame(win, padding=(10, 0, 10, 10))
        btns.pack(fill=tk.X)

        def excluir_selecionada(event=None):
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Excluir", "Selecione uma ordem.")
                return

            # exclui apenas a primeira selecionada (padrão)
            item_id = sel[0]
            v = tree.item(item_id)["values"]
            oid, numero = v[0], v[1]

            if not messagebox.askyesno("Confirmar", f"Deseja excluir a OP nº {numero} (ID {oid})?"):
                return

            ok = self.sistema.excluir_ordem_producao(int(oid))
            if ok:
                tree.delete(item_id)
                messagebox.showinfo("Exclusão", f"OP nº {numero} excluída.")
            else:
                messagebox.showerror("Erro", "Não foi possível excluir.")

        ttk.Button(btns, text="Excluir (DEL)", command=excluir_selecionada).pack(
            side=tk.RIGHT, padx=(0, 8))
        ttk.Button(btns, text="Fechar",
                   command=win.destroy).pack(side=tk.RIGHT)

        win.bind("<Escape>", lambda e: win.destroy())
        tree.bind("<Delete>", excluir_selecionada)

    # ============================================================
    # F11 - Finalizar pendentes
    # ============================================================
    def finalizar_producoes_pendentes(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "F11 - Finalizar", "Não há conexão com o banco.")
            return

        pendentes = self.sistema.listar_ordens_sem_data_fim()
        if not pendentes:
            messagebox.showinfo("F11 - Finalizar",
                                "Não há ordens pendentes sem data fim.")
            return

        win = tk.Toplevel(self)
        win.title("F11 - Finalizar Ordens Pendentes")
        win.geometry("1050x560")
        win.transient(self)
        win.grab_set()
        apply_window_icon(win)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("id", "numero", "produto_id", "produto_nome",
                "situacao", "quantidade", "data_inicio")
        tree = ttk.Treeview(frame, columns=cols,
                            show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        for c in cols:
            tree.heading(c, text=c)

        tree.column("id", width=70, anchor="center")
        tree.column("numero", width=90, anchor="center")
        tree.column("produto_id", width=110, anchor="center")
        tree.column("produto_nome", width=320, anchor="w")
        tree.column("situacao", width=180, anchor="w")
        tree.column("quantidade", width=120, anchor="e")
        tree.column("data_inicio", width=120, anchor="center")

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        def fmt(dt):
            if not dt:
                return ""
            if isinstance(dt, (datetime, date)):
                return dt.strftime("%d/%m/%Y")
            return str(dt)

        for (oid, numero, produto_id, produto_nome, _, situacao_nome, quantidade, data_inicio) in pendentes:
            tree.insert(
                "",
                tk.END,
                values=(
                    oid,
                    numero,
                    produto_id,
                    produto_nome or "",
                    situacao_nome or "",
                    f"{float(quantidade):.2f}" if quantidade is not None else "",
                    fmt(data_inicio),
                ),
            )

        btns = ttk.Frame(win, padding=(10, 0, 10, 10))
        btns.pack(fill=tk.X)

        def finalizar_selecionadas(event=None):
            sel = tree.selection()
            if not sel:
                messagebox.showwarning(
                    "Finalizar", "Selecione uma ou mais ordens.")
                return

            ordens_sel = []
            for item_id in sel:
                v = tree.item(item_id)["values"]
                ordens_sel.append((item_id, int(v[0]), v[1]))

            if len(ordens_sel) == 1:
                _, oid, numero = ordens_sel[0]
                msg = f"Deseja finalizar a OP nº {numero} (ID {oid}) com data fim hoje?"
            else:
                nums = ", ".join(str(x[2]) for x in ordens_sel)
                msg = f"Deseja finalizar {len(ordens_sel)} OPs (números: {nums}) com data fim hoje?"

            if not messagebox.askyesno("Confirmar", msg):
                return

            ok_count = 0
            for item_id, oid, _ in ordens_sel:
                ok = self.sistema.finalizar_ordem_individual(oid)
                if ok:
                    tree.delete(item_id)
                    ok_count += 1

            if ok_count:
                messagebox.showinfo(
                    "Finalização", f"{ok_count} ordem(ns) finalizada(s) com sucesso!")
            else:
                messagebox.showerror(
                    "Erro", "Não foi possível finalizar as ordens selecionadas.")
            self.sistema.conn.commit()

        ttk.Button(btns, text="Finalizar selecionadas (ENTER)",
                   command=finalizar_selecionadas).pack(side=tk.RIGHT, padx=(0, 8))
        ttk.Button(btns, text="Fechar",
                   command=win.destroy).pack(side=tk.RIGHT)

        win.bind("<Escape>", lambda e: win.destroy())
        tree.bind("<Return>", finalizar_selecionadas)
        tree.bind("<Double-Button-1>", finalizar_selecionadas)

    # ---------------- listas básicas ----------------

    def mostrar_produtos(self, event=None):
        if not self.connected:
            messagebox.showerror("Erro", "Não há conexão com o banco.")
            return

        produtos = self.sistema.listar_produtos_disponiveis()
        if not produtos:
            messagebox.showinfo("Produtos", "Nenhum produto encontrado.")
            return

        janela = tk.Toplevel(self)
        janela.title("Produtos - Duplo clique para selecionar")
        janela.geometry("980x520")
        janela.transient(self)
        janela.grab_set()

        frame = ttk.Frame(janela, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("ID", "Nome", "SKU", "Preço", "Tipo")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        for c in cols:
            tree.heading(c, text=c)

        tree.column("ID", width=120)
        tree.column("Nome", width=420)
        tree.column("SKU", width=170)
        tree.column("Preço", width=120, anchor="e")
        tree.column("Tipo", width=120)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        for p in produtos:
            preco = f"{float(p[3]):.2f}" if p[3] else "0.00"
            tree.insert("", tk.END, values=(p[0], p[1], p[2], preco, p[4]))

        def selecionar(_):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0])["values"]
            self.produto_id_var.set(str(vals[0]))
            self.valor_var.set(str(vals[3]).replace(",", "."))
            janela.destroy()
            self.atualizar_quantidade_producao()
            self.quantidade_entry.focus_set()

        tree.bind("<Double-Button-1>", selecionar)
        janela.bind("<Escape>", lambda e: janela.destroy())

    def mostrar_situacoes(self, event=None):
        if not self.connected:
            messagebox.showerror("Erro", "Não há conexão com o banco.")
            return

        situacoes = self.sistema.listar_situacoes_disponiveis()
        if not situacoes:
            messagebox.showinfo("Situações", "Nenhuma situação encontrada.")
            return

        janela = tk.Toplevel(self)
        janela.title("Situações - Duplo clique para selecionar")
        janela.geometry("620x440")
        janela.transient(self)
        janela.grab_set()

        frame = ttk.Frame(janela, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("ID", "Situação")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        tree.heading("ID", text="ID")
        tree.heading("Situação", text="Situação")
        tree.column("ID", width=140)
        tree.column("Situação", width=420)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        for s in situacoes:
            tree.insert("", tk.END, values=(s[0], s[1]))

        def selecionar(_):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0])["values"]
            self.situacao_id_var.set(str(vals[0]))
            janela.destroy()

        tree.bind("<Double-Button-1>", selecionar)
        janela.bind("<Escape>", lambda e: janela.destroy())

    def mostrar_depositos_origem(self, event=None):
        self._mostrar_depositos("origem")

    def mostrar_depositos_destino(self, event=None):
        self._mostrar_depositos("destino")

    def _mostrar_depositos(self, modo: str):
        if not self.connected:
            messagebox.showerror("Erro", "Não há conexão com o banco.")
            return

        depositos = self.sistema.listar_depositos_disponiveis()
        if not depositos:
            messagebox.showinfo("Depósitos", "Nenhum depósito encontrado.")
            return

        janela = tk.Toplevel(self)
        janela.title("Depósitos - Duplo clique para selecionar")
        janela.geometry("820x460")
        janela.transient(self)
        janela.grab_set()

        frame = ttk.Frame(janela, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("ID", "Descrição", "Situação", "Padrão", "Desconsiderar saldo")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        for c in cols:
            tree.heading(c, text=c)

        tree.column("ID", width=140)
        tree.column("Descrição", width=280)
        tree.column("Situação", width=120)
        tree.column("Padrão", width=100)
        tree.column("Desconsiderar saldo", width=160)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        for d in depositos:
            tree.insert("", tk.END, values=(d[0], d[1], d[2], d[3], d[4]))

        def selecionar(_):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0])["values"]
            dep_id = str(vals[0])
            if modo == "origem":
                self.deposito_origem_var.set(dep_id)
            else:
                self.deposito_destino_var.set(dep_id)
            janela.destroy()

        tree.bind("<Double-Button-1>", selecionar)
        janela.bind("<Escape>", lambda e: janela.destroy())

    # ---------------- salvar OP ----------------

    def salvar_ordem(self):
        if not self.connected:
            messagebox.showerror("Erro", "Não há conexão com o banco.")
            return

        try:
            dados: Dict[str, Any] = {}

            dados["numero"] = self.numero_var.get().strip()
            if not dados["numero"]:
                raise ValueError("O campo 'Número da Ordem' é obrigatório.")

            dados["deposito_id_origem"] = self.parse_int(
                self.deposito_origem_var.get(), "Depósito Origem")
            dados["deposito_id_destino"] = self.parse_int(
                self.deposito_destino_var.get(), "Depósito Destino")
            dados["situacao_id"] = self.parse_int(
                self.situacao_id_var.get(), "Situação")
            dados["fkprodutoid"] = self.parse_int(
                self.produto_id_var.get(), "Produto")
            dados["quantidade"] = self.parse_float(
                self.quantidade_var.get(), "Quantidade", obrigatorio=True)

            dados["responsavel"] = (self.responsavel_var.get().strip() or None)
            dados["valor"] = self.parse_float(
                self.valor_var.get(), "Valor", obrigatorio=False)

            obs = self.observacao_text.get("1.0", tk.END).strip()
            dados["observacao"] = obs if obs else None

            dados["id"] = None
            dados["data_previsao_inicio"] = None
            dados["data_previsao_final"] = None
            dados["data_inicio"] = None
            dados["data_fim"] = None
            dados["transmite_bling"] = 0  # opcional

            # --------------------------
            # Define o "tamanho do lote" pelo ARRANJO (via SKU)
            # --------------------------
            produto = self.sistema.validar_produto(int(dados["fkprodutoid"]))
            if not produto:
                raise ValueError("Produto não encontrado para gerar OP.")

            sku = (produto.get("sku") or "").strip()
            # Se vocês usam a regra do 'N' final, mantenha:
            if sku and sku[-1].upper() == "N":
                # normalmente o arranjo está com N (ex.: VIX9444N)
                sku_arranjo = sku
            else:
                sku_arranjo = sku

            qtd_arranjo = float(
                self.sistema.buscar_qtd_produzir_por_sku(sku_arranjo) or 0.0)
            qtd_total = float(dados["quantidade"] or 0.0)

            # Se não houver arranjo (0) ou a quantidade total já couber em 1 OP, segue como antes
            if qtd_arranjo <= 0:
                partes = [qtd_total]
            else:
                # Quebra em OPs com no máximo qtd_arranjo
                partes = []
                restante = qtd_total
                # evita loop infinito por número “quase zero”
                eps = 1e-9

                while restante > eps:
                    lote = min(qtd_arranjo, restante)
                    # normaliza lote (evita -0.00)
                    if lote < eps:
                        break
                    partes.append(lote)
                    restante -= lote

            # Converte o número base da OP para ir incrementando
            try:
                numero_base = int(str(dados["numero"]).strip())
            except Exception:
                numero_base = int(self.sistema.gerar_numero_ordem())

            # Confirmação mostrando como vai quebrar
            if len(partes) == 1:
                msg_conf = f"Confirma inserir OP nº {numero_base}?"
            else:
                resumo = ", ".join([f"{p:.2f}" for p in partes])
                msg_conf = (
                    f"Confirma inserir {len(partes)} OPs a partir do nº {numero_base}?\n\n"
                    f"Quantidade total: {qtd_total:.2f}\n"
                    f"Máximo por OP (arranjo): {qtd_arranjo:.2f}\n"
                    f"Lotes: {resumo}"
                )

            # --------------------------
            # VALIDA ESTOQUE DOS INSUMOS (bloqueia se negativo ou insuficiente)
            # - valida o PIOR caso (maior lote), ou valida cada lote.
            # - para não ter surpresa, vou validar cada lote.
            # --------------------------
            fkproduto_final = int(dados["fkprodutoid"])

            problemas_gerais = []
            for i, qtd_lote in enumerate(partes):
                res = self.sistema.validar_estoque_insumos_para_producao(
                    fkproduto=fkproduto_final,
                    qtd_produzir=float(qtd_lote),
                    bloquear_se_saldo_negativo=True,
                    bloquear_se_insuficiente=True,
                )
                if not res.get("ok"):
                    problemas_gerais.append(
                        (i, float(qtd_lote), res.get("problemas", [])))

            if problemas_gerais:
                # monta mensagem amigável
                linhas_msg = []
                for (idx, lote, probs) in problemas_gerais:
                    num_op = numero_base + idx
                    linhas_msg.append(f"\nOP {num_op} | Lote: {lote:.2f}")
                    for p in probs[:40]:  # evita msg gigante
                        comp = p.get("componente", "")
                        nome = (p.get("nome") or "")
                        necessario = float(p.get("necessario") or 0.0)
                        saldo = float(p.get("saldo") or 0.0)
                        falta = float(p.get("falta") or 0.0)
                        motivo = p.get("motivo", "Problema")
                        linhas_msg.append(
                            f" - {motivo}: {comp} {nome} | Nec: {necessario:.4f} | Saldo: {saldo:.4f} | Falta: {falta:.4f}"
                        )

                    if len(probs) > 40:
                        linhas_msg.append(f" ... (+{len(probs)-40} itens)")

                messagebox.showerror(
                    "Bloqueado",
                    "Não é possível salvar a OP.\n"
                    "Existem insumos com saldo negativo e/ou saldo menor que o necessário:\n"
                    + "\n".join(linhas_msg)
                )
                return

            if not messagebox.askyesno("Confirmar", msg_conf):
                return

            # --------------------------
            # Insere uma OP por lote
            # --------------------------
            op_criadas = []

            for i, qtd_lote in enumerate(partes):
                dados_lote = dict(dados)  # cópia
                dados_lote["id"] = None
                dados_lote["numero"] = str(numero_base + i)
                dados_lote["quantidade"] = float(qtd_lote)

                ok, err = self.sistema.inserir_ordem_producao(dados_lote)
                if not ok:
                    messagebox.showerror(
                        "Erro ao inserir",
                        f"Falha ao inserir OP nº {dados_lote['numero']}.\n\n{err}"
                    )
                    return

                # ✅ Envia para o Bling (assíncrono, não trava a UI)
    #            threading.Thread(
    #                target=enviar_op_bling_thread,
    #                args=(self, self.cfg, dados),
    #                daemon=True
    #            ).start()

                # webhook opcional (manda 1 payload por OP)
                if N8N_WEBHOOK_URL:
                    try:
                        payload = {
                            "numero": dados_lote["numero"],
                            "deposito_id_origem": dados_lote["deposito_id_origem"],
                            "deposito_id_destino": dados_lote["deposito_id_destino"],
                            "situacao_id": dados_lote["situacao_id"],
                            "fkprodutoid": dados_lote["fkprodutoid"],
                            "quantidade": float(dados_lote["quantidade"] or 0),
                            "responsavel": dados_lote.get("responsavel"),
                            "observacao": dados_lote.get("observacao"),
                        }
                        requests.post(N8N_WEBHOOK_URL,
                                      json=payload, timeout=10)
                    except Exception:
                        pass

                op_criadas.append(
                    (dados_lote["numero"], float(dados_lote["quantidade"])))

            if len(op_criadas) == 1:
                messagebox.showinfo(
                    "Sucesso", f"OP {op_criadas[0][0]} inserida com sucesso!")
            else:
                lista = "\n".join(
                    [f"OP {n} - {q:.2f}" for (n, q) in op_criadas])
                messagebox.showinfo(
                    "Sucesso",
                    f"{len(op_criadas)} OPs inseridas com sucesso!\n\n{lista}"
                )

            self.limpar_formulario()

        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def limpar_formulario(self):
        if self.connected:
            self.numero_var.set(str(self.sistema.gerar_numero_ordem()))
        else:
            self.numero_var.set("")

        self.deposito_origem_var.set("")
        self.deposito_destino_var.set("")
        self.situacao_id_var.set("")
        self.produto_id_var.set("")
        self.responsavel_var.set("")
        self.quantidade_var.set("")
        self.valor_var.set("")
        self.observacao_text.delete("1.0", tk.END)
        self.variaveis_quantidade = None

    # ---------------- close ----------------

    def on_closing(self):
        if self._closing:
            return
        self._closing = True
        try:
            if messagebox.askokcancel("Sair", "Deseja realmente sair?"):
                try:
                    self.sistema.desconectar()
                except Exception:
                    pass
                self.destroy()
            else:
                self._closing = False
        except Exception:
            try:
                self.sistema.desconectar()
            except Exception:
                pass
            try:
                self.destroy()
            except Exception:
                pass


# ============================================================
# MAIN
# ============================================================

def build_bling_payload(dados: Dict[str, Any]) -> Dict[str, Any]:
    """
    ⚠️ Ajuste os nomes dos campos conforme a API do Bling.
    Payload base conforme seu formulário.
    """
    return {
        "numero": str(dados["numero"]),
        "depositoOrigem": {"id": int(dados["deposito_id_origem"])},
        "depositoDestino": {"id": int(dados["deposito_id_destino"])},
        "itens": [
            {
                "produto": {"id": int(dados["fkprodutoid"])},
                "quantidade": float(dados["quantidade"]),
            }
        ],
        "observacoes": (dados.get("observacao") or ""),
        "responsavel": (dados.get("responsavel") or ""),
    }


def enviar_op_bling_thread(app: tk.Tk, cfg: AppConfig, dados: Dict[str, Any]) -> None:
    """
    Envia para o Bling em thread para não travar a interface.
    Mostra aviso caso falhe, mas não desfaz a OP do banco.
    """
    try:
        token = (cfg.bling_token or os.getenv("BLING_TOKEN", "")).strip()
        if not token:
            app.after(0, lambda: messagebox.showwarning(
                "Bling",
                "OP salva no banco, mas NÃO foi enviada ao Bling.\n"
                "Motivo: token não configurado (cfg.bling_token ou BLING_TOKEN).",
                parent=app
            ))
            return

        url = f"{cfg.bling_base_url.rstrip('/')}/ordens-producao"
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "Authorization": f"Bearer {token}",
        }

        payload = build_bling_payload(dados)

        resp = requests.post(
            url,
            json=payload,
            headers=headers,
            timeout=int(cfg.bling_timeout or 20)
        )

        if resp.status_code in (200, 201):
            return

        txt = (resp.text or "").strip()
        if len(txt) > 800:
            txt = txt[:800] + "..."

        app.after(0, lambda: messagebox.showwarning(
            "Bling",
            "OP salva no banco, mas FALHOU o envio ao Bling.\n\n"
            f"HTTP {resp.status_code}\n{txt}",
            parent=app
        ))

    except Exception as e:
        app.after(0, lambda: messagebox.showwarning(
            "Bling",
            "OP salva no banco, mas ocorreu erro ao enviar ao Bling.\n\n"
            f"{type(e).__name__}: {e}",
            parent=app
        ))


if __name__ == "__main__":
    app = OrdemProducaoApp()
    app.mainloop()
