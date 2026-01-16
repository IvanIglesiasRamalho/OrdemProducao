# -*- coding: utf-8 -*-
"""
Relatório de Componentes (BOM) + Custos
- Usa: Ekenox.produtos, Ekenox.estrutura, Ekenox.infoProduto, Ekenox.arranjo
- Objetivo:
  1) Para cada produto final (SKU), listar componentes (BOM 1 nível),
     custo unitário do componente e custos.
  2) Se existir "arranjo" (sku, quantidade), calcula também:
     - quantidade total de cada componente = qtd_bom * qtd_produzir
     - custo total do produto (para o lote do arranjo)
  3) Gera um Excel com abas: "Detalhe" e "Resumo".

Como usar (no seu app):
    from relatorio_componentes import gerar_relatorio_componentes_excel
    caminho = gerar_relatorio_componentes_excel(cfg, base_dir=BASE_DIR)

Autor: gerado pelo ChatGPT
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple

import psycopg2

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -------------------------
# Helpers
# -------------------------

def _sku_key(sku: str) -> str:
    """
    Normaliza o SKU para chave de comparação:
    - trim, upper
    - remove o "N" final (regra usada no app)
    """
    s = (sku or "").strip().upper()
    if s.endswith("N"):
        s = s[:-1]
    return s


def _auto_col_width(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            v = str(v)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def _money(v: Any) -> float:
    try:
        if v is None:
            return 0.0
        return float(v)
    except Exception:
        return 0.0


# -------------------------
# Consulta principal
# -------------------------

SQL_RELATORIO = r"""
WITH
prod AS (
    SELECT
        p."produtoId"::bigint AS produto_id,
        p."nomeProduto" AS produto_nome,
        p."sku" AS produto_sku,
        p."tipo" AS produto_tipo,
        -- sku_key (sem N no final)
        CASE
            WHEN RIGHT(UPPER(TRIM(COALESCE(p."sku", ''))), 1) = 'N'
            THEN LEFT(UPPER(TRIM(COALESCE(p."sku", ''))), GREATEST(LENGTH(UPPER(TRIM(COALESCE(p."sku", '')))) - 1, 0))
            ELSE UPPER(TRIM(COALESCE(p."sku", '')))
        END AS sku_key
    FROM "Ekenox"."produtos" p
),
arr AS (
    SELECT
        CASE
            WHEN RIGHT(UPPER(TRIM(COALESCE(a."sku", ''))), 1) = 'N'
            THEN LEFT(UPPER(TRIM(COALESCE(a."sku", ''))), GREATEST(LENGTH(UPPER(TRIM(COALESCE(a."sku", '')))) - 1, 0))
            ELSE UPPER(TRIM(COALESCE(a."sku", '')))
        END AS sku_key,
        COALESCE(SUM(a."quantidade"), 0) AS qtd_produzir
    FROM "Ekenox"."arranjo" a
    GROUP BY 1
),
base_prod AS (
    SELECT
        pr.produto_id,
        pr.produto_nome,
        pr.produto_sku,
        pr.produto_tipo,
        COALESCE(ar.qtd_produzir, 0) AS qtd_produzir
    FROM prod pr
    LEFT JOIN arr ar
      ON ar.sku_key = pr.sku_key
),
bom AS (
    SELECT
        e."fkproduto"::bigint AS produto_id,
        e."componente"::bigint AS componente_id,
        COALESCE(e."quantidade", 0) AS qtd_bom
    FROM "Ekenox"."estrutura" e
),
comp AS (
    SELECT
        p."produtoId"::bigint AS componente_id,
        p."nomeProduto" AS componente_nome,
        p."sku" AS componente_sku,
        p."custo" AS custo_produtos
    FROM "Ekenox"."produtos" p
),
info_comp AS (
    SELECT
        i."fkProduto"::bigint AS componente_id,
        i."precoCompra" AS preco_compra,
        i."unidade" AS unidade,
        i."unidadeMedida" AS unidade_medida
    FROM "Ekenox"."infoProduto" i
)
SELECT
    bp.produto_id,
    bp.produto_nome,
    bp.produto_sku,
    bp.produto_tipo,
    bp.qtd_produzir,

    b.componente_id,
    c.componente_nome,
    c.componente_sku,

    COALESCE(ic.unidade_medida, ic.unidade, '') AS componente_unidade,

    b.qtd_bom,
    -- custo unitário do componente (prioridade: produtos.custo, senão infoProduto.precoCompra)
    COALESCE(c.custo_produtos, ic.preco_compra, 0) AS custo_unitario,

    -- custo por unidade do produto (apenas do componente)
    (COALESCE(b.qtd_bom, 0) * COALESCE(c.custo_produtos, ic.preco_compra, 0)) AS custo_total_por_unidade,

    -- qtd total do componente para o arranjo
    (COALESCE(b.qtd_bom, 0) * COALESCE(bp.qtd_produzir, 0)) AS qtd_total_componente,

    -- custo total do componente no lote do arranjo
    (COALESCE(b.qtd_bom, 0) * COALESCE(bp.qtd_produzir, 0) * COALESCE(c.custo_produtos, ic.preco_compra, 0)) AS custo_total_lote
FROM base_prod bp
JOIN bom b
  ON b.produto_id = bp.produto_id
LEFT JOIN comp c
  ON c.componente_id = b.componente_id
LEFT JOIN info_comp ic
  ON ic.componente_id = b.componente_id
ORDER BY bp.produto_nome, c.componente_nome;
"""


# -------------------------
# API pública (para seu app)
# -------------------------

def gerar_relatorio_componentes_excel(
    cfg: Any,
    base_dir: str,
    somente_skus_do_arranjo: bool = True,
    tipos_produto_final: Optional[Iterable[str]] = None,
    nome_arquivo: Optional[str] = None,
) -> str:
    """
    Gera o Excel do relatório e devolve o caminho do arquivo.

    Params
    - cfg: seu AppConfig (precisa ter db_host/db_database/db_user/db_password/db_port)
    - base_dir: pasta onde salvar (ex.: BASE_DIR)
    - somente_skus_do_arranjo: se True, remove produtos com qtd_produzir=0
    - tipos_produto_final: se informado, filtra p.tipo (case-sensitive no Postgres se não estiver com UPPER)
    - nome_arquivo: se None, gera um nome com timestamp

    Saída
    - caminho .xlsx
    """
    os.makedirs(base_dir, exist_ok=True)

    if not nome_arquivo:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_componentes_{ts}.xlsx"

    caminho_saida = os.path.join(base_dir, nome_arquivo)

    conn = psycopg2.connect(
        host=cfg.db_host,
        database=cfg.db_database,
        user=cfg.db_user,
        password=cfg.db_password,
        port=int(cfg.db_port),
        connect_timeout=10,
    )

    try:
        cur = conn.cursor()
        cur.execute(SQL_RELATORIO)
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description]
        cur.close()

        # Converte em lista de dicts para facilitar
        itens: List[Dict[str, Any]] = [dict(zip(cols, r)) for r in rows]

        # Filtragens em Python (para não complicar o SQL)
        if somente_skus_do_arranjo:
            itens = [x for x in itens if _money(x.get("qtd_produzir")) > 0]

        if tipos_produto_final:
            tipos_set = {str(t) for t in tipos_produto_final}
            itens = [x for x in itens if str(x.get("produto_tipo") or "") in tipos_set]

        if not itens:
            # Ainda assim gera um arquivo vazio com cabeçalho explicativo
            wb = Workbook()
            ws = wb.active
            ws.title = "Detalhe"
            ws["A1"] = "Nenhum dado encontrado para o relatório (verifique filtros/arranjo/estrutura)."
            wb.save(caminho_saida)
            return caminho_saida

        # Agrupa por produto
        por_produto: Dict[int, List[Dict[str, Any]]] = {}
        for x in itens:
            pid = int(x["produto_id"])
            por_produto.setdefault(pid, []).append(x)

        # Monta workbook
        wb = Workbook()

        # --- Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        money_fmt = '"R$" #,##0.00'
        qty_fmt = '#,##0.0000'

        # ==========================
        # Aba DETALHE
        # ==========================
        ws = wb.active
        ws.title = "Detalhe"

        headers = [
            "Produto ID",
            "Produto",
            "SKU",
            "Tipo",
            "Qtd Produzir (Arranjo)",
            "Componente ID",
            "Componente",
            "SKU Componente",
            "Unidade",
            "Qtd por Unidade (BOM)",
            "Custo Unitário Comp",
            "Custo Comp por Unidade",
            "Qtd Total Comp (lote)",
            "Custo Total Comp (lote)",
        ]
        ws.append(headers)

        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        row_idx = 2
        for pid, linhas in por_produto.items():
            for x in linhas:
                ws.append([
                    int(x.get("produto_id") or 0),
                    x.get("produto_nome") or "",
                    x.get("produto_sku") or "",
                    x.get("produto_tipo") or "",
                    _money(x.get("qtd_produzir")),
                    int(x.get("componente_id") or 0),
                    x.get("componente_nome") or "",
                    x.get("componente_sku") or "",
                    x.get("componente_unidade") or "",
                    _money(x.get("qtd_bom")),
                    _money(x.get("custo_unitario")),
                    _money(x.get("custo_total_por_unidade")),
                    _money(x.get("qtd_total_componente")),
                    _money(x.get("custo_total_lote")),
                ])

                # formatação linha
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.border = border
                    if c in (5, 10, 13):  # quantidades
                        cell.number_format = qty_fmt
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif c in (11, 12, 14):  # dinheiro
                        cell.number_format = money_fmt
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif c in (1, 6):  # ids
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                row_idx += 1

            # linha subtotal por produto
            custo_unit_prod = sum(_money(a.get("custo_total_por_unidade")) for a in linhas)
            custo_total_lote = sum(_money(a.get("custo_total_lote")) for a in linhas)

            ws.append([
                "", f"TOTAL PRODUTO {pid}", "", "", "",
                "", "", "", "",
                "",
                "",
                custo_unit_prod,
                "",
                custo_total_lote,
            ])
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = border
                cell.font = Font(bold=True)
                if c in (12, 14):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            row_idx += 1

        ws.freeze_panes = "A2"
        _auto_col_width(ws)

        # ==========================
        # Aba RESUMO
        # ==========================
        ws2 = wb.create_sheet("Resumo")

        headers2 = [
            "Produto ID",
            "Produto",
            "SKU",
            "Tipo",
            "Qtd Produzir (Arranjo)",
            "Custo Total por Unidade",
            "Custo Total do Lote (Arranjo)",
        ]
        ws2.append(headers2)
        for col_idx, _ in enumerate(headers2, start=1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        r = 2
        total_geral = 0.0
        for pid, linhas in por_produto.items():
            first = linhas[0]
            qtd_prod = _money(first.get("qtd_produzir"))
            custo_unit_prod = sum(_money(a.get("custo_total_por_unidade")) for a in linhas)
            custo_total_lote = sum(_money(a.get("custo_total_lote")) for a in linhas)
            total_geral += custo_total_lote

            ws2.append([
                int(first.get("produto_id") or 0),
                first.get("produto_nome") or "",
                first.get("produto_sku") or "",
                first.get("produto_tipo") or "",
                qtd_prod,
                custo_unit_prod,
                custo_total_lote,
            ])

            for c in range(1, len(headers2) + 1):
                cell = ws2.cell(row=r, column=c)
                cell.border = border
                if c == 5:
                    cell.number_format = qty_fmt
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c in (6, 7):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            r += 1

        # linha total geral
        ws2.append(["", "TOTAL GERAL", "", "", "", "", total_geral])
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.border = border
            cell.font = Font(bold=True)
            if c == 7:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws2.freeze_panes = "A2"
        _auto_col_width(ws2)

        wb.save(caminho_saida)
        return caminho_saida

    finally:
        try:
            conn.close()
        except Exception:
            pass
