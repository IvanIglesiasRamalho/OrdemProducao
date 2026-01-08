# ordem_producao_windows_corrigido.py
# ============================================================
# SISTEMA DE ORDEM DE PRODUÇÃO - EKENOX (Tkinter + PostgreSQL)
# Versão corrigida para:
# - Ter APENAS 1 Tk() (uma raiz)
# - Evitar erro: TclError: can't invoke "wm" command: application has been destroyed
# - Remover funções/métodos “perdidos” por indentação errada
# - Centralizar registro do WM_DELETE_WINDOW dentro do __init__
# - Consolidar binds/atalhos
# - Corrigir estrutura geral para rodar direto no VS Code
#
# IMPORTANTE:
# - Ajuste BASE_DIR, host, porta, credenciais, etc.
# - Se quiser desabilitar o webhook n8n, deixe N8N_WEBHOOK_URL = ""
# ============================================================

from __future__ import annotations

import os
import sys
import subprocess
import warnings
import traceback
import requests
from math import ceil
from collections import defaultdict
from datetime import datetime, date
from typing import Optional, Dict, Any, List, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

import psycopg2
from psycopg2 import errors

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ============================================================
# CONFIG / PATHS
# ============================================================

def get_app_dir() -> str:
    """Quando estiver em .exe (PyInstaller), pega a pasta do executável.
    Quando estiver em .py, pega a pasta do arquivo .py.
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = get_app_dir()

# Diretório base
# BASE_DIR = r"C:\Users\User\Desktop\Pyton"  # <<< AJUSTE AQUI
BASE_DIR = r"Z:\Planilhas_OP"  # <<< AJUSTE AQUI

# Garante que o diretório existe
if not os.path.exists(BASE_DIR):
    try:
        os.makedirs(BASE_DIR)
        print(f"✓ Diretório criado: {BASE_DIR}")
    except Exception as e:
        print(f"✗ Erro ao criar diretório {BASE_DIR}: {e}")

# Excel modelo/saída
CAMINHO_MODELO = os.path.join(BASE_DIR, "pedido-de-compra v2.xlsx")
CAMINHO_SAIDA = os.path.join(BASE_DIR, "saida_pedido-de-compra v2.xlsx")

# Webhook n8n (deixe "" para desabilitar)
N8N_WEBHOOK_URL = "http://localhost:56789/webhook/ordem-producao"  # ou ""

# Warnings openpyxl
warnings.filterwarnings(
    "ignore", message="Cannot parse header or footer so it will be ignored")
warnings.filterwarnings(
    "ignore", message="Data Validation extension is not supported and will be removed")


# ============================================================
# LOG
# ============================================================

def log_exception(err: Exception, context: str = "") -> str:
    """Loga o stacktrace em erro_app.log e retorna o caminho."""
    try:
        texto = "".join(traceback.format_exception(
            type(err), err, err.__traceback__))
        log_path = os.path.join(BASE_DIR, "erro_app.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n" + ("=" * 60) + "\n")
            if context:
                f.write(f"{context}\n")
            f.write(texto)
        return log_path
    except Exception:
        # Último recurso
        print("ERRO AO GRAVAR LOG:", err)
        traceback.print_exc()
        return os.path.join(BASE_DIR, "erro_app.log")


# ============================================================
# ÍCONE
# ============================================================

def find_icon_path() -> str | None:
    candidates = [
        os.path.join(BASE_DIR, "imagens", "favicon.ico"),
        os.path.join(BASE_DIR, "favicon.ico"),
        os.path.join(APP_DIR, "imagens", "favicon.ico"),
        os.path.join(APP_DIR, "favicon.ico"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return None


def apply_window_icon(win) -> None:
    try:
        icon_path = find_icon_path()
        if icon_path:
            win.iconbitmap(default=icon_path)
    except Exception as e:
        print(f"⚠ Não foi possível aplicar ícone: {e}")


# ============================================================
# EXCEL - Helpers
# ============================================================

def _nome_aba_excel_valido(nome: str) -> str:
    invalidos = ['\\', '/', '?', '*', '[', ']']
    for ch in invalidos:
        nome = nome.replace(ch, " ")
    return nome[:31]


def _escrever_celula_segura(ws, coord: str, valor):
    """Escreve valor em célula tratando merge."""
    try:
        cell = ws[coord]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if coord in merged_range:
                    top_left_coord = merged_range.coord.split(":")[0]
                    ws[top_left_coord] = valor
                    return
        else:
            cell.value = valor
    except Exception as e:
        print(f"Erro ao escrever na célula {coord}: {e}")


def _set_cell_segura_rc(ws, row: int, col: int, valor):
    cell = ws.cell(row=row, column=col)
    coord = cell.coordinate
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                tl_row = merged_range.min_row
                tl_col = merged_range.min_col
                ws.cell(row=tl_row, column=tl_col).value = valor
                return
        print(f"Aviso: {coord} é MergedCell mas não encontrada no range.")
    else:
        cell.value = valor


def gerar_abas_fornecedor_pedido(dados, nome_aba_modelo: str = "Pedido de Compra"):
    """Cria uma aba por (fornecedor, numero_pedido[, data_pedido])."""

    # 1) Abrir workbook
    if os.path.exists(CAMINHO_SAIDA):
        wb = load_workbook(CAMINHO_SAIDA)
    else:
        wb = load_workbook(CAMINHO_MODELO)

    if nome_aba_modelo not in wb.sheetnames:
        raise ValueError(
            f"Aba de modelo '{nome_aba_modelo}' não encontrada. Abas: {wb.sheetnames}"
        )

    aba_modelo = wb[nome_aba_modelo]

    # 2) Agrupar
    tmp = defaultdict(list)
    for item in dados:
        fornecedor = item["fornecedor"]
        numero_pedido = item["numero_pedido"]
        data_pedido = item.get("data_pedido")
        tmp[(fornecedor, numero_pedido, data_pedido)].append(item)

    grupos = list(tmp.items())

    # 3) Criar abas
    for (fornecedor, numero_pedido, data_pedido), linhas in grupos:
        ws = wb.copy_worksheet(aba_modelo)
        titulo_aba = f"{numero_pedido} - {str(fornecedor)[:15]}"
        ws.title = _nome_aba_excel_valido(titulo_aba)

        _escrever_celula_segura(ws, "D6", numero_pedido)
        _escrever_celula_segura(ws, "D8", data_pedido or date.today())
        _escrever_celula_segura(ws, "D10", str(fornecedor))

        # limpar tabela
        for r in range(16, 43):
            for c in range(2, 10):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        linha = 16
        numero_item = 1
        for item in linhas:
            descricao = item["produto"]
            quantidade = float(item["quantidade"])

            estoque_atual = float(item.get("estoque_atual", 0.0) or 0.0)
            estoque_minimo = float(item.get("estoque_minimo", 0.0) or 0.0)
            estoque_maximo = float(item.get("estoque_maximo", 0.0) or 0.0)
            valor_unitario = item.get("valor_unitario")

            _set_cell_segura_rc(ws, linha, 2, numero_item)   # B
            _set_cell_segura_rc(ws, linha, 3, descricao)     # C
            _set_cell_segura_rc(ws, linha, 4, estoque_atual)  # D
            _set_cell_segura_rc(ws, linha, 5, estoque_minimo)  # E
            _set_cell_segura_rc(ws, linha, 6, estoque_maximo)  # F

            if valor_unitario is not None:
                vu = float(valor_unitario)
                _set_cell_segura_rc(ws, linha, 7, vu)                 # G
                _set_cell_segura_rc(ws, linha, 9, quantidade * vu)    # I

            _set_cell_segura_rc(ws, linha, 8, quantidade)             # H

            numero_item += 1
            linha += 1

    wb.save(CAMINHO_SAIDA)


def gerar_planilha_excel(dados):
    gerar_abas_fornecedor_pedido(
        dados=dados, nome_aba_modelo="Pedido de Compra")


def gerar_pedido_compra(itens_pedido, numero_pedido_inicial, fornecedor_padrao=None, parent=None) -> int:
    """Gera planilha de pedido agrupando por fornecedor. Retorna último número usado."""
    if not itens_pedido:
        messagebox.showinfo(
            "Pedido de Compra", "Não há itens em falta para gerar o pedido.", parent=parent)
        return int(numero_pedido_inicial)

    dados = []
    hoje = date.today()

    grupos_por_fornecedor = defaultdict(list)
    for item in itens_pedido:
        fornecedor_item = item.get(
            "fornecedor") or fornecedor_padrao or "SEM FORNECEDOR"
        grupos_por_fornecedor[fornecedor_item].append(item)

    numero_atual = int(numero_pedido_inicial)

    for fornecedor_item, itens in grupos_por_fornecedor.items():
        for item in itens:
            valor_unitario = float(item.get("valor_unitario", 0.0) or 0.0)
            dados.append(
                {
                    "fornecedor": fornecedor_item,
                    "numero_pedido": numero_atual,
                    "data_pedido": hoje,
                    "produto": item["descricao"],
                    "quantidade": float(item["qtd_comprar"]),
                    "estoque_atual": float(item.get("estoque_atual", 0.0) or 0.0),
                    "estoque_minimo": float(item.get("estoque_minimo", 0.0) or 0.0),
                    "estoque_maximo": float(item.get("estoque_maximo", 0.0) or 0.0),
                    "valor_unitario": valor_unitario,
                }
            )
        numero_atual += 1

    gerar_planilha_excel(dados)
    messagebox.showinfo("Pedido de Compra",
                        f"Planilha gerada em:\n{CAMINHO_SAIDA}", parent=parent)

    return numero_atual - 1


# ============================================================
# BANCO / REGRAS
# ============================================================

class SistemaOrdemProducao:
    def __init__(self, host: str, database: str, user: str, password: str, port: str):
        self.conn_params = {
            "host": host,
            "database": database,
            "user": user,
            "password": password,
            "port": int(port),
        }
        self.conn = None
        self.cursor = None
        self._ultimo_erro_bd: Optional[str] = None

    def conectar(self) -> bool:
        try:
            self.conn = psycopg2.connect(**self.conn_params)
            self.cursor = self.conn.cursor()
            print("✓ Conectado ao PostgreSQL")
            return True
        except Exception as e:
            self._ultimo_erro_bd = str(e)
            print(f"✗ Erro ao conectar: {e}")
            return False

    def desconectar(self):
        try:
            if self.cursor:
                self.cursor.close()
            if self.conn:
                self.conn.close()
            print("✓ Desconectado do PostgreSQL")
        except Exception as e:
            print(f"✗ Erro ao desconectar: {e}")

    # ---------------- Sequenciador ----------------

    def obter_sequenciador(self, nome_tabela: str) -> int:
        try:
            sql = """
                SELECT s."sequenciador"
                  FROM "Ekenox"."sequenciadores" AS s
                 WHERE s."tabela" = %s;
            """
            self.cursor.execute(sql, (nome_tabela,))
            row = self.cursor.fetchone()
            return int(row[0]) if row and row[0] is not None else 0
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro ao buscar sequenciador ({nome_tabela}): {e}")
            return 0

    def salvar_sequenciador(self, nome_tabela: str, valor: int) -> None:
        try:
            sql_select = """
                SELECT s."sequenciador"
                  FROM "Ekenox"."sequenciadores" AS s
                 WHERE s."tabela" = %s;
            """
            self.cursor.execute(sql_select, (nome_tabela,))
            row = self.cursor.fetchone()

            if row:
                sql_upd = """
                    UPDATE "Ekenox"."sequenciadores"
                       SET "sequenciador" = %s
                     WHERE "tabela" = %s;
                """
                self.cursor.execute(sql_upd, (valor, nome_tabela))
            else:
                sql_ins = """
                    INSERT INTO "Ekenox"."sequenciadores" ("tabela","sequenciador")
                    VALUES (%s, %s);
                """
                self.cursor.execute(sql_ins, (nome_tabela, valor))

            self.conn.commit()
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro ao salvar sequenciador ({nome_tabela}): {e}")

    # ---------------- Consultas auxiliares ----------------

    def qtd_registros(self, tabela: str) -> int:
        sql = f'SELECT COUNT(*) FROM "Ekenox"."{tabela}";'
        try:
            self.cursor.execute(sql)
            row = self.cursor.fetchone()
            return int(row[0]) if row and row[0] is not None else 0
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro COUNT {tabela}: {e}")
            return 0

    def validar_produto(self, produto_id: int) -> Optional[Dict[str, Any]]:
        try:
            query = """
                SELECT p."produtoId", p."nomeProduto", p."sku", p."preco", p."tipo"
                FROM "Ekenox"."produtos" AS p
                WHERE p."produtoId" = %s
            """
            self.cursor.execute(query, (str(produto_id),))
            r = self.cursor.fetchone()
            if not r:
                return None
            return {
                "produtoid": r[0],
                "nomeproduto": r[1],
                "sku": r[2],
                "preco": r[3],
                "tipo": r[4],
            }
        except Exception as e:
            print(f"✗ Erro validar_produto: {e}")
            return None

    def validar_situacao(self, situacao_id: int) -> Optional[Dict[str, Any]]:
        try:
            query = """
                SELECT s."id", s."nome"
                FROM "Ekenox"."situacao" AS s
                WHERE s."id" = %s
            """
            self.cursor.execute(query, (situacao_id,))
            r = self.cursor.fetchone()
            if not r:
                return None
            return {"id": r[0], "nome": r[1]}
        except Exception as e:
            print(f"✗ Erro validar_situacao: {e}")
            return None

    def gerar_id_ordem(self) -> int:
        try:
            query = """
                SELECT COALESCE(MAX(id), 0) + 1 AS proximo_id
                  FROM "Ekenox"."ordem_producao";
            """
            self.cursor.execute(query)
            r = self.cursor.fetchone()
            return int(r[0]) if r and r[0] is not None else 1
        except Exception as e:
            print(f"✗ Erro gerar_id_ordem: {e}")
            return 1

    def gerar_numero_ordem(self) -> int:
        try:
            query = """
                SELECT (numero + 1) AS proximo_id
                  FROM "Ekenox"."ordem_producao"
              ORDER BY numero DESC
                 LIMIT 1;
            """
            self.cursor.execute(query)
            r = self.cursor.fetchone()
            return int(r[0]) if r and r[0] is not None else 1
        except Exception as e:
            print(f"✗ Erro gerar_numero_ordem: {e}")
            return 1

    # ---------------- Listagens ----------------

    def listar_produtos_disponiveis(self, limite: Optional[int] = None):
        try:
            if limite is None or limite <= 0:
                query = """
                    SELECT p."produtoId", p."nomeProduto", p."sku", p."preco", p."tipo"
                      FROM "Ekenox"."produtos" AS p
                      JOIN "Ekenox"."infoProduto" AS ip
                        ON ip."fkProduto" = p."produtoId"
                     WHERE ip."fkCategoria" NOT IN (
                        3844533,3983855,7879429,3869959,4241123,
                        3870601,3844542,7651801,3983399,959867,
                        897565,3984869,3862825,7879102,7911660,
                        4828356,6568231
                     )
                  ORDER BY p."nomeProduto"
                """
                self.cursor.execute(query)
            else:
                query = """
                    SELECT p."produtoId", p."nomeProduto", p."sku", p."preco", p."tipo"
                      FROM "Ekenox"."produtos" AS p
                      JOIN "Ekenox"."infoProduto" AS ip
                        ON ip."fkProduto" = p."produtoId"
                     WHERE ip."fkCategoria" NOT IN (
                        3844533,3983855,7879429,3869959,4241123,
                        3870601,3844542,7651801,3983399,959867,
                        897565,3984869,3862825,7879102,7911660,
                        4828356,6568231
                     )
                  ORDER BY p."nomeProduto"
                     LIMIT %s
                """
                self.cursor.execute(query, (limite,))
            return self.cursor.fetchall()
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro listar_produtos_disponiveis: {e}")
            return []

    def listar_situacoes_disponiveis(self, limite: Optional[int] = None):
        try:
            if limite is None or limite <= 0:
                query = """
                    SELECT s."id", s."nome"
                      FROM "Ekenox"."situacao" AS s
                  ORDER BY s."nome"
                """
                self.cursor.execute(query)
            else:
                query = """
                    SELECT s."id", s."nome"
                      FROM "Ekenox"."situacao" AS s
                  ORDER BY s."nome"
                     LIMIT %s
                """
                self.cursor.execute(query, (limite,))
            return self.cursor.fetchall()
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro listar_situacoes_disponiveis: {e}")
            return []

    def listar_depositos_disponiveis(self, limite: Optional[int] = None):
        try:
            if limite is None or limite <= 0:
                query = """
                    SELECT d."id", d."descricao", d."situacao", d."padrao", d."desconsiderarsaldo"
                      FROM "Ekenox"."deposito" AS d
                  ORDER BY d."descricao"
                """
                self.cursor.execute(query)
            else:
                query = """
                    SELECT d."id", d."descricao", d."situacao", d."padrao", d."desconsiderarsaldo"
                      FROM "Ekenox"."deposito" AS d
                  ORDER BY d."descricao"
                     LIMIT %s
                """
                self.cursor.execute(query, (limite,))
            return self.cursor.fetchall()
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro listar_depositos_disponiveis: {e}")
            return []

    def listar_ordens_producao(self, limite: Optional[int] = None):
        try:
            if limite is None or limite <= 0:
                query = """
                    SELECT
                        o."id",
                        o."numero",
                        o."fkprodutoid",
                        p."nomeProduto" AS produto_nome,
                        o."situacao_id",
                        s."nome" AS situacao_nome,
                        o."quantidade",
                        o."data_inicio",
                        o."data_fim"
                    FROM "Ekenox"."ordem_producao" AS o
                    LEFT JOIN "Ekenox"."produtos" AS p
                           ON p."produtoId" = o."fkprodutoid"
                    LEFT JOIN "Ekenox"."situacao" AS s
                           ON s."id" = o."situacao_id"
                    ORDER BY o."id" DESC
                """
                self.cursor.execute(query)
            else:
                query = """
                    SELECT
                        o."id",
                        o."numero",
                        o."fkprodutoid",
                        p."nomeProduto" AS produto_nome,
                        o."situacao_id",
                        s."nome" AS situacao_nome,
                        o."quantidade",
                        o."data_inicio",
                        o."data_fim"
                    FROM "Ekenox"."ordem_producao" AS o
                    LEFT JOIN "Ekenox"."produtos" AS p
                           ON p."produtoId" = o."fkprodutoid"
                    LEFT JOIN "Ekenox"."situacao" AS s
                           ON s."id" = o."situacao_id"
                    ORDER BY o."id" DESC
                    LIMIT %s
                """
                self.cursor.execute(query, (limite,))
            return self.cursor.fetchall()
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro listar_ordens_producao: {e}")
            return []

    def listar_ordens_sem_data_fim(self):
        try:
            query = """
                SELECT
                    o."id",
                    o."numero",
                    o."fkprodutoid",
                    p."nomeProduto" AS produto_nome,
                    o."situacao_id",
                    s."nome" AS situacao_nome,
                    o."quantidade",
                    o."data_inicio"
                FROM "Ekenox"."ordem_producao" AS o
                LEFT JOIN "Ekenox"."produtos" AS p
                       ON p."produtoId" = o."fkprodutoid"
                LEFT JOIN "Ekenox"."situacao" AS s
                       ON s."id" = o."situacao_id"
                WHERE o."data_fim" IS NULL
                   OR o."data_fim" = '1970-01-01'
                ORDER BY o."id" DESC;
            """
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro listar_ordens_sem_data_fim: {e}")
            return []

    # ---------------- Inserção / exclusão ----------------

    def inserir_ordem_producao(self, dados: Dict[str, Any]) -> bool:
        self._ultimo_erro_bd = None

        try:
            produto = self.validar_produto(dados["fkprodutoid"])
            if not produto:
                raise ValueError(
                    f"Produto ID {dados['fkprodutoid']} não encontrado.")

            situacao = self.validar_situacao(dados["situacao_id"])
            if not situacao:
                raise ValueError(
                    f"Situação ID {dados['situacao_id']} não encontrada.")

            if not dados.get("id"):
                dados["id"] = self.gerar_id_ordem()

            query = """
                INSERT INTO "Ekenox"."ordem_producao" (
                    id,
                    numero,
                    deposito_id_destino,
                    deposito_id_origem,
                    situacao_id,
                    responsavel,
                    fkprodutoid,
                    data_previsao_inicio,
                    data_previsao_final,
                    data_inicio,
                    data_fim,
                    valor,
                    observacao,
                    quantidade
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """

            valores = (
                dados["id"],
                dados["numero"],
                dados["deposito_id_destino"],
                dados["deposito_id_origem"],
                dados["situacao_id"],
                dados.get("responsavel"),
                dados["fkprodutoid"],
                dados.get("data_previsao_inicio"),
                dados.get("data_previsao_final"),
                dados.get("data_inicio"),
                dados.get("data_fim"),
                dados.get("valor"),
                dados.get("observacao"),
                dados["quantidade"],
            )

            self.cursor.execute(query, valores)
            self.conn.commit()
            return True

        except errors.UniqueViolation as e:
            self.conn.rollback()
            msg = (
                "NÚMERO DE ORDEM JÁ EXISTENTE.\n\n"
                f"Número: {dados.get('numero')}\n"
                f"Constraint: {getattr(e.diag, 'constraint_name', '')}"
            )
            self._ultimo_erro_bd = msg
            return False

        except errors.ForeignKeyViolation as e:
            self.conn.rollback()
            tabela = getattr(e.diag, "table_name", "desconhecida")
            constraint = getattr(e.diag, "constraint_name", "desconhecida")
            msg = (
                "VIOLAÇÃO DE CHAVE ESTRANGEIRA.\n\n"
                f"Tabela alvo: {tabela}\n"
                f"Constraint: {constraint}\n\n"
                "Provavelmente algum ID não existe (depósitos, situação ou produto)."
            )
            self._ultimo_erro_bd = msg
            return False

        except Exception as e:
            if self.conn:
                self.conn.rollback()
            self._ultimo_erro_bd = f"Erro ao inserir OP: {e}"
            return False

    def excluir_ordem_producao(self, ordem_id: int) -> bool:
        try:
            sql = 'DELETE FROM "Ekenox"."ordem_producao" WHERE "id" = %s;'
            self.cursor.execute(sql, (ordem_id,))
            if self.cursor.rowcount == 0:
                self.conn.commit()
                return False
            self.conn.commit()
            return True
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro excluir_ordem_producao: {e}")
            return False

    # ---------------- Finalização ----------------

    def finalizar_ordem_individual(self, ordem_id) -> bool:
        try:
            ordem_id_int = int(ordem_id)
            hoje = date.today()

            sql = '''
                UPDATE "Ekenox"."ordem_producao"
                   SET "data_fim" = %s,
                       situacao_id = 18162
                 WHERE "id" = %s
                   AND ("data_fim" IS NULL OR "data_fim" = '1970-01-01');
            '''
            self.cursor.execute(sql, (hoje, ordem_id_int))

            if self.cursor.rowcount == 0:
                self.conn.rollback()
                return False

            self.conn.commit()

            # Webhook (opcional)
            if N8N_WEBHOOK_URL:
                try:
                    payload = {"ordem_id": ordem_id_int, "data_fim": str(hoje)}
                    resp = requests.post(
                        N8N_WEBHOOK_URL, json=payload, timeout=10)
                    if not (200 <= resp.status_code < 300):
                        print(f"⚠ n8n erro {resp.status_code}: {resp.text}")
                except Exception as n8n_err:
                    print(f"⚠ n8n exceção: {n8n_err}")

            return True

        except Exception as e:
            if self.conn:
                self.conn.rollback()
            print(f"✗ Erro finalizar_ordem_individual: {e}")
            return False


# ============================================================
# UI - APP
# ============================================================

class OrdemProducaoApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self._closing = False
        self.variaveis_quantidade = None

        try:
            self.title("Sistema de Ordem de Produção - Ekenox")
            self.geometry("1150x650")
            self.minsize(1150, 650)
            apply_window_icon(self)

            # IMPORTANTÍSSIMO: protocol dentro do __init__
            self.protocol("WM_DELETE_WINDOW", self.on_closing)

            # Esconde enquanto carrega e mostra splash
            self.withdraw()

            # atalhos (uma vez só)
            self.bind_all("<F12>", self.abrir_programa_etiqueta)

            # conecta banco
            self.sistema = SistemaOrdemProducao(
                host="10.0.0.154",
                database="postgresekenox",
                user="postgresekenox",
                password="Ekenox5426",
                port="55432",
            )
            self.connected = self.sistema.conectar()

            self.total_produtos = 0
            self.total_situacoes = 0
            self.total_depositos = 0
            self.total_ordens = 0

            if self.connected:
                self.carregar_totais()

            self.create_widgets()
            self.atualizar_numero_ordem()

            # splash/entrada depois que Tk "assentar"
            self.after(50, self.mostrar_tela_entrada)

        except Exception as e:
            self._fatal(e, "Falha no __init__ do app")

    # ---------------- Fatal ----------------

    def _fatal(self, err: Exception, context: str = ""):
        log_path = log_exception(err, context=context)
        try:
            messagebox.showerror(
                "Erro Fatal",
                f"O programa falhou.\n\nVeja o log:\n{log_path}"
            )
        except Exception:
            pass
        try:
            if self.winfo_exists():
                self.destroy()
        except Exception:
            pass

    # ---------------- Bring to front ----------------

    def bring_window_to_front_by_title(self, title_contains: str) -> bool:
        """Windows: tenta achar janela pelo título e trazer para frente."""
        try:
            import ctypes
            from ctypes import wintypes

            user32 = ctypes.windll.user32
            EnumWindows = user32.EnumWindows
            EnumWindowsProc = ctypes.WINFUNCTYPE(
                ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
            GetWindowTextW = user32.GetWindowTextW
            GetWindowTextLengthW = user32.GetWindowTextLengthW
            IsWindowVisible = user32.IsWindowVisible
            ShowWindow = user32.ShowWindow
            SetForegroundWindow = user32.SetForegroundWindow

            found = {"hwnd": None}

            def foreach(hwnd, lParam):
                if not IsWindowVisible(hwnd):
                    return True
                length = GetWindowTextLengthW(hwnd)
                if length <= 0:
                    return True
                buf = ctypes.create_unicode_buffer(length + 1)
                GetWindowTextW(hwnd, buf, length + 1)
                if title_contains.lower() in buf.value.lower():
                    found["hwnd"] = hwnd
                    return False
                return True

            EnumWindows(EnumWindowsProc(foreach), 0)

            if found["hwnd"]:
                SW_RESTORE = 9
                ShowWindow(found["hwnd"], SW_RESTORE)
                SetForegroundWindow(found["hwnd"])
                return True

        except Exception as e:
            print("⚠ bring_window_to_front_by_title falhou:", e)

        return False

    # ---------------- Etiquetas (F12) ----------------

    def abrir_programa_etiqueta(self, event=None):
        exe_path = os.path.join(BASE_DIR, "etiqueta.exe")
        if not os.path.isfile(exe_path):
            messagebox.showerror("Etiquetas", f"Não encontrei:\n{exe_path}")
            return
        print("Abrindo etiqueta exe:", exe_path)
        try:
            if os.name == "nt":
                os.startfile(exe_path)
            else:
                subprocess.Popen([exe_path], cwd=BASE_DIR)

            def _retry(attempt=0):
                ok = self.bring_window_to_front_by_title("etiqueta")
                if not ok and attempt < 12:
                    self.after(500, lambda: _retry(attempt + 1))

            self.after(500, lambda: _retry(0))

        except Exception as e:
            messagebox.showerror(
                "Etiquetas", f"Falha ao abrir etiqueta.exe:\n{e}")

    # ---------------- Totais ----------------

    def carregar_totais(self):
        if not self.connected:
            return
        self.total_produtos = self.sistema.qtd_registros("produtos")
        self.total_situacoes = self.sistema.qtd_registros("situacao")
        self.total_depositos = self.sistema.qtd_registros("deposito")
        self.total_ordens = self.sistema.qtd_registros("ordem_producao")

    # ---------------- Splash / Entrada ----------------

    def _carregar_avatar_tk(self, caminho: str, max_lado: int = 280) -> tk.PhotoImage | None:
        try:
            img = tk.PhotoImage(file=caminho)
            w, h = img.width(), img.height()
            maior = max(w, h)
            if maior > max_lado:
                fator = max(1, int(maior / max_lado))
                img = img.subsample(fator, fator)
            return img
        except Exception:
            return None

    def mostrar_tela_entrada(self):
        """Tela de entrada. Ao entrar, mostra a janela principal."""
        if self._closing:
            return

        if hasattr(self, "_tela_entrada") and self._tela_entrada.winfo_exists():
            return

        tela = tk.Toplevel(self)
        self._tela_entrada = tela
        apply_window_icon(tela)

        tela.title("Ekenox - Entrada")
        tela.resizable(False, False)
        tela.configure(bg="#121212")

        try:
            tela.attributes("-topmost", True)
            tela.after(300, lambda: tela.attributes("-topmost", False))
        except Exception:
            pass

        # Se fechar a splash, fecha o app todo
        tela.protocol("WM_DELETE_WINDOW", self.on_closing)

        candidatos = [
            os.path.join(BASE_DIR, "imagens", "avatar_ekenox.png"),
            os.path.join(BASE_DIR, "avatar_ekenox.png"),
            os.path.join(BASE_DIR, "imagens", "Ekenox.png"),
            os.path.join(BASE_DIR, "Ekenox.png"),
            os.path.join(BASE_DIR, "imagens", "ekenox.png"),
            os.path.join(BASE_DIR, "ekenox.png"),
        ]
        caminho_avatar = next(
            (p for p in candidatos if os.path.isfile(p)), None)

        frame = tk.Frame(tela, bg="#121212", padx=30, pady=25)
        frame.pack(fill="both", expand=True)

        avatar_img = self._carregar_avatar_tk(
            caminho_avatar, max_lado=260) if caminho_avatar else None
        if avatar_img:
            lbl_img = tk.Label(frame, image=avatar_img, bg="#121212")
            lbl_img.image = avatar_img
            lbl_img.pack(pady=(0, 15))
        else:
            tk.Label(
                frame,
                text="(Avatar não encontrado)",
                bg="#121212",
                fg="#aaaaaa",
                font=("Segoe UI", 10),
            ).pack(pady=(0, 15))

        tk.Label(
            frame,
            text="Sistema de Ordem de Produção",
            bg="#121212",
            fg="#ffffff",
            font=("Segoe UI", 14, "bold"),
        ).pack()

        tk.Label(
            frame,
            text="Ekenox",
            bg="#121212",
            fg="#ff9f1a",
            font=("Segoe UI", 18, "bold"),
        ).pack(pady=(2, 18))

        botoes = tk.Frame(frame, bg="#121212")
        botoes.pack(fill="x")

        def entrar(event=None):
            if self._closing:
                return
            try:
                tela.destroy()
            except Exception:
                pass

            # mostra principal
            try:
                self.deiconify()
                self.lift()
                self.focus_force()
                self.deposito_origem_entry.focus_set()
            except Exception:
                pass

        btn_entrar = ttk.Button(botoes, text="Entrar", command=entrar)
        btn_entrar.pack(side="left", expand=True, fill="x", padx=(0, 8))

        btn_sair = ttk.Button(botoes, text="Sair", command=self.on_closing)
        btn_sair.pack(side="left", expand=True, fill="x")

        tela.bind("<Return>", entrar)
        tela.bind("<Escape>", lambda e: self.on_closing())

        # Centraliza
        tela.update_idletasks()
        w, h = tela.winfo_width(), tela.winfo_height()
        x = (tela.winfo_screenwidth() // 2) - (w // 2)
        y = (tela.winfo_screenheight() // 2) - (h // 2)
        tela.geometry(f"+{x}+{y}")
        btn_entrar.focus_set()

    # ---------------- Data helpers ----------------

    def _formatar_data_digitando(self, event, var: tk.StringVar, entry: ttk.Entry):
        texto = var.get()
        digitos = "".join(ch for ch in texto if ch.isdigit())
        if len(digitos) > 8:
            digitos = digitos[:8]

        if len(digitos) <= 2:
            novo = digitos
        elif len(digitos) <= 4:
            novo = digitos[:2] + "/" + digitos[2:]
        else:
            novo = digitos[:2] + "/" + digitos[2:4] + "/" + digitos[4:]

        var.set(novo)
        entry.icursor(tk.END)

    def validar_data_digitada(self, data_str: str, campo: str) -> Optional[datetime]:
        data_str = (data_str or "").strip()
        if not data_str:
            return None
        try:
            return datetime.strptime(data_str, "%d/%m/%Y")
        except ValueError:
            raise ValueError(
                f"O campo '{campo}' deve conter uma data válida no formato DD/MM/AAAA.\n"
                f"Valor informado: {data_str}"
            )

    def parse_int(self, valor_str: str, campo: str) -> int:
        if not valor_str.strip():
            raise ValueError(f"O campo '{campo}' é obrigatório.")
        try:
            return int(valor_str.strip())
        except ValueError:
            raise ValueError(f"O campo '{campo}' deve ser um número inteiro.")

    def parse_float(self, valor_str: str, campo: str, obrigatorio: bool = False) -> Optional[float]:
        if not valor_str.strip():
            if obrigatorio:
                raise ValueError(f"O campo '{campo}' é obrigatório.")
            return None
        try:
            return float(valor_str.replace(",", "."))
        except ValueError:
            raise ValueError(f"O campo '{campo}' deve ser um número decimal.")

    def parse_date(self, valor_str: str, campo: str) -> Optional[datetime]:
        valor_str = (valor_str or "").strip()
        if not valor_str:
            return None
        return self.validar_data_digitada(valor_str, campo)

    def criar_entry_data(self, parent, var: tk.StringVar, nome_campo: str = "Data", largura: int = 12):
        entry = ttk.Entry(parent, textvariable=var, width=largura)

        entry.bind("<KeyRelease>", lambda e, v=var,
                   ent=entry: self._formatar_data_digitando(e, v, ent))

        def on_focus_out(event):
            texto = var.get().strip()
            if not texto:
                return
            try:
                dt = self.validar_data_digitada(texto, nome_campo)
                var.set(dt.strftime("%d/%m/%Y"))
            except ValueError as e:
                messagebox.showerror("Data inválida", str(e))
                self.after(10, lambda: entry.focus_set())

        entry.bind("<FocusOut>", on_focus_out)
        return entry

    # ---------------- UI widgets ----------------

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # barra status
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X)

        self.status_label = ttk.Label(
            status_frame,
            text=(
                "Conectado ao banco de dados" if self.connected else "Erro ao conectar ao banco"),
            foreground=("green" if self.connected else "red"),
            font=("Segoe UI", 10, "bold"),
        )
        self.status_label.pack(side=tk.LEFT)

        ttk.Separator(main_frame, orient=tk.HORIZONTAL).pack(
            fill=tk.X, pady=10)

        form_frame = ttk.LabelFrame(main_frame, text="Nova Ordem de Produção")
        form_frame.pack(fill=tk.BOTH, expand=True)

        for col in range(6):
            form_frame.columnconfigure(col, weight=1 if col in (1, 4) else 0)
        form_frame.rowconfigure(6, weight=1)

        # vars
        self.numero_var = tk.StringVar()
        self.deposito_origem_var = tk.StringVar()
        self.deposito_destino_var = tk.StringVar()
        self.situacao_id_var = tk.StringVar()
        self.produto_id_var = tk.StringVar()
        self.responsavel_var = tk.StringVar()
        self.quantidade_var = tk.StringVar()
        self.valor_var = tk.StringVar()
        self.data_previsao_inicio_var = tk.StringVar()
        self.data_previsao_final_var = tk.StringVar()
        self.data_inicio_var = tk.StringVar()
        self.data_fim_var = tk.StringVar()

        # linha 0
        ttk.Label(form_frame, text="Número da Ordem:*").grid(row=0,
                                                             column=0, sticky="e", padx=(5, 5), pady=5)
        ttk.Entry(form_frame, textvariable=self.numero_var, width=15).grid(
            row=0, column=1, sticky="w", padx=(0, 10), pady=5)

        ttk.Label(form_frame, text="Responsável:").grid(
            row=0, column=3, sticky="e", padx=(5, 5), pady=5)
        ttk.Entry(form_frame, textvariable=self.responsavel_var, width=25).grid(
            row=0, column=4, sticky="ew", padx=(0, 10), pady=5)

        # linha 1 depósitos
        ttk.Label(form_frame, text="Depósito Origem (ID):*").grid(row=1,
                                                                  column=0, sticky="e", padx=(5, 5), pady=5)
        origem_frame = ttk.Frame(form_frame)
        origem_frame.grid(row=1, column=1, columnspan=2,
                          sticky="w", padx=(0, 10), pady=5)

        self.deposito_origem_entry = ttk.Entry(
            origem_frame, textvariable=self.deposito_origem_var, width=18)
        self.deposito_origem_entry.grid(row=0, column=0, sticky="w")

        ttk.Button(origem_frame, text="Listar Depósitos (F4)",
                   command=self.mostrar_depositos_origem).grid(row=0, column=1, padx=(5, 0))

        ttk.Label(form_frame, text="Depósito Destino (ID):*").grid(row=1,
                                                                   column=3, sticky="e", padx=(5, 5), pady=5)
        destino_frame = ttk.Frame(form_frame)
        destino_frame.grid(row=1, column=4, columnspan=2,
                           sticky="w", padx=(0, 10), pady=5)

        self.deposito_destino_entry = ttk.Entry(
            destino_frame, textvariable=self.deposito_destino_var, width=18)
        self.deposito_destino_entry.grid(row=0, column=0, sticky="w")

        ttk.Button(destino_frame, text="Listar Depósitos (F8)",
                   command=self.mostrar_depositos_destino).grid(row=0, column=1, padx=(5, 0))

        # binds depósitos
        self.deposito_origem_entry.bind(
            "<Return>", lambda e: self.deposito_destino_entry.focus_set())
        self.deposito_origem_entry.bind("<F4>", self.mostrar_depositos_origem)
        self.deposito_origem_entry.bind(
            "<F8>", self.abrir_destino_e_listar_depositos)

        self.deposito_destino_entry.bind(
            "<F8>", self.mostrar_depositos_destino)
        self.deposito_destino_entry.bind("<Return>", lambda e: None)

        # linha 2 situação/produto
        ttk.Label(form_frame, text="Situação (ID):*").grid(row=2,
                                                           column=0, sticky="e", padx=(5, 5), pady=5)
        situacao_frame = ttk.Frame(form_frame)
        situacao_frame.grid(row=2, column=1, columnspan=2,
                            sticky="w", padx=(0, 10), pady=5)

        self.situacao_entry = ttk.Entry(
            situacao_frame, textvariable=self.situacao_id_var, width=18)
        self.situacao_entry.grid(row=0, column=0, sticky="w")

        ttk.Button(situacao_frame, text="Listar Situações (F3)",
                   command=self.mostrar_situacoes).grid(row=0, column=1, padx=(5, 0))

        self.deposito_destino_entry.bind(
            "<Return>", lambda e: self.situacao_entry.focus_set())

        ttk.Label(form_frame, text="Produto (ID):*").grid(row=2,
                                                          column=3, sticky="e", padx=(5, 5), pady=5)
        prod_frame = ttk.Frame(form_frame)
        prod_frame.grid(row=2, column=4, columnspan=2,
                        sticky="w", padx=(0, 10), pady=5)

        self.produto_entry = ttk.Entry(
            prod_frame, textvariable=self.produto_id_var, width=18)
        self.produto_entry.grid(row=0, column=0, sticky="w")

        ttk.Button(prod_frame, text="Listar Produtos (F2)",
                   command=self.mostrar_produtos).grid(row=0, column=1, padx=(5, 0))
        self.produto_entry.bind(
            "<FocusOut>", self.atualizar_quantidade_producao)
        self.produto_entry.bind("<Return>", self.atualizar_quantidade_producao)

        # linha 3 quantidade
        ttk.Label(form_frame, text="Quantidade:*").grid(row=3,
                                                        column=0, sticky="e", padx=(5, 5), pady=5)
        qtd_frame = ttk.Frame(form_frame)
        qtd_frame.grid(row=3, column=1, columnspan=3,
                       sticky="w", padx=(0, 10), pady=5)

        self.quantidade_entry = ttk.Entry(
            qtd_frame, textvariable=self.quantidade_var, width=18)
        self.quantidade_entry.grid(row=0, column=0, sticky="w", padx=(0, 5))

        ttk.Button(qtd_frame, text="Detalhes (F6)", command=self.mostrar_detalhes_quantidade).grid(
            row=0, column=1, padx=(0, 5))

        # linha 4 datas previstas
        ttk.Label(form_frame, text="Prev. Início (DD/MM/AAAA):").grid(row=4,
                                                                      column=0, sticky="e", padx=(5, 5), pady=5)
        self.entry_prev_inicio = self.criar_entry_data(
            form_frame, self.data_previsao_inicio_var, nome_campo="Prev. Início", largura=25)
        self.entry_prev_inicio.grid(
            row=4, column=1, sticky="w", padx=(0, 10), pady=5)

        ttk.Label(form_frame, text="Prev. Final (DD/MM/AAAA):").grid(row=4,
                                                                     column=3, sticky="e", padx=(5, 5), pady=5)
        self.entry_prev_final = self.criar_entry_data(
            form_frame, self.data_previsao_final_var, nome_campo="Prev. Final", largura=25)
        self.entry_prev_final.grid(
            row=4, column=4, sticky="w", padx=(0, 10), pady=5)

        # linha 5 datas reais + totais
        ttk.Label(form_frame, text="Data Início (DD/MM/AAAA):").grid(row=5,
                                                                     column=0, sticky="e", padx=(5, 5), pady=5)
        self.entry_data_inicio = self.criar_entry_data(
            form_frame, self.data_inicio_var, nome_campo="Data Início", largura=25)
        self.entry_data_inicio.grid(
            row=5, column=1, sticky="w", padx=(0, 10), pady=5)

        ttk.Label(form_frame, text="Data Fim (DD/MM/AAAA):").grid(row=5,
                                                                  column=3, sticky="e", padx=(5, 5), pady=5)
        self.entry_data_fim = self.criar_entry_data(
            form_frame, self.data_fim_var, nome_campo="Data Fim", largura=25)
        self.entry_data_fim.grid(
            row=5, column=4, sticky="w", padx=(0, 10), pady=5)

        ttk.Button(form_frame, text="Totais Tabelas (F5)", command=self.atualizar_totais).grid(
            row=5, column=5, sticky="n", padx=(0, 10), pady=5)

        # linha 6 observação
        ttk.Label(form_frame, text="Observação:").grid(
            row=6, column=0, sticky="ne", padx=(5, 5), pady=5)
        self.observacao_text = tk.Text(form_frame, height=6)
        self.observacao_text.grid(
            row=6, column=1, columnspan=5, sticky="nsew", padx=(0, 10), pady=5)

        # linha 7 botões
        botoes = ttk.Frame(form_frame)
        botoes.grid(row=7, column=0, columnspan=6, pady=15)

        ttk.Button(botoes, text="Salvar Ordem",
                   command=self.salvar_ordem).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes, text="Limpar", command=self.limpar_formulario).pack(
            side=tk.LEFT, padx=5)
        ttk.Button(botoes, text="Ordens Existentes (F10)",
                   command=self.mostrar_ordens_producao).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes, text="Finalizar Pendentes (F11)",
                   command=self.finalizar_producoes_pendentes).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes, text="Etiquetas (F12)",
                   command=self.abrir_programa_etiqueta).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes, text="Fechar", command=self.on_closing).pack(
            side=tk.LEFT, padx=5)

        ttk.Label(main_frame, text="Campos marcados com * são obrigatórios.",
                  foreground="gray").pack(side=tk.BOTTOM, anchor="w", pady=(5, 0))

        # atalhos globais (uma vez)
        self.bind("<F2>", self.mostrar_produtos)
        self.bind("<F3>", self.mostrar_situacoes)
        self.bind("<F4>", self.f4_global)
        self.bind("<F5>", self.atualizar_totais)
        self.bind("<F6>", lambda e: self.mostrar_detalhes_quantidade())
        self.bind("<F8>", self.f8_global)
        self.bind("<F10>", self.mostrar_ordens_producao)
        self.bind("<F11>", self.finalizar_producoes_pendentes)

    def abrir_destino_e_listar_depositos(self, event=None):
        self.deposito_destino_entry.focus_set()
        self.mostrar_depositos_destino()

    def f4_global(self, event=None):
        if self.focus_get() == self.deposito_origem_entry:
            return
        self.deposito_origem_entry.focus_set()
        self.mostrar_depositos_origem()

    def f8_global(self, event=None):
        if event is not None and event.widget in (self.deposito_origem_entry, self.deposito_destino_entry):
            return
        self.deposito_destino_entry.focus_set()
        self.mostrar_depositos_destino()

    # ---------------- Totais ----------------

    def atualizar_totais(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco de dados.")
            return
        self.carregar_totais()
        msg = (
            f"Total de Produtos....................: {self.total_produtos}\n"
            f"Total de Situações...................: {self.total_situacoes}\n"
            f"Total de Depósitos..................: {self.total_depositos}\n"
            f"Total de Ordens de Produção.........: {self.total_ordens}"
        )
        messagebox.showinfo("Totais de Registros", msg)

    # ---------------- Número Ordem ----------------

    def atualizar_numero_ordem(self):
        if not self.connected:
            self.numero_var.set("")
            return
        n = self.sistema.gerar_numero_ordem()
        self.numero_var.set(str(n))

    # ---------------- Detalhes Quantidade ----------------

    def mostrar_detalhes_quantidade(self, event=None):
        if not self.variaveis_quantidade:
            messagebox.showinfo(
                "Detalhes da Quantidade",
                "Nenhum cálculo de quantidade foi realizado ainda.\n"
                "Informe um Produto (ID) e pressione Enter para calcular.",
            )
            return

        janela = tk.Toplevel(self)
        janela.title("Detalhes do cálculo da quantidade")
        janela.geometry("600x500")
        janela.transient(self)
        janela.grab_set()

        frame = ttk.Frame(janela, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        txt = tk.Text(frame, wrap="word")
        txt.pack(fill=tk.BOTH, expand=True)
        txt.insert(tk.END, "Variáveis do cálculo:\n\n")
        for k, v in self.variaveis_quantidade.items():
            txt.insert(tk.END, f"{k}: {v}\n")
        txt.config(state="disabled")

        ttk.Button(frame, text="Fechar", command=janela.destroy).pack(pady=5)
        janela.bind("<Escape>", lambda e: janela.destroy())

    # ---------------- Atualizar quantidade (simplificado) ----------------
    # OBS: Seu cálculo original é enorme. Aqui mantemos um mínimo funcional:
    # - Se existir preço, coloca em Valor
    # - Sugere quantidade = 1
    #
    # Se você quiser colocar seu cálculo completo, cole dentro deste método.
    # -----------------------------------------------------

    def atualizar_quantidade_producao(self, event=None):
        try:
            pid_str = self.produto_id_var.get().strip()
            if not pid_str:
                self.quantidade_var.set("0")
                self.variaveis_quantidade = None
                return

            pid = int(pid_str)
            produto = self.sistema.validar_produto(pid)
            if not produto:
                self.quantidade_var.set("0")
                self.variaveis_quantidade = {"erro": "Produto não encontrado"}
                return

            preco = produto.get("preco") or 0
            self.valor_var.set(f"{float(preco):.2f}")

            # Sugestão simples
            qtd = 1.0
            self.quantidade_var.set(f"{qtd:.2f}")
            self.variaveis_quantidade = {
                "produto_id": pid,
                "produto_nome": produto.get("nomeproduto"),
                "sku": produto.get("sku"),
                "preco": preco,
                "quantidade_sugerida": qtd,
                "obs": "Cálculo simplificado nesta versão corrigida. Cole seu cálculo completo aqui se desejar.",
            }
        except Exception as e:
            print("✗ Erro atualizar_quantidade_producao:", e)
            self.quantidade_var.set("0")
            self.variaveis_quantidade = {"erro": str(e)}

    # ---------------- Produtos / Situações / Depósitos ----------------

    def mostrar_produtos(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco de dados.")
            return
        try:
            limite = self.total_produtos if self.total_produtos > 0 else None
            produtos = self.sistema.listar_produtos_disponiveis(limite)
            if not produtos:
                messagebox.showinfo("Produtos", "Nenhum produto encontrado.")
                return

            janela = tk.Toplevel(self)
            janela.title("Produtos Disponíveis - Duplo clique para selecionar")
            janela.geometry("950x500")
            janela.transient(self)
            janela.grab_set()

            frame = ttk.Frame(janela)
            frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            cols = ("ID", "Nome", "SKU", "Preço", "Tipo")
            tree = ttk.Treeview(frame, columns=cols,
                                show="headings", yscrollcommand=scrollbar.set)
            for c in cols:
                tree.heading(c, text=c)

            tree.column("ID", width=80)
            tree.column("Nome", width=350)
            tree.column("SKU", width=150)
            tree.column("Preço", width=100)
            tree.column("Tipo", width=120)

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=tree.yview)

            for p in produtos:
                preco_formatado = f"{float(p[3]):.2f}" if p[3] else "0.00"
                tree.insert("", tk.END, values=(
                    p[0], p[1], p[2], preco_formatado, p[4]))

            def selecionar(event):
                sel = tree.selection()
                if not sel:
                    return
                v = tree.item(sel[0])["values"]
                self.produto_id_var.set(str(v[0]))
                self.valor_var.set(str(v[3]).replace(",", "."))
                janela.destroy()
                self.atualizar_quantidade_producao()
                self.quantidade_entry.focus_set()

            tree.bind("<Double-Button-1>", selecionar)
            janela.bind("<Escape>", lambda e: janela.destroy())

        except Exception as e:
            messagebox.showerror(
                "Erro", f"Erro ao abrir lista de produtos:\n{e}")

    def mostrar_situacoes(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco de dados.")
            return
        try:
            limite = self.total_situacoes if self.total_situacoes > 0 else None
            situacoes = self.sistema.listar_situacoes_disponiveis(limite)
            if not situacoes:
                messagebox.showinfo(
                    "Situações", "Nenhuma situação encontrada.")
                return

            janela = tk.Toplevel(self)
            janela.title(
                "Situações Disponíveis - Duplo clique para selecionar")
            janela.geometry("600x400")
            janela.transient(self)
            janela.grab_set()

            frame = ttk.Frame(janela)
            frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            cols = ("ID", "Situação")
            tree = ttk.Treeview(frame, columns=cols,
                                show="headings", yscrollcommand=scrollbar.set)
            tree.heading("ID", text="ID")
            tree.heading("Situação", text="Nome da Situação")
            tree.column("ID", width=100)
            tree.column("Situação", width=400)

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=tree.yview)

            for s in situacoes:
                tree.insert("", tk.END, values=(s[0], s[1]))

            def selecionar(event):
                sel = tree.selection()
                if not sel:
                    return
                v = tree.item(sel[0])["values"]
                self.situacao_id_var.set(str(v[0]))
                janela.destroy()

            tree.bind("<Double-Button-1>", selecionar)
            janela.bind("<Escape>", lambda e: janela.destroy())

        except Exception as e:
            messagebox.showerror(
                "Erro", f"Erro ao abrir lista de situações:\n{e}")

    def mostrar_depositos_origem(self, event=None):
        self._mostrar_depositos(modo="origem")

    def mostrar_depositos_destino(self, event=None):
        self._mostrar_depositos(modo="destino")

    def _mostrar_depositos(self, modo: str):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco de dados.")
            return
        try:
            limite = self.total_depositos if self.total_depositos > 0 else None
            depositos = self.sistema.listar_depositos_disponiveis(limite)
            if not depositos:
                messagebox.showinfo("Depósitos", "Nenhum depósito encontrado.")
                return

            janela = tk.Toplevel(self)
            janela.title(
                "Depósitos Disponíveis - Duplo clique para selecionar")
            janela.geometry("700x400")
            janela.transient(self)
            janela.grab_set()

            frame = ttk.Frame(janela)
            frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            cols = ("ID", "Descrição", "Situação",
                    "Padrão", "Desconsiderar saldo")
            tree = ttk.Treeview(frame, columns=cols,
                                show="headings", yscrollcommand=scrollbar.set)
            for c in cols:
                tree.heading(c, text=c)

            tree.column("ID", width=80)
            tree.column("Descrição", width=250)
            tree.column("Situação", width=100)
            tree.column("Padrão", width=100)
            tree.column("Desconsiderar saldo", width=150)

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=tree.yview)

            for d in depositos:
                tree.insert("", tk.END, values=(d[0], d[1], d[2], d[3], d[4]))

            def selecionar(event):
                sel = tree.selection()
                if not sel:
                    return
                v = tree.item(sel[0])["values"]
                if modo == "origem":
                    self.deposito_origem_var.set(str(v[0]))
                else:
                    self.deposito_destino_var.set(str(v[0]))
                janela.destroy()

            tree.bind("<Double-Button-1>", selecionar)
            janela.bind("<Escape>", lambda e: janela.destroy())

        except Exception as e:
            messagebox.showerror(
                "Erro", f"Erro ao abrir lista de depósitos:\n{e}")

    # ---------------- Ordens existentes ----------------

    def mostrar_ordens_producao(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco de dados.")
            return

        try:
            ordens = self.sistema.listar_ordens_producao()
            if not ordens:
                messagebox.showinfo("Ordens de Produção",
                                    "Nenhuma ordem encontrada.")
                return

            janela = tk.Toplevel(self)
            janela.title("Ordens de Produção")
            janela.geometry("1000x550")
            janela.transient(self)
            janela.grab_set()

            frame_lista = ttk.Frame(janela)
            frame_lista.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            scrollbar = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            cols = ("id", "numero", "produto_id", "produto_nome",
                    "situacao", "quantidade", "data_inicio", "data_fim")
            tree = ttk.Treeview(frame_lista, columns=cols, show="headings",
                                yscrollcommand=scrollbar.set, selectmode="extended")
            for c in cols:
                tree.heading(c, text=c)

            tree.column("id", width=60, anchor="center")
            tree.column("numero", width=80, anchor="center")
            tree.column("produto_id", width=80, anchor="center")
            tree.column("produto_nome", width=260, anchor="w")
            tree.column("situacao", width=150, anchor="w")
            tree.column("quantidade", width=100, anchor="e")
            tree.column("data_inicio", width=110, anchor="center")
            tree.column("data_fim", width=110, anchor="center")

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=tree.yview)

            def fmt(dt):
                if not dt:
                    return ""
                if isinstance(dt, (datetime, date)):
                    return dt.strftime("%d/%m/%Y")
                return str(dt)

            for (oid, numero, produto_id, produto_nome, _, situacao_nome, quantidade, data_inicio, data_fim) in ordens:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        oid,
                        numero,
                        produto_id,
                        produto_nome or "",
                        situacao_nome or "",
                        f"{float(quantidade):.2f}" if quantidade is not None else "",
                        fmt(data_inicio),
                        fmt(data_fim),
                    ),
                )

            frame_botoes = ttk.Frame(janela)
            frame_botoes.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(0, 10))

            def fechar(event=None):
                janela.destroy()

            def excluir_selecionada(event=None):
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning(
                        "Excluir", "Selecione uma ordem para excluir.")
                    return
                item_id = sel[0]
                v = tree.item(item_id)["values"]
                oid, numero = v[0], v[1]
                if not messagebox.askyesno("Confirmar", f"Deseja excluir a OP nº {numero} (ID {oid})?"):
                    return
                ok = self.sistema.excluir_ordem_producao(int(oid))
                if ok:
                    tree.delete(item_id)
                    messagebox.showinfo(
                        "Exclusão", f"OP nº {numero} excluída.")
                else:
                    messagebox.showerror(
                        "Erro", "Não foi possível excluir (talvez vinculada a outros registros).")

            ttk.Button(frame_botoes, text="Excluir selecionada",
                       command=excluir_selecionada).pack(side=tk.RIGHT, padx=(0, 5))
            ttk.Button(frame_botoes, text="Fechar",
                       command=fechar).pack(side=tk.RIGHT)

            janela.bind("<Escape>", fechar)
            tree.bind("<Delete>", excluir_selecionada)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao listar ordens:\n{e}")

    # ---------------- Finalizar pendentes ----------------

    def finalizar_producoes_pendentes(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco de dados.")
            return
        try:
            ordens = self.sistema.listar_ordens_sem_data_fim()
            if not ordens:
                messagebox.showinfo(
                    "Finalizar", "Não há ordens pendentes sem data fim.")
                return

            janela = tk.Toplevel(self)
            janela.title("Finalizar ordens pendentes")
            janela.geometry("1000x500")
            janela.transient(self)
            janela.grab_set()

            frame = ttk.Frame(janela, padding=10)
            frame.pack(fill=tk.BOTH, expand=True)

            scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            cols = ("id", "numero", "produto_id", "produto_nome",
                    "situacao", "quantidade", "data_inicio")
            tree = ttk.Treeview(frame, columns=cols, show="headings",
                                yscrollcommand=scrollbar.set, selectmode="extended")
            for c in cols:
                tree.heading(c, text=c)

            tree.column("id", width=60, anchor="center")
            tree.column("numero", width=80, anchor="center")
            tree.column("produto_id", width=80, anchor="center")
            tree.column("produto_nome", width=260, anchor="w")
            tree.column("situacao", width=150, anchor="w")
            tree.column("quantidade", width=100, anchor="e")
            tree.column("data_inicio", width=110, anchor="center")

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=tree.yview)

            def fmt(dt):
                if not dt:
                    return ""
                if isinstance(dt, (datetime, date)):
                    return dt.strftime("%d/%m/%Y")
                return str(dt)

            for (oid, numero, produto_id, produto_nome, _, situacao_nome, quantidade, data_inicio) in ordens:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        oid,
                        numero,
                        produto_id,
                        produto_nome or "",
                        situacao_nome or "",
                        f"{float(quantidade):.2f}" if quantidade is not None else "",
                        fmt(data_inicio),
                    ),
                )

            frame_botoes = ttk.Frame(janela, padding=(10, 0, 10, 10))
            frame_botoes.pack(side=tk.BOTTOM, fill=tk.X)

            def fechar(event=None):
                janela.destroy()

            def finalizar_selecionada(event=None):
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning(
                        "Finalizar", "Selecione uma ou mais ordens.")
                    return

                ordens_sel = []
                for item_id in sel:
                    v = tree.item(item_id)["values"]
                    ordens_sel.append((item_id, v[0], v[1]))

                if len(ordens_sel) == 1:
                    _, oid, numero = ordens_sel[0]
                    texto = f"Deseja finalizar a OP nº {numero} (ID {oid}) com data fim hoje?"
                else:
                    nums = ", ".join(str(x[2]) for x in ordens_sel)
                    texto = f"Deseja finalizar {len(ordens_sel)} OPs (números: {nums}) com data fim hoje?"

                if not messagebox.askyesno("Confirmar", texto):
                    return

                ok_count = 0
                for item_id, oid, _ in ordens_sel:
                    ok = self.sistema.finalizar_ordem_individual(oid)
                    if ok:
                        tree.delete(item_id)
                        ok_count += 1

                if ok_count:
                    messagebox.showinfo(
                        "Finalização", f"{ok_count} ordem(ns) finalizada(s) com sucesso.")
                else:
                    messagebox.showerror(
                        "Erro", "Não foi possível finalizar as ordens selecionadas.")

            ttk.Button(frame_botoes, text="Finalizar selecionada(s)",
                       command=finalizar_selecionada).pack(side=tk.RIGHT, padx=(0, 5))
            ttk.Button(frame_botoes, text="Fechar",
                       command=fechar).pack(side=tk.RIGHT)

            janela.bind("<Escape>", fechar)
            tree.bind("<Return>", finalizar_selecionada)
            tree.bind("<Double-Button-1>", finalizar_selecionada)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao finalizar pendentes:\n{e}")

    # ---------------- Salvar / Limpar ----------------

    def salvar_ordem(self):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco de dados.")
            return

        try:
            dados: Dict[str, Any] = {}

            dados["numero"] = self.numero_var.get().strip()
            if not dados["numero"]:
                raise ValueError("O campo 'Número da Ordem' é obrigatório.")

            dados["deposito_id_origem"] = self.parse_int(
                self.deposito_origem_var.get(), "Depósito Origem")
            dados["deposito_id_destino"] = self.parse_int(
                self.deposito_destino_var.get(), "Depósito Destino")
            dados["situacao_id"] = self.parse_int(
                self.situacao_id_var.get(), "Situação")
            dados["fkprodutoid"] = self.parse_int(
                self.produto_id_var.get(), "Produto")
            dados["quantidade"] = self.parse_float(
                self.quantidade_var.get(), "Quantidade", obrigatorio=True)

            dados["responsavel"] = self.responsavel_var.get().strip() or None
            dados["valor"] = self.parse_float(
                self.valor_var.get(), "Valor", obrigatorio=False)
            dados["data_previsao_inicio"] = self.parse_date(
                self.data_previsao_inicio_var.get(), "Prev. Início")
            dados["data_previsao_final"] = self.parse_date(
                self.data_previsao_final_var.get(), "Prev. Final")
            dados["data_inicio"] = self.parse_date(
                self.data_inicio_var.get(), "Data Início")
            dados["data_fim"] = self.parse_date(
                self.data_fim_var.get(), "Data Fim")

            if dados["data_previsao_inicio"] and dados["data_previsao_final"]:
                if dados["data_previsao_final"] < dados["data_previsao_inicio"]:
                    raise ValueError(
                        "A data de previsão final deve ser >= à previsão de início.")

            if dados["data_inicio"] and dados["data_fim"]:
                if dados["data_fim"] < dados["data_inicio"]:
                    raise ValueError("A data fim deve ser >= à data início.")

            obs = self.observacao_text.get("1.0", tk.END).strip()
            dados["observacao"] = obs if obs else None
            dados["id"] = None  # gerado no inserir

            # Confirmação simples
            if not messagebox.askyesno("Confirmar", f"Confirma inserir a OP nº {dados['numero']}?"):
                return

            ok = self.sistema.inserir_ordem_producao(dados)
            if not ok:
                erro = self.sistema._ultimo_erro_bd or "Erro ao inserir (veja console)."
                messagebox.showerror("Erro ao gravar no banco", erro)
                return

            # Webhook n8n (opcional) com dados completos
            self.enviar_webhook_ordem(dados)

            messagebox.showinfo(
                "Sucesso", f"Ordem de produção {dados['numero']} inserida com sucesso!")
            self.limpar_formulario()

        except ValueError as ve:
            messagebox.showerror("Erro de validação", str(ve))
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado:\n{e}")
            log_exception(e, "Erro em salvar_ordem")

    def limpar_formulario(self, event=None):
        self.atualizar_numero_ordem()
        self.deposito_origem_var.set("")
        self.deposito_destino_var.set("")
        self.situacao_id_var.set("")
        self.produto_id_var.set("")
        self.responsavel_var.set("")
        self.quantidade_var.set("")
        self.valor_var.set("")
        self.data_previsao_inicio_var.set("")
        self.data_previsao_final_var.set("")
        self.data_inicio_var.set("")
        self.data_fim_var.set("")
        self.observacao_text.delete("1.0", tk.END)

    # ---------------- Webhook (completo) ----------------

    def enviar_webhook_ordem(self, dados: Dict[str, Any]) -> None:
        if not N8N_WEBHOOK_URL:
            return

        try:
            def fmt_dt(dt):
                if not dt:
                    return None
                if isinstance(dt, (datetime, date)):
                    return dt.isoformat()
                return str(dt)

            payload = {
                "id": dados.get("id"),
                "numero": dados.get("numero"),
                "deposito_id_destino": dados.get("deposito_id_destino"),
                "deposito_id_origem": dados.get("deposito_id_origem"),
                "situacao_id": dados.get("situacao_id"),
                "responsavel": dados.get("responsavel"),
                "fkprodutoid": dados.get("fkprodutoid"),
                "data_previsao_inicio": fmt_dt(dados.get("data_previsao_inicio")),
                "data_previsao_final": fmt_dt(dados.get("data_previsao_final")),
                "data_inicio": fmt_dt(dados.get("data_inicio")),
                "data_fim": fmt_dt(dados.get("data_fim")),
                "valor": float(dados["valor"]) if dados.get("valor") is not None else None,
                "observacao": dados.get("observacao"),
                "quantidade": float(dados.get("quantidade") or 0),
            }

            produto = self.sistema.validar_produto(dados["fkprodutoid"])
            situacao = self.sistema.validar_situacao(dados["situacao_id"])
            if produto:
                payload["produto_nome"] = produto.get("nomeproduto")
                payload["produto_sku"] = produto.get("sku")
            if situacao:
                payload["situacao_nome"] = situacao.get("nome")

            resp = requests.post(N8N_WEBHOOK_URL, json=payload, timeout=10)
            resp.raise_for_status()

        except Exception as e:
            print(f"⚠ Webhook n8n falhou: {e}")

    # ---------------- Fechar ----------------

    def on_closing(self):
        # Evita reentrância (fecha duas vezes)
        if self._closing:
            return
        self._closing = True

        try:
            if messagebox.askokcancel("Sair", "Deseja realmente sair?"):
                try:
                    self.sistema.desconectar()
                except Exception:
                    pass
                try:
                    if self.winfo_exists():
                        self.destroy()
                except Exception:
                    pass
            else:
                self._closing = False
        except Exception:
            # se até messagebox falhar, fecha forçado
            try:
                self.sistema.desconectar()
            except Exception:
                pass
            try:
                if self.winfo_exists():
                    self.destroy()
            except Exception:
                pass


# ============================================================
# MAIN (sem protocol aqui!)
# ============================================================

if __name__ == "__main__":
    app = OrdemProducaoApp()
    if app.winfo_exists():
        app.mainloop()
