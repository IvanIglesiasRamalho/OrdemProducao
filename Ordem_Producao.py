from __future__ import annotations

# ============================================================
# Ordem_Producao.py (arquivo único corrigido)
# - Sem splash
# - Fecha voltando para menu_principal.py (--skip-entrada)
# ============================================================

import os
import sys
import json
import traceback
import subprocess
from dataclasses import dataclass, asdict
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
from math import ceil
from collections import defaultdict
from typing import Optional, Dict, Any, List, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

import requests
import psycopg2
from psycopg2 import errors

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import mm
from reportlab.pdfgen import canvas


# ============================================================
# PATHS / CONFIG
# ============================================================

def get_app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = get_app_dir()

BASE_DIR = r"Z:\Planilhas_OP"
if not os.path.exists(BASE_DIR):
    try:
        os.makedirs(BASE_DIR, exist_ok=True)
    except Exception:
        BASE_DIR = os.path.join(APP_DIR, "Planilhas_OP")
        os.makedirs(BASE_DIR, exist_ok=True)

CAMINHO_MODELO = os.path.join(BASE_DIR, "pedido-de-compra v2.xlsx")
CAMINHO_SAIDA = os.path.join(BASE_DIR, "saida_pedido-de-compra v2.xlsx")

# webhook opcional (n8n); use "" para desabilitar
N8N_WEBHOOK_URL = "http://localhost:56789/webhook/ordem-producao"


def log_exception(err: Exception, context: str = "") -> str:
    try:
        texto = "".join(traceback.format_exception(
            type(err), err, err.__traceback__))
        log_path = os.path.join(BASE_DIR, "erro_app.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n" + ("=" * 80) + "\n")
            if context:
                f.write(f"{context}\n")
            f.write(texto)
        return log_path
    except Exception:
        return os.path.join(BASE_DIR, "erro_app.log")


# ============================================================
# ÍCONE
# ============================================================

def find_icon_path() -> Optional[str]:
    candidates = [
        os.path.join(BASE_DIR, "imagens", "favicon.ico"),
        os.path.join(BASE_DIR, "favicon.ico"),
        os.path.join(APP_DIR, "imagens", "favicon.ico"),
        os.path.join(APP_DIR, "favicon.ico"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return None


def apply_window_icon(win) -> None:
    try:
        ico = find_icon_path()
        if ico:
            try:
                win.iconbitmap(default=ico)
            except Exception:
                pass

        png_candidates = [
            os.path.join(BASE_DIR, "imagens", "favicon.png"),
            os.path.join(APP_DIR, "imagens", "favicon.png"),
        ]
        png = next((p for p in png_candidates if os.path.isfile(p)), None)
        if png:
            img = tk.PhotoImage(file=png)
            win.iconphoto(True, img)
            win._icon_img = img  # mantém referência
    except Exception:
        pass


# ============================================================
# EXECUÇÃO DO MENU PRINCIPAL (sem entrada)
# ============================================================

def pick_python_gui_cmd() -> List[str]:
    if os.name != "nt":
        return [sys.executable]

    candidates = [
        ["pyw", "-3.12"],
        ["pyw"],
        ["pythonw"],
        ["python"],
    ]
    for cmd in candidates:
        try:
            subprocess.run(cmd + ["-c", "print('ok')"],
                           capture_output=True, text=True, timeout=2)
            return cmd
        except Exception:
            continue
    return ["python"]


def build_run_cmd(script_path: str, extra_args: Optional[List[str]] = None) -> List[str]:
    extra_args = extra_args or []
    if script_path.lower().endswith(".exe"):
        return [script_path] + extra_args
    if os.name == "nt" and script_path.lower().endswith(".py"):
        return pick_python_gui_cmd() + [script_path] + extra_args
    return [sys.executable, script_path] + extra_args


def abrir_menu_principal_skip_entrada(parent: Optional[tk.Misc] = None) -> None:
    """
    Abre menu_principal.py com --skip-entrada.
    """
    menu_path = os.path.join(APP_DIR, "menu_principal.py")
    if not os.path.isfile(menu_path):
        messagebox.showerror(
            "Menu Principal", f"Não encontrei:\n{menu_path}", parent=parent)
        return

    cmd = build_run_cmd(menu_path, ["--skip-entrada"])
    try:
        subprocess.Popen(cmd, cwd=os.path.dirname(menu_path) or APP_DIR)
    except Exception as e:
        messagebox.showerror(
            "Menu Principal", f"Falha ao abrir o menu:\n{e}", parent=parent)


# ============================================================
# CONFIG
# ============================================================

@dataclass
class AppConfig:
    db_host: str = "10.0.0.154"
    db_database: str = "postgresekenox"
    db_user: str = "postgresekenox"
    db_password: str = "Ekenox5426"
    db_port: int = 55432

    f7_geometry: str = "1100x560"

    caminho_modelo: str = CAMINHO_MODELO
    caminho_saida: str = CAMINHO_SAIDA

    bling_base_url: str = "https://api.bling.com.br/Api/v3"
    bling_token: str = ""
    bling_timeout: int = 20


def config_path() -> str:
    return os.path.join(BASE_DIR, "config_op.json")


def load_config() -> AppConfig:
    p = config_path()
    if not os.path.exists(p):
        return AppConfig()
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        return AppConfig(**data)
    except Exception:
        return AppConfig()


def save_config(cfg: AppConfig) -> None:
    p = config_path()
    with open(p, "w", encoding="utf-8") as f:
        json.dump(asdict(cfg), f, ensure_ascii=False, indent=2)


# ============================================================
# EXCEL HELPERS (Pedido de Compra)
# ============================================================

def _nome_aba_excel_valido(nome: str) -> str:
    invalidos = ['\\', '/', '?', '*', '[', ']']
    for ch in invalidos:
        nome = nome.replace(ch, " ")
    return nome[:31]


def _escrever_celula_segura(ws, coord: str, valor):
    try:
        cell = ws[coord]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if coord in merged_range:
                    top_left_coord = merged_range.coord.split(":")[0]
                    ws[top_left_coord] = valor
                    return
        else:
            cell.value = valor
    except Exception:
        pass


def _set_cell_segura_rc(ws, row: int, col: int, valor):
    cell = ws.cell(row=row, column=col)
    coord = cell.coordinate
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                ws.cell(row=merged_range.min_row,
                        column=merged_range.min_col).value = valor
                return
    else:
        cell.value = valor


def gerar_abas_fornecedor_pedido(
    dados: List[Dict[str, Any]],
    nome_aba_modelo: str = "Pedido de Compra",
    caminho_modelo: str = CAMINHO_MODELO,
    caminho_saida: str = CAMINHO_SAIDA,
):
    if os.path.exists(caminho_saida):
        wb = load_workbook(caminho_saida)
    else:
        wb = load_workbook(caminho_modelo)

    if nome_aba_modelo not in wb.sheetnames:
        raise ValueError(
            f"Aba de modelo '{nome_aba_modelo}' não encontrada. Abas: {wb.sheetnames}")

    aba_modelo = wb[nome_aba_modelo]

    tmp = defaultdict(list)
    for item in dados:
        fornecedor = item["fornecedor"]
        numero_pedido = item["numero_pedido"]
        data_pedido = item.get("data_pedido")
        tmp[(fornecedor, numero_pedido, data_pedido)].append(item)

    for (fornecedor, numero_pedido, data_pedido), linhas in tmp.items():
        ws = wb.copy_worksheet(aba_modelo)
        titulo_aba = f"{numero_pedido} - {str(fornecedor)[:15]}"
        ws.title = _nome_aba_excel_valido(titulo_aba)

        _escrever_celula_segura(ws, "D6", numero_pedido)
        _escrever_celula_segura(ws, "D8", data_pedido or date.today())
        _escrever_celula_segura(ws, "D10", str(fornecedor))

        for r in range(16, 43):
            for c in range(2, 10):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        linha = 16
        numero_item = 1

        for item in linhas:
            descricao = item["produto"]
            quantidade = float(item["quantidade"])

            estoque_atual = float(item.get("estoque_atual", 0.0) or 0.0)
            estoque_minimo = float(item.get("estoque_minimo", 0.0) or 0.0)
            estoque_maximo = float(item.get("estoque_maximo", 0.0) or 0.0)
            valor_unitario = item.get("valor_unitario")

            _set_cell_segura_rc(ws, linha, 2, numero_item)       # B
            _set_cell_segura_rc(ws, linha, 3, descricao)         # C
            _set_cell_segura_rc(ws, linha, 4, estoque_atual)     # D
            _set_cell_segura_rc(ws, linha, 5, estoque_minimo)    # E
            _set_cell_segura_rc(ws, linha, 6, estoque_maximo)    # F

            if valor_unitario is not None:
                vu = float(valor_unitario)
                _set_cell_segura_rc(ws, linha, 7, vu)            # G
                _set_cell_segura_rc(ws, linha, 9, quantidade * vu)  # I

            _set_cell_segura_rc(ws, linha, 8, quantidade)        # H

            numero_item += 1
            linha += 1

    wb.save(caminho_saida)


# ============================================================
# DB
# ============================================================

class SistemaOrdemProducao:
    def __init__(self, cfg: AppConfig):
        self.cfg = cfg
        self.conn: Optional[psycopg2.extensions.connection] = None
        self.cursor: Optional[psycopg2.extensions.cursor] = None
        self.ultimo_erro: Optional[str] = None

    def conectar(self) -> bool:
        self.ultimo_erro = None
        try:
            self.conn = psycopg2.connect(
                host=self.cfg.db_host,
                database=self.cfg.db_database,
                user=self.cfg.db_user,
                password=self.cfg.db_password,
                port=int(self.cfg.db_port),
                connect_timeout=5,
            )
            self.cursor = self.conn.cursor()
            return True
        except Exception as e:
            self.ultimo_erro = f"{type(e).__name__}: {e}"
            return False

    def desconectar(self):
        try:
            if self.cursor:
                self.cursor.close()
        except Exception:
            pass
        try:
            if self.conn:
                self.conn.close()
        except Exception:
            pass

    def _q(self, sql: str, params: Tuple = ()) -> None:
        if not self.cursor:
            raise RuntimeError("Sem cursor (não conectado).")
        self.cursor.execute(sql, params)

    def gerar_numero_ordem(self) -> int:
        try:
            self._q(
                """SELECT COALESCE(MAX(numero),0) + 1 FROM "Ekenox"."ordem_producao";""")
            r = self.cursor.fetchone()
            return int(r[0]) if r and r[0] is not None else 1
        except Exception:
            return 1

    def gerar_id_ordem(self) -> int:
        try:
            self._q(
                """SELECT COALESCE(MAX(id),0) + 1 FROM "Ekenox"."ordem_producao";""")
            r = self.cursor.fetchone()
            return int(r[0]) if r and r[0] is not None else 1
        except Exception:
            return 1

    def validar_produto(self, produto_id: int) -> Optional[Dict[str, Any]]:
        try:
            sql = """
                SELECT p."produtoId", p."nomeProduto", p."sku", p."preco", p."tipo"
                  FROM "Ekenox"."produtos" p
                 WHERE p."produtoId" = %s;
            """
            self._q(sql, (int(produto_id),))
            r = self.cursor.fetchone()
            if not r:
                return None
            return {"produtoid": r[0], "nomeproduto": r[1], "sku": r[2], "preco": r[3], "tipo": r[4]}
        except Exception:
            if self.conn:
                self.conn.rollback()
            return None

    def validar_situacao(self, situacao_id: int) -> Optional[Dict[str, Any]]:
        try:
            self._q(
                """SELECT s."id", s."nome" FROM "Ekenox"."situacao" s WHERE s."id" = %s;""",
                (int(situacao_id),),
            )
            r = self.cursor.fetchone()
            if not r:
                return None
            return {"id": r[0], "nome": r[1]}
        except Exception:
            if self.conn:
                self.conn.rollback()
            return None

    def saldo_fisico(self, produto_id: int) -> float:
        try:
            sql = """
                SELECT COALESCE(SUM(e."saldoFisico"), 0)
                  FROM "Ekenox"."estoque" e
                 WHERE e."fkProduto"::bigint = %s;
            """
            self._q(sql, (int(produto_id),))
            r = self.cursor.fetchone()
            return float(r[0]) if r and r[0] is not None else 0.0
        except Exception:
            if self.conn:
                self.conn.rollback()
            return 0.0

    def media_vendas_mensal(self, fk_produto: int, ano: int | None = None, mes: int | None = None) -> float:
        try:
            if ano is None or mes is None:
                hoje = date.today()
                ano = hoje.year
                mes = hoje.month

            mes_inicio = date(int(ano), int(mes), 1)
            sql = """
                SELECT v.media_vendas
                FROM vw_media_vendas_mensal v
                WHERE v.fkProduto = %s
                  AND date_trunc('month', v.dataVenda)::date = %s
                LIMIT 1
            """
            self._q(sql, (int(fk_produto), mes_inicio))
            row = self.cursor.fetchone()
            return float(row[0]) if row and row[0] is not None else 0.0
        except Exception:
            if self.conn:
                self.conn.rollback()
            return 0.0

    def buscar_qtd_produzir_por_sku(self, sku: str) -> float:
        try:
            sku_norm = (sku or "").strip().upper()
            if not sku_norm:
                return 0.0
            candidatos = {sku_norm}
            if sku_norm.endswith("N"):
                candidatos.add(sku_norm[:-1])
            else:
                candidatos.add(sku_norm + "N")

            sql = """
                SELECT COALESCE(SUM(a."quantidade"), 0)
                  FROM "Ekenox"."arranjo" a
                 WHERE UPPER(TRIM(a."sku")) = ANY(%s);
            """
            self._q(sql, (sorted(candidatos),))
            row = self.cursor.fetchone()
            return float(row[0]) if row and row[0] is not None else 0.0
        except Exception:
            if self.conn:
                self.conn.rollback()
            return 0.0

    def listar_produtos_disponiveis(self, limite: Optional[int] = None):
        try:
            base_sql = """
                SELECT p."produtoId", p."nomeProduto", p."sku", p."preco", p."tipo"
                  FROM "Ekenox"."produtos" AS p
                  JOIN "Ekenox"."infoProduto" AS ip
                    ON ip."fkProduto" = p."produtoId"
                 WHERE ip."fkCategoria" NOT IN (
                    3844533,3983855,7879429,3869959,4241123,
                    3870601,3844542,7651801,3983399,959867,
                    897565,3984869,3862825,7879102,7911660,
                    4828356,6568231
                 )
                 ORDER BY p."nomeProduto"
            """
            if limite and limite > 0:
                self._q(base_sql + " LIMIT %s", (int(limite),))
            else:
                self._q(base_sql)
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def listar_situacoes_disponiveis(self, limite: Optional[int] = None):
        try:
            sql = """SELECT s."id", s."nome" FROM "Ekenox"."situacao" s ORDER BY s."nome" """
            if limite and limite > 0:
                sql += " LIMIT %s"
                self._q(sql, (int(limite),))
            else:
                self._q(sql)
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def listar_depositos_disponiveis(self, limite: Optional[int] = None):
        try:
            sql = """
                SELECT d."id", d."descricao", d."situacao", d."padrao", d."desconsiderarsaldo"
                  FROM "Ekenox"."deposito" d
                 ORDER BY d."descricao"
            """
            if limite and limite > 0:
                sql += " LIMIT %s"
                self._q(sql, (int(limite),))
            else:
                self._q(sql)
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def inserir_ordem_producao(self, dados: Dict[str, Any]) -> tuple[bool, str]:
        try:
            produto = self.validar_produto(int(dados["fkprodutoid"]))
            if not produto:
                return False, f"Produto ID {dados['fkprodutoid']} não encontrado."

            situacao = self.validar_situacao(int(dados["situacao_id"]))
            if not situacao:
                return False, f"Situação ID {dados['situacao_id']} não encontrada."

            if not dados.get("id"):
                dados["id"] = self.gerar_id_ordem()

            query = """
                INSERT INTO "Ekenox"."ordem_producao" (
                    id, numero, deposito_destino, deposito_origem, situacao_id,
                    responsavel, fkprodutoid, data_previsao_inicio, data_previsao_final,
                    data_inicio, data_fim, valor, observacao, quantidade
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """
            valores = (
                int(dados["id"]),
                int(dados["numero"]),
                int(dados["deposito_id_destino"]),
                int(dados["deposito_id_origem"]),
                int(dados["situacao_id"]),
                dados.get("responsavel"),
                int(dados["fkprodutoid"]),
                dados.get("data_previsao_inicio"),
                dados.get("data_previsao_final"),
                dados.get("data_inicio"),
                dados.get("data_fim"),
                dados.get("valor"),
                dados.get("observacao"),
                float(dados["quantidade"]),
            )

            self.cursor.execute(query, valores)
            self.conn.commit()
            return True, ""

        except errors.UniqueViolation as e:
            if self.conn:
                self.conn.rollback()
            cn = getattr(e.diag, 'constraint_name', '')
            return False, f"NÚMERO DE ORDEM JÁ EXISTENTE.\nNúmero: {dados.get('numero')}\nConstraint: {cn}"

        except errors.ForeignKeyViolation as e:
            if self.conn:
                self.conn.rollback()
            tabela = getattr(e.diag, "table_name", "desconhecida")
            constraint = getattr(e.diag, "constraint_name", "desconhecida")
            return False, (
                "VIOLAÇÃO DE CHAVE ESTRANGEIRA.\n\n"
                f"Tabela alvo: {tabela}\n"
                f"Constraint: {constraint}\n\n"
                "Provavelmente algum ID não existe (depósitos, situação ou produto)."
            )

        except Exception as e:
            if self.conn:
                self.conn.rollback()
            return False, f"Erro ao inserir OP: {type(e).__name__}: {e}"

    def buscar_ordem_producao_por_numero(self, numero: str | int):
        try:
            num_int = int(str(numero).strip())
            sql = """
                SELECT
                    o."id", o."numero", o."deposito_destino", o."deposito_origem",
                    o."situacao_id", o."responsavel", o."fkprodutoid",
                    o."data_previsao_inicio", o."data_previsao_final",
                    o."data_inicio", o."data_fim", o."valor", o."observacao", o."quantidade"
                FROM "Ekenox"."ordem_producao" o
                WHERE o."numero" = %s
                LIMIT 1;
            """
            self._q(sql, (num_int,))
            r = self.cursor.fetchone()
            if not r:
                return None

            return {
                "id": r[0],
                "numero": r[1],
                "deposito_destino": r[2],
                "deposito_origem": r[3],
                "situacao_id": r[4],
                "responsavel": r[5],
                "fkprodutoid": r[6],
                "data_previsao_inicio": r[7],
                "data_previsao_final": r[8],
                "data_inicio": r[9],
                "data_fim": r[10],
                "valor": r[11],
                "observacao": r[12],
                "quantidade": r[13],
            }
        except Exception:
            if self.conn:
                self.conn.rollback()
            return None

    def listar_ordens_producao(self):
        try:
            sql = """
                SELECT
                    o."id",
                    o."numero",
                    o."fkprodutoid",
                    p."nomeProduto" AS produto_nome,
                    o."situacao_id",
                    s."nome" AS situacao_nome,
                    o."quantidade",
                    o."data_inicio",
                    o."data_fim"
                FROM "Ekenox"."ordem_producao" o
                LEFT JOIN "Ekenox"."produtos" p
                       ON p."produtoId" = o."fkprodutoid"
                LEFT JOIN "Ekenox"."situacao" s
                       ON s."id" = o."situacao_id"
                ORDER BY o."id" DESC;
            """
            self._q(sql)
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def excluir_ordem_producao(self, ordem_id: int) -> bool:
        try:
            self._q(
                'DELETE FROM "Ekenox"."ordem_producao" WHERE "id" = %s;', (int(ordem_id),))
            ok = (self.cursor.rowcount or 0) > 0
            self.conn.commit()
            return ok
        except Exception:
            if self.conn:
                self.conn.rollback()
            return False

    def listar_ordens_sem_data_fim(self):
        try:
            sql = """
                SELECT
                    o."id",
                    o."numero",
                    o."fkprodutoid",
                    p."nomeProduto" AS produto_nome,
                    o."situacao_id",
                    s."nome" AS situacao_nome,
                    o."quantidade",
                    o."data_inicio"
                FROM "Ekenox"."ordem_producao" o
                LEFT JOIN "Ekenox"."produtos" p
                       ON p."produtoId" = o."fkprodutoid"
                LEFT JOIN "Ekenox"."situacao" s
                       ON s."id" = o."situacao_id"
                WHERE o."data_fim" IS NULL OR o."data_fim" = '1970-01-01'
                ORDER BY o."id" DESC;
            """
            self._q(sql)
            return self.cursor.fetchall() or []
        except Exception:
            if self.conn:
                self.conn.rollback()
            return []

    def finalizar_ordem_individual(self, ordem_id: int) -> bool:
        try:
            hoje = date.today()
            sql = """
                UPDATE "Ekenox"."ordem_producao"
                   SET "data_fim" = %s,
                       situacao_id = 18162
                 WHERE "id" = %s
                   AND ("data_fim" IS NULL OR "data_fim" = '1970-01-01');
            """
            self._q(sql, (hoje, int(ordem_id)))
            if (self.cursor.rowcount or 0) == 0:
                self.conn.rollback()
                return False
            self.conn.commit()
            return True
        except Exception:
            if self.conn:
                self.conn.rollback()
            return False

    def buscar_estoque_maximo(self, fkproduto: int) -> float:
        try:
            sql = """
                SELECT ip."estoqueMaximo"
                  FROM "Ekenox"."infoProduto" ip
                 WHERE ip."fkProduto"::bigint = %s
                 LIMIT 1;
            """
            self._q(sql, (int(fkproduto),))
            r = self.cursor.fetchone()
            return float(r[0]) if r and r[0] is not None else 0.0
        except Exception:
            if self.conn:
                self.conn.rollback()
            return 0.0

    def f7_buscar_estrutura(self, fkproduto: int):
        sql = """
            SELECT e."componente", e."quantidade"
              FROM "Ekenox"."estrutura" e
             WHERE e."fkproduto"::bigint = %s
             ORDER BY e."componente";
        """
        self._q(sql, (int(fkproduto),))
        return self.cursor.fetchall() or []

    def f7_buscar_info_produto(self, fkproduto: int) -> Dict[str, Any]:
        sql = """
            SELECT
                i."estoqueMinimo",
                i."estoqueMaximo",
                i."precoCompra",
                i."fkFornecedor",
                i."fkProduto"
              FROM "Ekenox"."infoProduto" i
             WHERE i."fkProduto"::bigint = %s;
        """
        self._q(sql, (int(fkproduto),))
        r = self.cursor.fetchone()
        if not r:
            return {}
        return {
            "estoqueMinimo": float(r[0] or 0),
            "estoqueMaximo": float(r[1] or 0),
            "precoCompra": float(r[2] or 0),
            "fkFornecedor": int(r[3] or 0) if r[3] is not None else 0,
            "fkProduto": int(r[4] or fkproduto),
        }

    def f7_buscar_fornecedor(self, id_fornecedor: int) -> Dict[str, Any]:
        if not id_fornecedor:
            return {}
        sql = """
            SELECT f."idFornecedor", f."nome", f."codigo", f."telefone", f."celular"
              FROM "Ekenox"."fornecedor" f
             WHERE f."idFornecedor" = %s;
        """
        self._q(sql, (int(id_fornecedor),))
        r = self.cursor.fetchone()
        if not r:
            return {}
        return {"idFornecedor": r[0], "nome": r[1], "codigo": r[2], "telefone": r[3], "celular": r[4]}

    def relatorio_bling_insumos_produto(self, produto_id: int, qtd_produzir: float):
        sql = """
            SELECT
                COALESCE(pcomp."nomeProduto", '')                         AS descricao,
                pcomp."produtoId"::bigint                                 AS codigo,
                CASE
                  WHEN ipcomp."precoCompra" IS NOT NULL AND ipcomp."precoCompra" > 0
                    THEN ipcomp."precoCompra"::numeric
                  WHEN pcomp."custo" IS NOT NULL AND pcomp."custo" > 0
                    THEN pcomp."custo"::numeric
                  ELSE 0::numeric
                END                                                      AS valor_unit,
                COALESCE(ipcomp."unidadeMedida", ipcomp."unidade", '')    AS unidade,
                COALESCE(e."quantidade", 0)::numeric                      AS qtde_un,
                (COALESCE(e."quantidade", 0) * %s)::numeric               AS qtde_total
            FROM "Ekenox"."estrutura" e
            LEFT JOIN "Ekenox"."produtos" pcomp
              ON pcomp."produtoId"::bigint = e."componente"::bigint
            LEFT JOIN "Ekenox"."infoProduto" ipcomp
              ON ipcomp."fkProduto"::bigint = e."componente"::bigint
            WHERE e."fkproduto"::bigint = %s
            ORDER BY descricao;
        """
        self._q(sql, (float(qtd_produzir), int(produto_id)))
        return self.cursor.fetchall() or []

    def validar_estoque_insumos_para_producao(
        self,
        fkproduto: int,
        qtd_produzir: float,
        bloquear_se_saldo_negativo: bool = True,
        bloquear_se_insuficiente: bool = True,
    ) -> Dict[str, Any]:
        problemas: List[Dict[str, Any]] = []
        try:
            itens = self.f7_buscar_estrutura(int(fkproduto))
            if not itens:
                return {"ok": True, "problemas": []}

            for (componente, qtd_base) in itens:
                comp_id = int(componente)
                qtd_base_f = float(qtd_base or 0.0)
                necessario = qtd_base_f * float(qtd_produzir)

                saldo = float(self.saldo_fisico(comp_id) or 0.0)
                prod_comp = self.validar_produto(comp_id) or {}
                nome = (prod_comp.get("nomeproduto") or "").strip()
                falta = max(0.0, necessario - saldo)

                if bloquear_se_saldo_negativo and saldo < 0:
                    problemas.append({
                        "componente": comp_id,
                        "nome": nome,
                        "qtd_base": qtd_base_f,
                        "necessario": necessario,
                        "saldo": saldo,
                        "falta": falta,
                        "motivo": "Saldo negativo",
                    })
                    continue

                if bloquear_se_insuficiente and saldo < necessario:
                    problemas.append({
                        "componente": comp_id,
                        "nome": nome,
                        "qtd_base": qtd_base_f,
                        "necessario": necessario,
                        "saldo": saldo,
                        "falta": falta,
                        "motivo": "Saldo insuficiente",
                    })

            return {"ok": (len(problemas) == 0), "problemas": problemas}
        except Exception:
            if self.conn:
                self.conn.rollback()
            return {"ok": False, "problemas": [{"motivo": "Erro ao validar estrutura/estoque"}]}


# ============================================================
# ETIQUETAS (F12)
# ============================================================

class EtiquetasModule:
    def __init__(self, master: tk.Tk):
        self.master = master
        self.win: Optional[tk.Toplevel] = None

    def open(self, event=None):
        if self.win is not None and self.win.winfo_exists():
            try:
                self.win.deiconify()
                self.win.lift()
                self.win.focus_force()
            except Exception:
                pass
            return

        self.win = tk.Toplevel(self.master)
        apply_window_icon(self.win)
        self.win.title("Gerador de Etiquetas EKENOX")
        self.win.geometry("680x720")
        self.win.transient(self.master)
        self.win.grab_set()
        self.win.bind("<Escape>", lambda e: self.close())
        self._montar_interface(self.win)

    def close(self):
        try:
            if self.win and self.win.winfo_exists():
                self.win.destroy()
        except Exception:
            pass
        self.win = None

    def gerar_etiquetas(self):
        if not self.win or not self.win.winfo_exists():
            return
        try:
            empresa = {
                "company_name": self.entry_empresa.get().strip(),
                "company_address": self.entry_endereco.get().strip(),
                "company_district": self.entry_bairro.get().strip(),
                "company_city": self.entry_cidade.get().strip(),
                "company_state": self.entry_estado.get().strip(),
                "company_cep": self.entry_cep.get().strip(),
                "company_phone": self.entry_telefone.get().strip(),
                "company_email": self.entry_email.get().strip(),
            }

            produto = {
                "product_title": self.entry_produto.get().strip(),
                "product_model": self.entry_modelo.get().strip(),
                "product_classe": self.entry_classe.get().strip(),
                "voltage": self.combo_tensao.get().strip(),
                "power": self.combo_potencia.get().strip(),
                "temperature": self.entry_temperatura.get().strip(),
                "frequency": self.entry_frequencia.get().strip(),
            }

            if not produto["product_title"]:
                messagebox.showerror(
                    "Erro", "O campo 'Produto' deve ser preenchido.", parent=self.win)
                return

            try:
                quantidade = int(self.entry_quantidade.get().strip())
                if quantidade <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror(
                    "Erro", "A quantidade deve ser um inteiro > 0.", parent=self.win)
                return

            serie_base = self.entry_serie.get().strip()
            if not serie_base:
                messagebox.showerror(
                    "Erro", "Preencha o 'Número de Série (prefixo/base)'.", parent=self.win)
                return

            pdf_path = os.path.join(BASE_DIR, "etiquetas.pdf")
            largura, altura = 100 * mm, 75 * mm
            c = canvas.Canvas(pdf_path, pagesize=(largura, altura))

            x_titulo = 10
            x_valor = 70
            espaco = 10

            for i in range(quantidade):
                serial = f"{serie_base}-{i+1:03d}"

                c.setLineWidth(1)
                c.rect(5, 5, largura - 10, altura - 10)

                c.setFont("Helvetica-Bold", 12)
                c.drawCentredString(largura / 2, altura -
                                    15, empresa["company_name"])

                y = altura - 30

                campos_empresa = [
                    ("Endereço:", empresa["company_address"]),
                    ("Bairro:", empresa["company_district"]),
                    ("Cidade:",
                     f"{empresa['company_city']} - {empresa['company_state']}"),
                    ("CEP:", empresa["company_cep"]),
                    ("Telefone:", empresa["company_phone"]),
                    ("Email SAC:", empresa["company_email"]),
                ]

                for titulo, valor in campos_empresa:
                    c.setFont("Helvetica-Bold", 9)
                    c.drawString(x_titulo, y, titulo)
                    c.setFont("Helvetica", 9)
                    c.drawString(x_valor, y, valor)
                    y -= espaco

                c.line(x_titulo, y, largura - 10, y)
                y -= espaco

                produto_campos = [
                    ("Produto:", produto["product_title"]),
                    ("Modelo:", produto["product_model"]),
                    ("Classe:", produto["product_classe"]),
                    ("Tensão:", produto["voltage"]),
                    ("Potência:", produto["power"]),
                    ("Temp:", produto["temperature"]),
                    ("Freq:", produto["frequency"]),
                ]

                for titulo, valor in produto_campos:
                    c.setFont("Helvetica-Bold", 9)
                    c.drawString(x_titulo, y, titulo)
                    c.setFont("Helvetica", 9)
                    c.drawString(x_valor, y, valor)
                    y -= espaco

                c.line(x_titulo, y, largura - 10, y)
                y -= espaco * 2

                c.setFont("Helvetica-Bold", 12)
                c.drawCentredString(largura / 2, y, f"Nº Série: {serial}")

                c.showPage()

            c.save()

            messagebox.showinfo(
                "Sucesso", f"PDF gerado:\n{pdf_path}", parent=self.win)
            try:
                if os.name == "nt":
                    os.startfile(pdf_path)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror(
                "Erro", f"{type(e).__name__}: {e}", parent=self.win)

    def _montar_interface(self, root):
        frame_empresa = tk.LabelFrame(
            root, text="Informações da Empresa", padx=10, pady=10)
        frame_empresa.pack(fill="both", padx=10, pady=5)

        def add_row(r, label, default=""):
            tk.Label(frame_empresa, text=label).grid(
                row=r, column=0, sticky="e")
            e = tk.Entry(frame_empresa, width=50)
            if default:
                e.insert(0, default)
            e.grid(row=r, column=1, pady=2)
            return e

        self.entry_empresa = add_row(
            0, "Nome da Empresa:", "EKENOX DISTRIBUIDORA DE COZ. IND. LTDA")
        self.entry_endereco = add_row(
            1, "Endereço:", "Rua: José de Ribamar Souza, 499")
        self.entry_bairro = add_row(2, "Bairro:", "Pq. Industrial")
        self.entry_cidade = add_row(3, "Cidade:", "Catanduva")
        self.entry_estado = add_row(4, "Estado:", "SP")
        self.entry_cep = add_row(5, "CEP:", "15803-290")
        self.entry_telefone = add_row(6, "Telefone:", "(11)98740-3669")
        self.entry_email = add_row(7, "Email SAC:", "sac@ekenox.com.br")

        frame_prod = tk.LabelFrame(
            root, text="Informações do Produto", padx=10, pady=10)
        frame_prod.pack(fill="both", padx=10, pady=5)

        tk.Label(frame_prod, text="Produto:").grid(row=0, column=0, sticky="e")
        self.entry_produto = tk.Entry(frame_prod, width=50)
        self.entry_produto.insert(0, "BUFFET TÉRMICO")
        self.entry_produto.grid(row=0, column=1, pady=2,
                                sticky="w", columnspan=2)

        tk.Label(frame_prod, text="Classe:").grid(row=1, column=0, sticky="e")
        self.entry_classe = tk.Entry(frame_prod, width=50)
        self.entry_classe.insert(0, "IPX4")
        self.entry_classe.grid(
            row=1, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Modelo (SKU):").grid(
            row=2, column=0, sticky="e")
        self.entry_modelo = tk.Entry(frame_prod, width=50)
        self.entry_modelo.insert(0, "VIX8368")
        self.entry_modelo.grid(
            row=2, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Tensão:").grid(row=3, column=0, sticky="e")
        self.combo_tensao = ttk.Combobox(
            frame_prod, values=["127V", "220V"], state="readonly", width=47)
        self.combo_tensao.grid(
            row=3, column=1, columnspan=2, pady=2, sticky="w")
        self.combo_tensao.set("127V")

        tk.Label(frame_prod, text="Potência:").grid(
            row=4, column=0, sticky="e")
        self.combo_potencia = ttk.Combobox(
            frame_prod, values=["1000W", "2000W"], state="readonly", width=47)
        self.combo_potencia.grid(
            row=4, column=1, columnspan=2, pady=2, sticky="w")
        self.combo_potencia.set("2000W")

        tk.Label(frame_prod, text="Temperatura:").grid(
            row=5, column=0, sticky="e")
        self.entry_temperatura = tk.Entry(frame_prod, width=50)
        self.entry_temperatura.insert(0, "30°C a 120°C")
        self.entry_temperatura.grid(
            row=5, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Frequência:").grid(
            row=6, column=0, sticky="e")
        self.entry_frequencia = tk.Entry(frame_prod, width=50)
        self.entry_frequencia.insert(0, "60Hz")
        self.entry_frequencia.grid(
            row=6, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Número de Série (prefixo/base):").grid(row=7,
                                                                          column=0, sticky="e")
        self.entry_serie = tk.Entry(frame_prod, width=50)
        self.entry_serie.insert(0, "EKX2024")
        self.entry_serie.grid(
            row=7, column=1, columnspan=2, pady=2, sticky="w")

        tk.Label(frame_prod, text="Quantidade de etiquetas:").grid(
            row=8, column=0, sticky="e")
        self.entry_quantidade = tk.Entry(frame_prod, width=50)
        self.entry_quantidade.insert(0, "5")
        self.entry_quantidade.grid(
            row=8, column=1, columnspan=2, pady=2, sticky="w")

        frame_btn = tk.Frame(root, pady=10)
        frame_btn.pack(fill="x")

        tk.Button(
            frame_btn, text="Gerar PDF", command=self.gerar_etiquetas,
            bg="#2563eb", fg="white", font=("Arial", 12, "bold"), width=15
        ).pack(side="left", padx=(40, 10))

        tk.Button(
            frame_btn, text="Fechar", command=self.close,
            bg="#ef4444", fg="white", font=("Arial", 12, "bold"), width=15
        ).pack(side="left")


# ============================================================
# APP PRINCIPAL
# ============================================================

class OrdemProducaoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        apply_window_icon(self)

        self._closing = False
        self.cfg = load_config()
        self.sistema = SistemaOrdemProducao(self.cfg)
        self.connected = self.sistema.conectar()

        self.title("Sistema de Ordem de Produção - Ekenox")
        self.geometry("1150x650")
        self.minsize(1150, 650)

        # ✅ CORREÇÃO PRINCIPAL: método existe e está dentro da classe
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        self.mod_etiquetas = EtiquetasModule(self)

        if not self.connected:
            messagebox.showerror(
                "Banco",
                f"Falha ao conectar:\n{self.sistema.ultimo_erro or 'Erro desconhecido'}",
                parent=self,
            )
            self.after(0, self.destroy)
            return

        self.variaveis_quantidade: Optional[Dict[str, Any]] = None

        self._build_ui()

        try:
            self.numero_var.set(str(self.sistema.gerar_numero_ordem()))
        except Exception:
            pass

    # ---------------- helpers ----------------

    def parse_int(self, valor_str: str, campo: str) -> int:
        if not valor_str.strip():
            raise ValueError(f"O campo '{campo}' é obrigatório.")
        return int(valor_str.strip())

    def parse_float(self, valor_str: str, campo: str, obrigatorio: bool = False) -> Optional[float]:
        if not valor_str.strip():
            if obrigatorio:
                raise ValueError(f"O campo '{campo}' é obrigatório.")
            return None
        return float(valor_str.replace(",", "."))

    # ---------------- UI ----------------

    def _build_ui(self):
        try:
            main_frame = ttk.Frame(self, padding=10)
            main_frame.pack(fill=tk.BOTH, expand=True)

            status_frame = ttk.Frame(main_frame)
            status_frame.pack(fill=tk.X)

            self.status_label = ttk.Label(
                status_frame,
                text="Conectado ao banco de dados" if self.connected else "Erro ao conectar ao banco",
                foreground=("green" if self.connected else "red"),
                font=("Segoe UI", 10, "bold"),
            )
            self.status_label.pack(side=tk.LEFT)

            ttk.Separator(main_frame, orient=tk.HORIZONTAL).pack(
                fill=tk.X, pady=10)

            form_frame = ttk.LabelFrame(
                main_frame, text="Nova Ordem de Produção")
            form_frame.pack(fill=tk.BOTH, expand=True)

            for col in range(6):
                form_frame.columnconfigure(
                    col, weight=1 if col in (1, 4) else 0)
            form_frame.rowconfigure(6, weight=1)

            # vars
            self.numero_var = tk.StringVar()
            self.deposito_origem_var = tk.StringVar()
            self.deposito_destino_var = tk.StringVar()
            self.situacao_id_var = tk.StringVar()
            self.produto_id_var = tk.StringVar()
            self.responsavel_var = tk.StringVar()
            self.quantidade_var = tk.StringVar()
            self.valor_var = tk.StringVar()

            # linha 0
            ttk.Label(form_frame, text="Número da Ordem:*").grid(row=0,
                                                                 column=0, sticky="e", padx=5, pady=5)
            num_frame = ttk.Frame(form_frame)
            num_frame.grid(row=0, column=1, sticky="w", padx=(0, 10), pady=5)

            self.numero_entry = ttk.Entry(
                num_frame, textvariable=self.numero_var, width=15)
            self.numero_entry.grid(row=0, column=0, sticky="w")

            ttk.Button(num_frame, text="Buscar (F5)", width=12, command=self.buscar_ordem_por_numero).grid(
                row=0, column=1, padx=(6, 0), sticky="w"
            )
            self.numero_entry.bind("<Return>", self.buscar_ordem_por_numero)

            ttk.Label(form_frame, text="Responsável:").grid(
                row=0, column=3, sticky="e", padx=5, pady=5)
            ttk.Entry(form_frame, textvariable=self.responsavel_var, width=25).grid(
                row=0, column=4, sticky="ew", padx=(0, 10), pady=5
            )

            # linha 1 depósitos
            ttk.Label(form_frame, text="Depósito Origem (ID):*").grid(row=1,
                                                                      column=0, sticky="e", padx=5, pady=5)
            origem_frame = ttk.Frame(form_frame)
            origem_frame.grid(row=1, column=1, columnspan=2,
                              sticky="w", padx=(0, 10), pady=5)

            self.deposito_origem_entry = ttk.Entry(
                origem_frame, textvariable=self.deposito_origem_var, width=18)
            self.deposito_origem_entry.grid(row=0, column=0, sticky="w")
            ttk.Button(origem_frame, text="Listar (F4)", command=self.mostrar_depositos_origem).grid(
                row=0, column=1, padx=(5, 0)
            )

            ttk.Label(form_frame, text="Depósito Destino (ID):*").grid(row=1,
                                                                       column=3, sticky="e", padx=5, pady=5)
            destino_frame = ttk.Frame(form_frame)
            destino_frame.grid(row=1, column=4, columnspan=2,
                               sticky="w", padx=(0, 10), pady=5)

            self.deposito_destino_entry = ttk.Entry(
                destino_frame, textvariable=self.deposito_destino_var, width=18)
            self.deposito_destino_entry.grid(row=0, column=0, sticky="w")
            ttk.Button(destino_frame, text="Listar (F8)", command=self.mostrar_depositos_destino).grid(
                row=0, column=1, padx=(5, 0)
            )

            # linha 2 situação / produto
            ttk.Label(form_frame, text="Situação (ID):*").grid(row=2,
                                                               column=0, sticky="e", padx=5, pady=5)
            situacao_frame = ttk.Frame(form_frame)
            situacao_frame.grid(row=2, column=1, columnspan=2,
                                sticky="w", padx=(0, 10), pady=5)

            self.situacao_entry = ttk.Entry(
                situacao_frame, textvariable=self.situacao_id_var, width=18)
            self.situacao_entry.grid(row=0, column=0, sticky="w")
            ttk.Button(situacao_frame, text="Listar (F3)", command=self.mostrar_situacoes).grid(
                row=0, column=1, padx=(5, 0)
            )

            ttk.Label(form_frame, text="Produto (ID):*").grid(row=2,
                                                              column=3, sticky="e", padx=5, pady=5)
            prod_frame = ttk.Frame(form_frame)
            prod_frame.grid(row=2, column=4, columnspan=2,
                            sticky="w", padx=(0, 10), pady=5)

            self.produto_entry = ttk.Entry(
                prod_frame, textvariable=self.produto_id_var, width=18)
            self.produto_entry.grid(row=0, column=0, sticky="w")
            ttk.Button(prod_frame, text="Listar (F2)", command=self.mostrar_produtos).grid(
                row=0, column=1, padx=(5, 0)
            )
            self.produto_entry.bind(
                "<Return>", self.atualizar_quantidade_producao)
            self.produto_entry.bind(
                "<FocusOut>", self.atualizar_quantidade_producao)

            # linha 3 quantidade
            ttk.Label(form_frame, text="Quantidade:*").grid(row=3,
                                                            column=0, sticky="e", padx=5, pady=5)
            qtd_frame = ttk.Frame(form_frame)
            qtd_frame.grid(row=3, column=1, columnspan=5,
                           sticky="w", padx=(0, 10), pady=5)

            self.quantidade_entry = ttk.Entry(
                qtd_frame, textvariable=self.quantidade_var, width=18)
            self.quantidade_entry.grid(
                row=0, column=0, sticky="w", padx=(0, 6))

            ttk.Button(qtd_frame, text="Detalhes (F6)", command=self.mostrar_detalhes_quantidade).grid(
                row=0, column=1, padx=(0, 6)
            )
            ttk.Button(qtd_frame, text="Analisar Estrutura (F7)", command=self.analisar_estrutura_f7).grid(
                row=0, column=2
            )

            # linha 6 observação
            ttk.Label(form_frame, text="Observação:").grid(
                row=6, column=0, sticky="ne", padx=5, pady=5)
            self.observacao_text = tk.Text(form_frame, height=6)
            self.observacao_text.grid(
                row=6, column=1, columnspan=5, sticky="nsew", padx=(0, 10), pady=5)

            # botões
            botoes = ttk.Frame(form_frame)
            botoes.grid(row=7, column=0, columnspan=6, pady=15)

            ttk.Button(botoes, text="Salvar Ordem",
                       command=self.salvar_ordem).pack(side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Limpar", command=self.limpar_formulario).pack(
                side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Relatório (F9)", command=self.exportar_relatorio_bling_excel_f9).pack(
                side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Ordens (F10)", command=self.mostrar_ordens_producao).pack(
                side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Finalizar (F11)",
                       command=self.finalizar_producoes_pendentes).pack(side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Etiquetas (F12)",
                       command=self.mod_etiquetas.open).pack(side=tk.LEFT, padx=5)

            ttk.Button(botoes, text="Voltar ao Menu",
                       command=self.voltar_menu).pack(side=tk.LEFT, padx=5)
            ttk.Button(botoes, text="Fechar", command=self.on_close).pack(
                side=tk.LEFT, padx=5)

            ttk.Label(main_frame, text="Campos marcados com * são obrigatórios.", foreground="gray").pack(
                side=tk.BOTTOM, anchor="w", pady=(5, 0)
            )

            # binds globais
            for seq, func in [
                ("<F2>",  self.mostrar_produtos),
                ("<F3>",  self.mostrar_situacoes),
                ("<F4>",  self.mostrar_depositos_origem),
                ("<F5>",  self.buscar_ordem_por_numero),
                ("<F6>",  self.mostrar_detalhes_quantidade),
                ("<F7>",  self.analisar_estrutura_f7),
                ("<F8>",  self.mostrar_depositos_destino),
                ("<F9>",  self.exportar_relatorio_bling_excel_f9),
                ("<F10>", self.mostrar_ordens_producao),
                ("<F11>", self.finalizar_producoes_pendentes),
                ("<F12>", self.mod_etiquetas.open),
                ("<Escape>", lambda e: self.on_close()),
            ]:
                self.bind_all(seq, func)

        except Exception as e:
            log_path = log_exception(e, "Falha ao montar UI")
            messagebox.showerror(
                "Erro", f"Falha ao montar interface.\nLog: {log_path}", parent=self)

    # ============================================================
    # F5 - Buscar OP
    # ============================================================

    def buscar_ordem_por_numero(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "Buscar OP", "Sem conexão com o banco.", parent=self)
            return

        num_txt = (self.numero_var.get() or "").strip()
        if not num_txt:
            messagebox.showwarning(
                "Buscar OP", "Informe o Número da Ordem.", parent=self)
            return

        op = self.sistema.buscar_ordem_producao_por_numero(num_txt)
        if not op:
            messagebox.showinfo(
                "Buscar OP", f"OP nº {num_txt} não encontrada.", parent=self)
            return

        self.numero_var.set(str(op.get("numero") or ""))
        self.deposito_destino_var.set(str(op.get("deposito_destino") or ""))
        self.deposito_origem_var.set(str(op.get("deposito_origem") or ""))
        self.situacao_id_var.set(str(op.get("situacao_id") or ""))
        self.produto_id_var.set(str(op.get("fkprodutoid") or ""))
        self.responsavel_var.set(str(op.get("responsavel") or ""))

        qtd = op.get("quantidade")
        self.quantidade_var.set(f"{float(qtd):.2f}" if qtd is not None else "")

        val = op.get("valor")
        self.valor_var.set(f"{float(val):.2f}" if val is not None else "")

        obs = (op.get("observacao") or "").strip()
        self.observacao_text.delete("1.0", tk.END)
        if obs:
            self.observacao_text.insert("1.0", obs)

        messagebox.showinfo(
            "Buscar OP", f"OP nº {op.get('numero')} carregada com sucesso.", parent=self)

    # ============================================================
    # Atualizar quantidade (auto)
    # ============================================================

    def atualizar_quantidade_producao(self, event=None):
        try:
            pid_str = (self.produto_id_var.get() or "").strip()
            if not pid_str:
                self.quantidade_var.set("")
                self.variaveis_quantidade = None
                return

            pid = int(pid_str)
            produto = self.sistema.validar_produto(pid)
            if not produto:
                self.quantidade_var.set("")
                self.variaveis_quantidade = {"erro": "Produto não encontrado"}
                return

            preco = float(produto.get("preco") or 0.0)
            self.valor_var.set(f"{preco:.2f}")

            saldo = float(self.sistema.saldo_fisico(pid) or 0.0)
            media_vendas = float(self.sistema.media_vendas_mensal(pid) or 0.0)
            estoque_max = float(self.sistema.buscar_estoque_maximo(pid) or 0.0)

            dia = max(1, int(date.today().day or 1))
            media_dia = (media_vendas / dia) if media_vendas > 0 else 1.0

            multiplicador = 7.0
            producao_media = media_dia * multiplicador
            sugestao_calc = max(0.0, estoque_max - producao_media)

            sku = (produto.get("sku") or "").strip()
            qtd_arranjo = float(
                self.sistema.buscar_qtd_produzir_por_sku(sku) or 0.0)

            if sugestao_calc <= 0:
                sugestao_final = 0.0
            elif qtd_arranjo > 0:
                sugestao_final = ceil(
                    sugestao_calc / qtd_arranjo) * qtd_arranjo
            else:
                sugestao_final = sugestao_calc

            self.quantidade_var.set(f"{sugestao_final:.2f}")

            self.variaveis_quantidade = {
                "Produto id": pid,
                "Produto nome": produto.get("nomeproduto"),
                "SKU": produto.get("sku"),
                "Preço": preco,
                "Saldo": saldo,
                "Média vendas mês": media_vendas,
                "Dia atual": dia,
                "Média/dia": media_dia,
                "Estoque máximo (infoProduto)": estoque_max,
                "Sugestão calculada": sugestao_calc,
                "Arranjo (lote)": qtd_arranjo,
                "Sugestão final (múltiplo arranjo)": sugestao_final,
                "Obs": "Sugestão = estoqueMax - (média/dia * 7), arredondando para cima no múltiplo do arranjo.",
            }
        except Exception as e:
            self.variaveis_quantidade = {"erro": f"{type(e).__name__}: {e}"}

    def mostrar_detalhes_quantidade(self, event=None):
        if not self.variaveis_quantidade:
            messagebox.showinfo(
                "Detalhes", "Nenhum cálculo realizado ainda.", parent=self)
            return

        win = tk.Toplevel(self)
        apply_window_icon(win)
        win.title("Detalhes do cálculo da quantidade")
        win.geometry("650x520")
        win.transient(self)
        win.grab_set()

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        txt = tk.Text(frame, wrap="word")
        txt.pack(fill=tk.BOTH, expand=True)

        for k, v in self.variaveis_quantidade.items():
            txt.insert(tk.END, f"{k}: {v}\n")

        txt.config(state="disabled")
        ttk.Button(frame, text="Fechar", command=win.destroy).pack(pady=5)
        win.bind("<Escape>", lambda e: win.destroy())

    # ============================================================
    # F7 - Estrutura + pedido compra
    # ============================================================

    def analisar_estrutura_f7(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "F7 - Estrutura", "Sem conexão com o banco.", parent=self)
            return

        prod = (self.produto_id_var.get() or "").strip()
        qtd_txt = (self.quantidade_var.get() or "").strip()

        if not prod:
            messagebox.showerror(
                "F7 - Estrutura", "Informe o Produto (ID).", parent=self)
            return

        try:
            produto_id = int(prod)
        except ValueError:
            messagebox.showerror(
                "F7 - Estrutura", "Produto (ID) inválido.", parent=self)
            return

        try:
            qtd_produzir = float(qtd_txt.replace(",", ".")) if qtd_txt else 0.0
        except Exception:
            qtd_produzir = 0.0

        if qtd_produzir <= 0:
            messagebox.showerror(
                "F7 - Estrutura", "Quantidade para produzir deve ser > 0.", parent=self)
            return

        try:
            itens = self.sistema.f7_buscar_estrutura(produto_id)
        except Exception as e:
            messagebox.showerror(
                "F7 - Estrutura", f"Erro ao ler estrutura:\n{e}", parent=self)
            return

        if not itens:
            messagebox.showinfo(
                "F7 - Estrutura", "Sem estrutura cadastrada para este produto.", parent=self)
            return

        linhas = []
        faltantes = 0
        itens_faltantes_para_pedido: List[Dict[str, Any]] = []

        for (componente, qtd_base) in itens:
            qtd_necessaria = float(qtd_base or 0.0) * float(qtd_produzir)
            saldo = float(self.sistema.saldo_fisico(int(componente)) or 0.0)
            falta = max(0.0, qtd_necessaria - saldo)

            info = self.sistema.f7_buscar_info_produto(int(componente))
            est_min = float(info.get("estoqueMinimo", 0.0) or 0.0)
            est_max = float(info.get("estoqueMaximo", 0.0) or 0.0)
            preco_compra = float(info.get("precoCompra", 0.0) or 0.0)

            fornecedor_nome = ""
            fk_fornecedor = int(info.get("fkFornecedor", 0) or 0)
            if fk_fornecedor:
                forn = self.sistema.f7_buscar_fornecedor(fk_fornecedor)
                fornecedor_nome = (forn.get("nome") or "").strip()

            prod_comp = self.sistema.validar_produto(int(componente))
            comp_nome = (prod_comp.get("nomeproduto")
                         if prod_comp else "") or ""

            if falta > 0:
                faltantes += 1
                qtd_comprar = ceil(falta)

                itens_faltantes_para_pedido.append({
                    "fornecedor": fornecedor_nome or "SEM FORNECEDOR",
                    "descricao": f"{componente} - {comp_nome}".strip(" -"),
                    "qtd_comprar": float(qtd_comprar),
                    "estoque_atual": float(saldo),
                    "estoque_minimo": float(est_min),
                    "estoque_maximo": float(est_max),
                    "valor_unitario": float(preco_compra) if preco_compra > 0 else None,
                })

            linhas.append((
                int(componente),
                comp_nome,
                float(qtd_base or 0.0),
                float(qtd_necessaria),
                float(saldo),
                float(falta),
                float(est_min),
                float(est_max),
                fornecedor_nome,
                float(preco_compra),
            ))

        win = tk.Toplevel(self)
        apply_window_icon(win)
        win.title(f"F7 - Estrutura | Produto {produto_id}")
        win.geometry(self.cfg.f7_geometry)
        win.minsize(1050, 520)
        win.transient(self)
        win.grab_set()

        top = ttk.Frame(win, padding=10)
        top.pack(fill=tk.X)

        ttk.Label(
            top,
            text=f"Produto: {produto_id} | Produzir: {qtd_produzir:.2f} | Itens: {len(linhas)} | Faltando: {faltantes}",
            font=("Segoe UI", 10, "bold"),
        ).pack(side=tk.LEFT)

        def gerar_pedido():
            if not itens_faltantes_para_pedido:
                messagebox.showinfo(
                    "Pedido de Compra", "Não há faltantes para gerar pedido.", parent=win)
                return
            if not os.path.exists(self.cfg.caminho_modelo):
                messagebox.showerror(
                    "Pedido de Compra", f"Modelo não encontrado:\n{self.cfg.caminho_modelo}", parent=win)
                return

            numero_inicial = simpledialog.askinteger(
                "Pedido de Compra",
                "Informe o número inicial do pedido:",
                parent=win,
                minvalue=1
            )
            if not numero_inicial:
                return

            dados_excel = []
            hoje = date.today()
            numero_atual = int(numero_inicial)

            grupos = defaultdict(list)
            for it in itens_faltantes_para_pedido:
                grupos[it["fornecedor"]].append(it)

            for fornecedor, itens_f in grupos.items():
                for it in itens_f:
                    dados_excel.append({
                        "fornecedor": fornecedor,
                        "numero_pedido": numero_atual,
                        "data_pedido": hoje,
                        "produto": it["descricao"],
                        "quantidade": float(it["qtd_comprar"]),
                        "estoque_atual": float(it.get("estoque_atual", 0.0) or 0.0),
                        "estoque_minimo": float(it.get("estoque_minimo", 0.0) or 0.0),
                        "estoque_maximo": float(it.get("estoque_maximo", 0.0) or 0.0),
                        "valor_unitario": it.get("valor_unitario"),
                    })
                numero_atual += 1

            try:
                gerar_abas_fornecedor_pedido(
                    dados=dados_excel,
                    nome_aba_modelo="Pedido de Compra",
                    caminho_modelo=self.cfg.caminho_modelo,
                    caminho_saida=self.cfg.caminho_saida,
                )
                messagebox.showinfo(
                    "Pedido de Compra", f"Gerado em:\n{self.cfg.caminho_saida}", parent=win)
                try:
                    if os.name == "nt":
                        os.startfile(self.cfg.caminho_saida)
                except Exception:
                    pass
            except Exception as e:
                messagebox.showerror("Pedido de Compra",
                                     f"Falha ao gerar Excel:\n{e}", parent=win)

        ttk.Button(top, text="Gerar Pedido (faltantes)",
                   command=gerar_pedido).pack(side=tk.RIGHT)

        frame = ttk.Frame(win, padding=(10, 0, 10, 10))
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("comp", "nome", "base", "necessaria", "saldo",
                "falta", "min", "max", "forn", "preco")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        headings = {
            "comp": "Componente (ID)",
            "nome": "Nome",
            "base": "Qtd Estrutura",
            "necessaria": "Qtd Necessária",
            "saldo": "Saldo",
            "falta": "Falta",
            "min": "Est. Mín",
            "max": "Est. Máx",
            "forn": "Fornecedor",
            "preco": "Preço Compra",
        }
        for c in cols:
            tree.heading(c, text=headings[c])

        tree.column("comp", width=130, anchor="w")
        tree.column("nome", width=260, anchor="w")
        tree.column("base", width=120, anchor="e")
        tree.column("necessaria", width=130, anchor="e")
        tree.column("saldo", width=110, anchor="e")
        tree.column("falta", width=110, anchor="e")
        tree.column("min", width=90, anchor="e")
        tree.column("max", width=90, anchor="e")
        tree.column("forn", width=210, anchor="w")
        tree.column("preco", width=110, anchor="e")

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree.tag_configure("faltando", foreground="red")

        for (comp, nome, base, nec, saldo, falta, est_min, est_max, forn, preco) in linhas:
            vals = (
                str(comp),
                nome or "",
                f"{base:.4f}",
                f"{nec:.4f}",
                f"{saldo:.4f}",
                f"{falta:.4f}",
                f"{est_min:.2f}",
                f"{est_max:.2f}",
                forn or "",
                f"{preco:.2f}" if preco else "",
            )
            if falta > 0:
                tree.insert("", tk.END, values=vals, tags=("faltando",))
            else:
                tree.insert("", tk.END, values=vals)

        def close():
            try:
                self.cfg.f7_geometry = win.geometry()
                save_config(self.cfg)
            except Exception:
                pass
            win.destroy()

        win.bind("<Escape>", lambda e: close())
        win.protocol("WM_DELETE_WINDOW", close)

    # ============================================================
    # F9 - Relatório Excel
    # ============================================================

    def exportar_relatorio_bling_excel_f9(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "F9 - Relatório", "Sem conexão com o banco.", parent=self)
            return

        prod_txt = (self.produto_id_var.get() or "").strip()
        qtd_txt = (self.quantidade_var.get() or "").strip()

        if not prod_txt:
            messagebox.showerror(
                "F9 - Relatório", "Informe o Produto (ID).", parent=self)
            return

        try:
            produto_id = int(prod_txt)
        except ValueError:
            messagebox.showerror(
                "F9 - Relatório", "Produto (ID) inválido.", parent=self)
            return

        try:
            qtd_produzir = float(qtd_txt.replace(",", ".")) if qtd_txt else 0.0
        except Exception:
            qtd_produzir = 0.0

        if qtd_produzir <= 0:
            messagebox.showerror(
                "F9 - Relatório", "Quantidade para produzir deve ser > 0.", parent=self)
            return

        produto = self.sistema.validar_produto(produto_id) or {}
        if not produto:
            messagebox.showerror(
                "F9 - Relatório", "Produto não encontrado.", parent=self)
            return

        prod_nome = (produto.get("nomeproduto") or "").strip()
        prod_codigo = int(produto_id)

        try:
            insumos = self.sistema.relatorio_bling_insumos_produto(
                produto_id, qtd_produzir)
        except Exception as e:
            messagebox.showerror(
                "F9 - Relatório", f"Erro ao consultar estrutura:\n{e}", parent=self)
            return

        if not insumos:
            messagebox.showinfo(
                "F9 - Relatório", "Sem estrutura cadastrada para este produto.", parent=self)
            return

        def D(x) -> Decimal:
            if x is None:
                return Decimal("0")
            if isinstance(x, Decimal):
                return x
            try:
                return Decimal(str(x).replace(",", "."))
            except Exception:
                return Decimal("0")

        def money2(x: Decimal) -> Decimal:
            return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        qtd_prod_D = D(qtd_produzir)

        total_geral = Decimal("0")
        for (_, _, valor_unit, _un, _qt_un, qt_total) in insumos:
            v = D(valor_unit)
            qt_tot = D(qt_total)
            total_item = money2(v * qt_tot)
            total_geral += total_item

        unitario = money2(
            total_geral / qtd_prod_D) if qtd_prod_D > 0 else Decimal("0")

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        def auto_col_width(ws_, min_w=10, max_w=60):
            for col in ws_.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    v = cell.value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws_.column_dimensions[col_letter].width = max(
                    min_w, min(max_w, max_len + 2))

        blue = PatternFill("solid", fgColor="1F4E79")
        gray = PatternFill("solid", fgColor="F2F2F2")
        header_font = Font(bold=True, color="FFFFFF")
        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        qty_fmt = '#,##0.00'
        money_fmt = '"R$" #,##0.00'

        ws.merge_cells("A1:G1")
        ws["A1"] = "Relatório de Produção"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.append([""] * 7)
        ws.append(["Item para produção"] + [""] * 6)
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=7)
        ws.cell(row=ws.max_row, column=1).font = bold

        ws.append(["Descrição", "Código",
                  "Quantidade a produzir", "", "", "", ""])
        r = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

        ws.append([prod_nome, prod_codigo, float(
            qtd_produzir), "", "", "", ""])
        r = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c == 3:
                cell.number_format = qty_fmt
                cell.alignment = Alignment(
                    horizontal="right", vertical="center")

        ws.append([""] * 7)
        ws.append(["Insumos (matéria prima)"] + [""] * 6)
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=7)
        ws.cell(row=ws.max_row, column=1).font = bold

        headers = [
            "Descrição", "Código", "Valor do insumo", "Qtde un.",
            "Valor (Qtde un * compra)", "Qtde total", "Total insumo"
        ]
        ws.append(headers)

        header_row = ws.max_row
        for c in range(1, 8):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = blue
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

        start_data = ws.max_row + 1

        for (desc, cod, valor_unit, _un, qt_un, qt_total) in insumos:
            v = D(valor_unit)
            q_un = D(qt_un)
            q_tot = D(qt_total)

            valor_qt_un = money2(v * q_un)
            total_item = money2(v * q_tot)

            ws.append([desc or "", int(cod or 0), float(v), float(
                q_un), float(valor_qt_un), float(q_tot), float(total_item)])

            rr = ws.max_row
            for c in range(1, 8):
                cell = ws.cell(row=rr, column=c)
                cell.border = border
                if c in (3, 5, 7):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center")
                elif c in (4, 6):
                    cell.number_format = qty_fmt
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True)

        end_data = ws.max_row
        ws.auto_filter.ref = f"A{header_row}:G{end_data}"
        ws.freeze_panes = f"A{start_data}"

        ws.append(["", "", "", "", "", "TOTAL:", float(total_geral)])
        rr = ws.max_row
        for c in range(1, 8):
            ws.cell(row=rr, column=c).border = border
        ws.cell(row=rr, column=6).font = bold
        ws.cell(row=rr, column=7).font = bold
        ws.cell(row=rr, column=7).number_format = money_fmt
        ws.cell(row=rr, column=7).alignment = Alignment(
            horizontal="right", vertical="center")

        ws.append([""] * 7)
        ws.append(["Observação"] + [""] * 6)
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=7)
        ws.cell(row=ws.max_row, column=1).font = bold

        ws.append(["", "", "Total Unitário:", "", float(
            unitario), "Total Geral:", float(total_geral)])
        obs_row = ws.max_row
        ws.cell(row=obs_row, column=5).number_format = money_fmt
        ws.cell(row=obs_row, column=7).number_format = money_fmt

        for c in range(1, 8):
            cell = ws.cell(row=obs_row, column=c)
            cell.border = border
            cell.fill = gray
            cell.font = bold
            if c in (5, 7):
                cell.alignment = Alignment(
                    horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center")

        auto_col_width(ws)
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 14

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_bling_{produto_id}_{ts}.xlsx"
        caminho = os.path.join(BASE_DIR, nome_arquivo)
        wb.save(caminho)

        messagebox.showinfo(
            "F9 - Relatório", f"Relatório gerado em:\n{caminho}", parent=self)
        try:
            if os.name == "nt":
                os.startfile(caminho)
        except Exception:
            pass

    # ============================================================
    # Listas F2/F3/F4/F8
    # ============================================================

    def mostrar_produtos(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco.", parent=self)
            return

        produtos = self.sistema.listar_produtos_disponiveis()
        if not produtos:
            messagebox.showinfo(
                "Produtos", "Nenhum produto encontrado.", parent=self)
            return

        win = tk.Toplevel(self)
        apply_window_icon(win)
        win.title("Produtos - Duplo clique para selecionar")
        win.geometry("980x520")
        win.transient(self)
        win.grab_set()

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("ID", "Nome", "SKU", "Preço", "Tipo")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        for c in cols:
            tree.heading(c, text=c)

        tree.column("ID", width=120)
        tree.column("Nome", width=420)
        tree.column("SKU", width=170)
        tree.column("Preço", width=120, anchor="e")
        tree.column("Tipo", width=120)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        for p in produtos:
            preco = f"{float(p[3]):.2f}" if p[3] else "0.00"
            tree.insert("", tk.END, values=(p[0], p[1], p[2], preco, p[4]))

        def selecionar(_=None):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0])["values"]
            self.produto_id_var.set(str(vals[0]))
            win.destroy()
            self.atualizar_quantidade_producao()
            self.quantidade_entry.focus_set()

        tree.bind("<Double-Button-1>", selecionar)
        tree.bind("<Return>", selecionar)
        win.bind("<Escape>", lambda e: win.destroy())

    def mostrar_situacoes(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco.", parent=self)
            return

        situacoes = self.sistema.listar_situacoes_disponiveis()
        if not situacoes:
            messagebox.showinfo(
                "Situações", "Nenhuma situação encontrada.", parent=self)
            return

        win = tk.Toplevel(self)
        apply_window_icon(win)
        win.title("Situações - Duplo clique para selecionar")
        win.geometry("620x440")
        win.transient(self)
        win.grab_set()

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("ID", "Situação")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        tree.heading("ID", text="ID")
        tree.heading("Situação", text="Situação")
        tree.column("ID", width=140)
        tree.column("Situação", width=420)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        for s in situacoes:
            tree.insert("", tk.END, values=(s[0], s[1]))

        def selecionar(_=None):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0])["values"]
            self.situacao_id_var.set(str(vals[0]))
            win.destroy()

        tree.bind("<Double-Button-1>", selecionar)
        tree.bind("<Return>", selecionar)
        win.bind("<Escape>", lambda e: win.destroy())

    def mostrar_depositos_origem(self, event=None):
        self._mostrar_depositos("origem")

    def mostrar_depositos_destino(self, event=None):
        self._mostrar_depositos("destino")

    def _mostrar_depositos(self, modo: str):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco.", parent=self)
            return

        depositos = self.sistema.listar_depositos_disponiveis()
        if not depositos:
            messagebox.showinfo(
                "Depósitos", "Nenhum depósito encontrado.", parent=self)
            return

        win = tk.Toplevel(self)
        apply_window_icon(win)
        win.title("Depósitos - Duplo clique para selecionar")
        win.geometry("820x460")
        win.transient(self)
        win.grab_set()

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("ID", "Descrição", "Situação", "Padrão", "Desconsiderar saldo")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        for c in cols:
            tree.heading(c, text=c)

        tree.column("ID", width=140)
        tree.column("Descrição", width=280)
        tree.column("Situação", width=120)
        tree.column("Padrão", width=100)
        tree.column("Desconsiderar saldo", width=160)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        for d in depositos:
            tree.insert("", tk.END, values=(d[0], d[1], d[2], d[3], d[4]))

        def selecionar(_=None):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0])["values"]
            dep_id = str(vals[0])
            if modo == "origem":
                self.deposito_origem_var.set(dep_id)
            else:
                self.deposito_destino_var.set(dep_id)
            win.destroy()

        tree.bind("<Double-Button-1>", selecionar)
        tree.bind("<Return>", selecionar)
        win.bind("<Escape>", lambda e: win.destroy())

    # ============================================================
    # Salvar OP (valida insumos + quebra por arranjo)
    # ============================================================

    def salvar_ordem(self):
        if not self.connected:
            messagebox.showerror(
                "Erro", "Não há conexão com o banco.", parent=self)
            return

        try:
            dados: Dict[str, Any] = {}
            dados["numero"] = self.numero_var.get().strip()
            if not dados["numero"]:
                raise ValueError("O campo 'Número da Ordem' é obrigatório.")

            dados["deposito_id_origem"] = self.parse_int(
                self.deposito_origem_var.get(), "Depósito Origem")
            dados["deposito_id_destino"] = self.parse_int(
                self.deposito_destino_var.get(), "Depósito Destino")
            dados["situacao_id"] = self.parse_int(
                self.situacao_id_var.get(), "Situação")
            dados["fkprodutoid"] = self.parse_int(
                self.produto_id_var.get(), "Produto")
            dados["quantidade"] = self.parse_float(
                self.quantidade_var.get(), "Quantidade", obrigatorio=True)

            dados["responsavel"] = (self.responsavel_var.get().strip() or None)
            dados["valor"] = self.parse_float(
                self.valor_var.get(), "Valor", obrigatorio=False)

            obs = self.observacao_text.get("1.0", tk.END).strip()
            dados["observacao"] = obs if obs else None

            dados["id"] = None
            dados["data_previsao_inicio"] = None
            dados["data_previsao_final"] = None
            dados["data_inicio"] = None
            dados["data_fim"] = None

            produto = self.sistema.validar_produto(int(dados["fkprodutoid"]))
            if not produto:
                raise ValueError("Produto não encontrado para gerar OP.")

            sku = (produto.get("sku") or "").strip()
            qtd_arranjo = float(
                self.sistema.buscar_qtd_produzir_por_sku(sku) or 0.0)
            qtd_total = float(dados["quantidade"] or 0.0)

            if qtd_arranjo <= 0:
                partes = [qtd_total]
            else:
                partes = []
                restante = qtd_total
                eps = 1e-9
                while restante > eps:
                    lote = min(qtd_arranjo, restante)
                    if lote < eps:
                        break
                    partes.append(lote)
                    restante -= lote

            try:
                numero_base = int(str(dados["numero"]).strip())
            except Exception:
                numero_base = int(self.sistema.gerar_numero_ordem())

            fkproduto_final = int(dados["fkprodutoid"])
            problemas_gerais = []
            for i, qtd_lote in enumerate(partes):
                res = self.sistema.validar_estoque_insumos_para_producao(
                    fkproduto=fkproduto_final,
                    qtd_produzir=float(qtd_lote),
                    bloquear_se_saldo_negativo=True,
                    bloquear_se_insuficiente=True,
                )
                if not res.get("ok"):
                    problemas_gerais.append(
                        (i, float(qtd_lote), res.get("problemas", [])))

            if problemas_gerais:
                linhas_msg = []
                for (idx, lote, probs) in problemas_gerais:
                    num_op = numero_base + idx
                    linhas_msg.append(f"\nOP {num_op} | Lote: {lote:.2f}")
                    for p in probs[:40]:
                        comp = p.get("componente", "")
                        nome = (p.get("nome") or "")
                        necessario = float(p.get("necessario") or 0.0)
                        saldo = float(p.get("saldo") or 0.0)
                        falta = float(p.get("falta") or 0.0)
                        motivo = p.get("motivo", "Problema")
                        linhas_msg.append(
                            f" - {motivo}: {comp} {nome} | Nec: {necessario:.4f} | Saldo: {saldo:.4f} | Falta: {falta:.4f}"
                        )
                    if len(probs) > 40:
                        linhas_msg.append(f" ... (+{len(probs)-40} itens)")

                messagebox.showerror(
                    "Bloqueado",
                    "Não é possível salvar a OP.\n"
                    "Existem insumos com saldo negativo e/ou saldo menor que o necessário:\n"
                    + "\n".join(linhas_msg),
                    parent=self
                )
                return

            if len(partes) == 1:
                msg_conf = f"Confirma inserir OP nº {numero_base}?"
            else:
                resumo = ", ".join([f"{p:.2f}" for p in partes])
                msg_conf = (
                    f"Confirma inserir {len(partes)} OPs a partir do nº {numero_base}?\n\n"
                    f"Quantidade total: {qtd_total:.2f}\n"
                    f"Máximo por OP (arranjo): {qtd_arranjo:.2f}\n"
                    f"Lotes: {resumo}"
                )

            if not messagebox.askyesno("Confirmar", msg_conf, parent=self):
                return

            op_criadas = []
            for i, qtd_lote in enumerate(partes):
                dados_lote = dict(dados)
                dados_lote["id"] = None
                dados_lote["numero"] = str(numero_base + i)
                dados_lote["quantidade"] = float(qtd_lote)

                ok, err = self.sistema.inserir_ordem_producao(dados_lote)
                if not ok:
                    messagebox.showerror(
                        "Erro ao inserir", f"Falha ao inserir OP nº {dados_lote['numero']}.\n\n{err}", parent=self)
                    return

                if N8N_WEBHOOK_URL:
                    try:
                        payload = {
                            "numero": dados_lote["numero"],
                            "deposito_id_origem": dados_lote["deposito_id_origem"],
                            "deposito_id_destino": dados_lote["deposito_id_destino"],
                            "situacao_id": dados_lote["situacao_id"],
                            "fkprodutoid": dados_lote["fkprodutoid"],
                            "quantidade": float(dados_lote["quantidade"] or 0),
                            "responsavel": dados_lote.get("responsavel"),
                            "observacao": dados_lote.get("observacao"),
                        }
                        requests.post(N8N_WEBHOOK_URL,
                                      json=payload, timeout=10)
                    except Exception:
                        pass

                op_criadas.append(
                    (dados_lote["numero"], float(dados_lote["quantidade"])))

            if len(op_criadas) == 1:
                messagebox.showinfo(
                    "Sucesso", f"OP {op_criadas[0][0]} inserida com sucesso!", parent=self)
            else:
                lista = "\n".join(
                    [f"OP {n} - {q:.2f}" for (n, q) in op_criadas])
                messagebox.showinfo(
                    "Sucesso", f"{len(op_criadas)} OPs inseridas com sucesso!\n\n{lista}", parent=self)

            self.limpar_formulario()

        except Exception as e:
            messagebox.showerror(
                "Erro", f"{type(e).__name__}: {e}", parent=self)

    def limpar_formulario(self):
        try:
            self.numero_var.set(str(self.sistema.gerar_numero_ordem()))
        except Exception:
            self.numero_var.set("")

        self.deposito_origem_var.set("")
        self.deposito_destino_var.set("")
        self.situacao_id_var.set("")
        self.produto_id_var.set("")
        self.responsavel_var.set("")
        self.quantidade_var.set("")
        self.valor_var.set("")
        self.observacao_text.delete("1.0", tk.END)
        self.variaveis_quantidade = None

    # ============================================================
    # F10 - Ordens existentes
    # ============================================================

    def mostrar_ordens_producao(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "F10 - Ordens", "Não há conexão com o banco.", parent=self)
            return

        ordens = self.sistema.listar_ordens_producao()
        if not ordens:
            messagebox.showinfo(
                "F10 - Ordens", "Nenhuma ordem encontrada.", parent=self)
            return

        win = tk.Toplevel(self)
        apply_window_icon(win)
        win.title("F10 - Ordens Existentes")
        win.geometry("1050x560")
        win.transient(self)
        win.grab_set()

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("id", "numero", "produto_id", "produto_nome",
                "situacao", "quantidade", "data_inicio", "data_fim")
        tree = ttk.Treeview(frame, columns=cols,
                            show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        for c in cols:
            tree.heading(c, text=c)

        tree.column("id", width=70, anchor="center")
        tree.column("numero", width=90, anchor="center")
        tree.column("produto_id", width=110, anchor="center")
        tree.column("produto_nome", width=280, anchor="w")
        tree.column("situacao", width=180, anchor="w")
        tree.column("quantidade", width=120, anchor="e")
        tree.column("data_inicio", width=120, anchor="center")
        tree.column("data_fim", width=120, anchor="center")

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        def fmt(dt):
            if not dt:
                return ""
            if isinstance(dt, (datetime, date)):
                return dt.strftime("%d/%m/%Y")
            return str(dt)

        for (oid, numero, produto_id, produto_nome, _sid, situacao_nome, quantidade, data_inicio, data_fim) in ordens:
            tree.insert("", tk.END, values=(
                oid,
                numero,
                produto_id,
                produto_nome or "",
                situacao_nome or "",
                f"{float(quantidade):.2f}" if quantidade is not None else "",
                fmt(data_inicio),
                fmt(data_fim),
            ))

        btns = ttk.Frame(win, padding=(10, 0, 10, 10))
        btns.pack(fill=tk.X)

        def excluir_selecionada(event=None):
            sel = tree.selection()
            if not sel:
                messagebox.showwarning(
                    "Excluir", "Selecione uma ordem.", parent=win)
                return

            item_id = sel[0]
            v = tree.item(item_id)["values"]
            oid, numero = v[0], v[1]

            if not messagebox.askyesno("Confirmar", f"Deseja excluir a OP nº {numero} (ID {oid})?", parent=win):
                return

            ok = self.sistema.excluir_ordem_producao(int(oid))
            if ok:
                tree.delete(item_id)
                messagebox.showinfo(
                    "Exclusão", f"OP nº {numero} excluída.", parent=win)
            else:
                messagebox.showerror(
                    "Erro", "Não foi possível excluir.", parent=win)

        ttk.Button(btns, text="Excluir (DEL)", command=excluir_selecionada).pack(
            side=tk.RIGHT, padx=(0, 8))
        ttk.Button(btns, text="Fechar",
                   command=win.destroy).pack(side=tk.RIGHT)

        win.bind("<Escape>", lambda e: win.destroy())
        tree.bind("<Delete>", excluir_selecionada)

    # ============================================================
    # F11 - Finalizar pendentes
    # ============================================================

    def finalizar_producoes_pendentes(self, event=None):
        if not self.connected:
            messagebox.showerror(
                "F11 - Finalizar", "Não há conexão com o banco.", parent=self)
            return

        pendentes = self.sistema.listar_ordens_sem_data_fim()
        if not pendentes:
            messagebox.showinfo(
                "F11 - Finalizar", "Não há ordens pendentes sem data fim.", parent=self)
            return

        win = tk.Toplevel(self)
        apply_window_icon(win)
        win.title("F11 - Finalizar Ordens Pendentes")
        win.geometry("1050x560")
        win.transient(self)
        win.grab_set()

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("id", "numero", "produto_id", "produto_nome",
                "situacao", "quantidade", "data_inicio")
        tree = ttk.Treeview(frame, columns=cols,
                            show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        for c in cols:
            tree.heading(c, text=c)

        tree.column("id", width=70, anchor="center")
        tree.column("numero", width=90, anchor="center")
        tree.column("produto_id", width=110, anchor="center")
        tree.column("produto_nome", width=320, anchor="w")
        tree.column("situacao", width=180, anchor="w")
        tree.column("quantidade", width=120, anchor="e")
        tree.column("data_inicio", width=120, anchor="center")

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        def fmt(dt):
            if not dt:
                return ""
            if isinstance(dt, (datetime, date)):
                return dt.strftime("%d/%m/%Y")
            return str(dt)

        for (oid, numero, produto_id, produto_nome, _sid, situacao_nome, quantidade, data_inicio) in pendentes:
            tree.insert("", tk.END, values=(
                oid, numero, produto_id, produto_nome or "", situacao_nome or "",
                f"{float(quantidade):.2f}" if quantidade is not None else "",
                fmt(data_inicio),
            ))

        btns = ttk.Frame(win, padding=(10, 0, 10, 10))
        btns.pack(fill=tk.X)

        def finalizar_selecionadas(event=None):
            sel = tree.selection()
            if not sel:
                messagebox.showwarning(
                    "Finalizar", "Selecione uma ou mais ordens.", parent=win)
                return

            ordens_sel = []
            for item_id in sel:
                v = tree.item(item_id)["values"]
                ordens_sel.append((item_id, int(v[0]), v[1]))

            if len(ordens_sel) == 1:
                _, oid, numero = ordens_sel[0]
                msg = f"Deseja finalizar a OP nº {numero} (ID {oid}) com data fim hoje?"
            else:
                nums = ", ".join(str(x[2]) for x in ordens_sel)
                msg = f"Deseja finalizar {len(ordens_sel)} OPs (números: {nums}) com data fim hoje?"

            if not messagebox.askyesno("Confirmar", msg, parent=win):
                return

            ok_count = 0
            for item_id, oid, _ in ordens_sel:
                ok = self.sistema.finalizar_ordem_individual(oid)
                if ok:
                    tree.delete(item_id)
                    ok_count += 1

            if ok_count:
                messagebox.showinfo(
                    "Finalização", f"{ok_count} ordem(ns) finalizada(s) com sucesso!", parent=win)
            else:
                messagebox.showerror(
                    "Erro", "Não foi possível finalizar as ordens selecionadas.", parent=win)

        ttk.Button(btns, text="Finalizar selecionadas (ENTER)",
                   command=finalizar_selecionadas).pack(side=tk.RIGHT, padx=(0, 8))
        ttk.Button(btns, text="Fechar",
                   command=win.destroy).pack(side=tk.RIGHT)

        win.bind("<Escape>", lambda e: win.destroy())
        tree.bind("<Return>", finalizar_selecionadas)
        tree.bind("<Double-Button-1>", finalizar_selecionadas)

    # ============================================================
    # VOLTAR / FECHAR
    # ============================================================

    def voltar_menu(self):
        """
        Volta para o menu principal (sem entrada) e fecha este sistema.
        """
        if self._closing:
            return
        self._closing = True
        try:
            self.withdraw()
            self.update_idletasks()
        except Exception:
            pass
        self.after(50, self._open_menu_then_close)

    def _open_menu_then_close(self):
        try:
            abrir_menu_principal_skip_entrada(parent=None)
        finally:
            try:
                self.sistema.desconectar()
            except Exception:
                pass
            try:
                self.destroy()
            except Exception:
                pass

    def on_close(self):
        """
        Ao fechar (X, ESC, botão Fechar): pergunta e volta ao menu.
        """
        if self._closing:
            return
        try:
            if messagebox.askokcancel("Sair", "Deseja realmente sair e voltar ao menu?", parent=self):
                self.voltar_menu()
        except Exception:
            self.voltar_menu()

    # alias para compatibilidade com trechos antigos
    def on_closing(self):
        self.on_close()


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    app = OrdemProducaoApp()
    app.mainloop()
